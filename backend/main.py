from fastapi import FastAPI, Depends, HTTPException, status, UploadFile, File
from fastapi.security import OAuth2PasswordBearer, OAuth2PasswordRequestForm
from fastapi.middleware.cors import CORSMiddleware
from sqlalchemy.orm import Session
from sqlalchemy import text, String, func
from sqlalchemy.exc import IntegrityError
from pydantic import BaseModel
import uuid
import os
import logging
import json
import time
import urllib.request
import smtplib
import hashlib
import requests
import models, schemas, auth, database, ai_engine, wellness_utils, gemini_utils
from typing import List, Optional, Dict, Any
import csv
import io
from email.message import EmailMessage
from fastapi.responses import StreamingResponse
import math
from datetime import datetime, timedelta
import secrets
from receipt_generator import generate_receipt_pdf
from pesapal_utils import PesapalAPI

logger = logging.getLogger(__name__)

pesapal = PesapalAPI()
gemini_advisor = gemini_utils.GeminiAdvisor()

def bootstrap_admin_user():
    admin_email = os.getenv("ADMIN_EMAIL", "").strip().lower()
    admin_password = os.getenv("ADMIN_PASSWORD", "")
    admin_name = os.getenv("ADMIN_NAME", "Lovedogs360 Admin").strip() or "Lovedogs360 Admin"

    if not admin_email or not admin_password:
        logger.info("ADMIN_EMAIL or ADMIN_PASSWORD not set; skipping admin bootstrap.")
        return

    if len(admin_password) < 8:
        logger.error("ADMIN_PASSWORD must be at least 8 characters; skipping admin bootstrap.")
        return

    db = database.SessionLocal()
    try:
        user = db.query(models.User).filter(models.User.email == admin_email).first()
        if user:
            user.role = models.UserRole.ADMIN.value
            user.full_name = user.full_name or admin_name
            user.hashed_password = auth.get_password_hash(admin_password)
            logger.info("Updated existing user %s as admin.", admin_email)
        else:
            user = models.User(
                id=str(uuid.uuid4()),
                email=admin_email,
                full_name=admin_name,
                hashed_password=auth.get_password_hash(admin_password),
                role=models.UserRole.ADMIN.value,
            )
            db.add(user)
            logger.info("Created bootstrap admin user %s.", admin_email)
        db.commit()
    except Exception as e:
        db.rollback()
        logger.error("Admin bootstrap failed: %s", e)
    finally:
        db.close()

def get_smtp_settings():
    smtp_host = (os.getenv("SMTP_HOST") or "").strip()
    smtp_port = int((os.getenv("SMTP_PORT") or "587").strip())
    smtp_user = (os.getenv("SMTP_USER") or "").strip()
    smtp_password = (os.getenv("SMTP_PASSWORD") or "").strip()
    from_email = (os.getenv("SMTP_FROM_EMAIL") or smtp_user).strip()
    frontend_url = (os.getenv("FRONTEND_URL") or "https://hunter-v9qj-lovebirds-project.vercel.app").strip()
    use_ssl = (os.getenv("SMTP_USE_SSL") or "").strip().lower() in {"1", "true", "yes"} or smtp_port == 465
    use_tls = (os.getenv("SMTP_USE_TLS") or "true").strip().lower() not in {"0", "false", "no"}
    timeout = int((os.getenv("SMTP_TIMEOUT") or "10").strip())

    # Google shows app passwords grouped with spaces; SMTP auth expects the 16 characters.
    if "gmail.com" in smtp_host.lower():
        smtp_password = smtp_password.replace(" ", "")

    return {
        "host": smtp_host,
        "port": smtp_port,
        "user": smtp_user,
        "password": smtp_password,
        "from_email": from_email,
        "frontend_url": frontend_url,
        "use_ssl": use_ssl,
        "use_tls": use_tls,
        "timeout": timeout,
    }

def is_smtp_configured():
    settings = get_smtp_settings()
    return bool(
        os.getenv("RESEND_API_KEY")
        or (settings["host"] and settings["user"] and settings["password"] and settings["from_email"])
    )

def get_smtp_attempts(settings: dict):
    attempts = [settings]

    if "gmail.com" in settings["host"].lower():
        fallbacks = [
            {**settings, "port": 465, "use_ssl": True, "use_tls": False},
            {**settings, "port": 587, "use_ssl": False, "use_tls": True},
        ]
        for fallback in fallbacks:
            duplicate = any(
                attempt["host"] == fallback["host"]
                and attempt["port"] == fallback["port"]
                and attempt["use_ssl"] == fallback["use_ssl"]
                for attempt in attempts
            )
            if not duplicate:
                attempts.append(fallback)

    return attempts

def send_email_via_resend(to_email: str, subject: str, body: str, from_email: str):
    api_key = (os.getenv("RESEND_API_KEY") or "").strip()
    if not api_key:
        raise RuntimeError("RESEND_API_KEY is not configured")

    resend_from_email = (os.getenv("RESEND_FROM_EMAIL") or from_email).strip()
    response = requests.post(
        "https://api.resend.com/emails",
        headers={
            "Authorization": f"Bearer {api_key}",
            "Content-Type": "application/json",
        },
        json={
            "from": resend_from_email,
            "to": [to_email],
            "subject": subject,
            "text": body,
        },
        timeout=10,
    )
    response.raise_for_status()
    logger.info("Password reset email sent via Resend to %s", to_email)
    return True

def send_email_via_smtp(to_email: str, subject: str, body: str, settings: dict):
    if not (settings["host"] and settings["user"] and settings["password"] and settings["from_email"]):
        raise RuntimeError("SMTP is not fully configured")

    last_error = None
    for attempt in get_smtp_attempts(settings):
        message = EmailMessage()
        message["Subject"] = subject
        message["From"] = attempt["from_email"]
        message["To"] = to_email
        message.set_content(body)

        try:
            logger.info(
                "Sending password reset email via SMTP host=%s port=%s ssl=%s tls=%s from=%s",
                attempt["host"],
                attempt["port"],
                attempt["use_ssl"],
                attempt["use_tls"],
                attempt["from_email"],
            )
            smtp_class = smtplib.SMTP_SSL if attempt["use_ssl"] else smtplib.SMTP
            with smtp_class(attempt["host"], attempt["port"], timeout=attempt["timeout"]) as server:
                if attempt["use_tls"] and not attempt["use_ssl"]:
                    server.starttls()
                server.login(attempt["user"], attempt["password"])
                server.send_message(message)
            return True
        except Exception as exc:
            last_error = exc
            logger.warning(
                "SMTP password reset email attempt failed host=%s port=%s ssl=%s error_type=%s error=%s",
                attempt["host"],
                attempt["port"],
                attempt["use_ssl"],
                type(exc).__name__,
                exc,
            )

    raise last_error

def hash_reset_token(token: str):
    return hashlib.sha256(token.encode("utf-8")).hexdigest()

def send_password_reset_email(email: str, token: str):
    smtp_settings = get_smtp_settings()
    subject = "Reset your Lovedogs360 password"
    body = (
        "Use this reset code to update your Lovedogs360 password:\n\n"
        f"{token}\n\n"
        f"Open {smtp_settings['frontend_url']}/forgot-password and paste the code. "
        "This code expires in 30 minutes."
    )

    if os.getenv("RESEND_API_KEY"):
        try:
            return send_email_via_resend(email, subject, body, smtp_settings["from_email"])
        except Exception as exc:
            logger.warning("Resend password reset email failed; trying SMTP fallback. error_type=%s error=%s", type(exc).__name__, exc)

    return send_email_via_smtp(email, subject, body, smtp_settings)

def calculate_distance(lat1, lon1, lat2, lon2):
    # Haversine formula
    R = 6371  # Earth radius in kilometers
    dlat = math.radians(lat2 - lat1)
    dlon = math.radians(lon2 - lon1)
    a = (math.sin(dlat / 2) * math.sin(dlat / 2) +
         math.cos(math.radians(lat1)) * math.cos(math.radians(lat2)) *
         math.sin(dlon / 2) * math.sin(dlon / 2))
    c = 2 * math.asin(math.sqrt(a))
    return R * c

KARMA_REDEMPTION_TARGET = 100
KARMA_POINT_VALUE = 1.0
KARMA_MAX_ORDER_DISCOUNT_RATE = 0.20
KARMA_REWARD_AMOUNT_STEP = 100
KARMA_MIN_REWARD_PER_ORDER = 5
KARMA_MAX_REWARD_PER_ORDER = 500
KARMA_CASE_REPORT_REWARD = 10
KARMA_CASE_COMMENT_REWARD = 2
MARKETPLACE_MARKUP_RATE = 0.235
MARKETPLACE_PRICE_MULTIPLIER = 1 + MARKETPLACE_MARKUP_RATE
MIN_MARKETPLACE_LISTING_PRICE_KES = 500.0
FALLBACK_EXCHANGE_RATES = {
    "USD": 1.0,
    "KES": 129.0,
    "EUR": 0.92,
    "GBP": 0.78,
}

ADMIN_ROLE_VALUES = {models.UserRole.ADMIN.value, models.UserRole.SUPER_ADMIN.value}
PINNABLE_TARGET_TYPES = {"event", "service", "case", "community"}
PIN_ROUTE_BY_TARGET = {
    "event": "EventDetail",
    "service": "Marketplace",
    "case": "CaseDetail",
    "community": "Community",
}

SCORECARD_CATEGORIES = [
    "Human Wellbeing",
    "Animal Welfare",
    "Environment",
    "Social Cohesion",
    "Indigenous/Local Knowledge",
]

SCORECARD_BASELINE_LIKERT = [
    ("Human Wellbeing", "I feel comfortable interacting with dogs in my community."),
    ("Human Wellbeing", "I understand how to safely approach or interact with a dog."),
    ("Human Wellbeing", "I know what to do if I encounter an unfamiliar dog."),
    ("Human Wellbeing", "I understand how rabies affects both people and animals."),
    ("Human Wellbeing", "I believe dogs contribute positively to community wellbeing."),
    ("Animal Welfare", "Dogs deserve humane treatment and care."),
    ("Animal Welfare", "I understand the basic welfare needs of a dog."),
    ("Animal Welfare", "I believe regular vaccination is important."),
    ("Animal Welfare", "I believe responsible ownership benefits both dogs and people."),
    ("Animal Welfare", "I know where to seek help for a dog welfare concern."),
    ("Environment", "People and dogs can safely share public spaces."),
    ("Environment", "Responsible dog ownership contributes to cleaner communities."),
    ("Environment", "Dog welfare is connected to environmental wellbeing."),
    ("Environment", "Community spaces should consider both people and animals."),
    ("Social Cohesion", "Conversations about dogs can bring communities together."),
    ("Social Cohesion", "I am willing to learn from others about living with dogs."),
    ("Social Cohesion", "I feel my experiences with dogs are valued."),
    ("Social Cohesion", "Different generations can learn from one another about dog care."),
    ("Indigenous/Local Knowledge", "Local and traditional knowledge can help improve relationships between people and dogs."),
    ("Indigenous/Local Knowledge", "Stories and lived experiences are valuable sources of learning."),
]

SCORECARD_BASELINE_OPEN = [
    "What is one challenge involving dogs in your community?",
    "What is one thing you would like to learn about dogs?",
    "Tell us about a positive or difficult experience you have had with a dog.",
]

SCORECARD_FOLLOWUP_OPEN = [
    "What new knowledge have you gained?",
    "Has your attitude toward dogs changed? If yes, how?",
    "Have you changed any behavior relating to dogs?",
    "What action have you taken since participating?",
    "What story or lesson stayed with you most?",
    "What additional support would help your community?",
]

DEFAULT_REPORTING_FIELDS = {
    "community_members_engaged": 0,
    "trainings_story_labs_conducted": 0,
    "animals_indirectly_benefiting": 0,
    "materials_tools_produced": "",
    "human_wellbeing_outcome_notes": "",
    "animal_welfare_outcome_notes": "",
    "environmental_benefit_notes": "",
    "social_cohesion_notes": "",
    "evidence_links_or_uploaded_files": "",
}


def is_admin_user(user: models.User) -> bool:
    return user.role in ADMIN_ROLE_VALUES


def parse_datetime_value(value):
    if isinstance(value, datetime):
        return value
    if not value:
        return value
    return datetime.fromisoformat(str(value).replace("Z", "+00:00"))


def get_scorecard_seed_questions():
    questions = []
    order = 0
    for category, prompt in SCORECARD_BASELINE_LIKERT:
        questions.append({
            "id": f"baseline_likert_{order + 1:02d}",
            "survey_type": "baseline",
            "category": category,
            "question_type": "likert",
            "prompt": prompt,
            "sort_order": order,
        })
        order += 1
    for prompt in SCORECARD_BASELINE_OPEN:
        questions.append({
            "id": f"baseline_open_{order + 1:02d}",
            "survey_type": "baseline",
            "category": None,
            "question_type": "open",
            "prompt": prompt,
            "sort_order": order,
        })
        order += 1

    order = 0
    # Follow-up repeats the Likert scorecard so the app can calculate change.
    for category, prompt in SCORECARD_BASELINE_LIKERT:
        questions.append({
            "id": f"followup_likert_{order + 1:02d}",
            "survey_type": "followup",
            "category": category,
            "question_type": "likert",
            "prompt": prompt,
            "sort_order": order,
        })
        order += 1
    for prompt in SCORECARD_FOLLOWUP_OPEN:
        questions.append({
            "id": f"followup_open_{order + 1:02d}",
            "survey_type": "followup",
            "category": None,
            "question_type": "open",
            "prompt": prompt,
            "sort_order": order,
        })
        order += 1
    return questions


def seed_scorecard_questions():
    db = database.SessionLocal()
    try:
        for item in get_scorecard_seed_questions():
            existing = db.query(models.ScorecardQuestion).filter(models.ScorecardQuestion.id == item["id"]).first()
            if existing:
                existing.survey_type = item["survey_type"]
                existing.category = item["category"]
                existing.question_type = item["question_type"]
                existing.prompt = item["prompt"]
                existing.sort_order = item["sort_order"]
                existing.is_active = True
            else:
                db.add(models.ScorecardQuestion(**item))
        db.commit()
    except Exception as exc:
        db.rollback()
        logger.warning("Scorecard question seed failed: %s", exc)
    finally:
        db.close()


def active_pin_filter(now=None):
    from sqlalchemy import or_
    now = now or datetime.utcnow()
    return (
        models.ContentPin.is_active == True,
        or_(models.ContentPin.expires_at.is_(None), models.ContentPin.expires_at > now),
    )


def get_active_pin_map(db: Session, target_type: str) -> Dict[str, models.ContentPin]:
    pins = db.query(models.ContentPin).filter(
        models.ContentPin.target_type == target_type,
        *active_pin_filter(),
    ).all()
    return {pin.target_id: pin for pin in pins}


def apply_pin_metadata(items, pin_map: Dict[str, models.ContentPin]):
    for item in items:
        item_id = str(getattr(item, "id", ""))
        pin = pin_map.get(item_id)
        setattr(item, "is_pinned", pin is not None)
        setattr(item, "pin_priority", pin.priority if pin else None)
    return items


def sort_items_with_pins(items, pin_map: Dict[str, models.ContentPin], secondary_key=None, reverse_secondary=False):
    def key(item):
        item_id = str(getattr(item, "id", ""))
        pin = pin_map.get(item_id)
        secondary = secondary_key(item) if secondary_key else getattr(item, "created_at", datetime.min)
        if isinstance(secondary, datetime):
            secondary_value = secondary.timestamp()
        else:
            secondary_value = secondary or 0
        if reverse_secondary:
            secondary_value = -secondary_value
        return (0 if pin else 1, -(pin.priority if pin else 0), secondary_value)
    return sorted(items, key=key)


def get_target_pin_payload(db: Session, target_type: str, target_id: str) -> Dict[str, Any]:
    if target_type == "event":
        item = db.query(models.Event).filter(models.Event.id == target_id).first()
        if not item:
            raise HTTPException(status_code=404, detail="Event not found")
        return {"title": item.title, "description": item.description, "image_url": item.poster_url}
    if target_type == "service":
        item = db.query(models.Service).filter(models.Service.id == target_id).first()
        if not item:
            raise HTTPException(status_code=404, detail="Service not found")
        return {"title": item.title, "description": item.description, "image_url": item.image_url}
    if target_type == "case":
        item = db.query(models.CaseReport).filter(models.CaseReport.id == target_id).first()
        if not item:
            raise HTTPException(status_code=404, detail="Case report not found")
        return {"title": item.title, "description": item.description, "image_url": item.image_url}
    if target_type == "community":
        item = db.query(models.CommunityMessage).filter(models.CommunityMessage.id == target_id).first()
        if not item:
            raise HTTPException(status_code=404, detail="Community post not found")
        title = (item.content or "Community post").strip()
        return {"title": title[:80], "description": item.content, "image_url": None}
    raise HTTPException(status_code=400, detail="Unsupported pin target type")


def ensure_content_pin(
    db: Session,
    target_type: str,
    target_id: str,
    admin: models.User,
    title: Optional[str] = None,
    description: Optional[str] = None,
    image_url: Optional[str] = None,
    priority: int = 100,
    expires_at: Optional[datetime] = None,
    commit: bool = True,
):
    if target_type not in PINNABLE_TARGET_TYPES:
        raise HTTPException(status_code=400, detail="Unsupported pin target type")

    target_payload = get_target_pin_payload(db, target_type, target_id)
    pin = db.query(models.ContentPin).filter(
        models.ContentPin.target_type == target_type,
        models.ContentPin.target_id == target_id,
        models.ContentPin.is_active == True,
    ).first()
    if not pin:
        pin = models.ContentPin(
            id=str(uuid.uuid4()),
            target_type=target_type,
            target_id=target_id,
            created_by_id=admin.id if admin else None,
        )
        db.add(pin)

    pin.title = title or target_payload["title"]
    pin.description = description if description is not None else target_payload.get("description")
    pin.image_url = image_url if image_url is not None else target_payload.get("image_url")
    pin.priority = int(priority or 100)
    pin.expires_at = expires_at
    pin.is_active = True
    pin.updated_at = datetime.utcnow()
    if commit:
        db.commit()
        db.refresh(pin)
    else:
        db.flush()
    return pin


def deactivate_content_pin(db: Session, target_type: str, target_id: str):
    pin = db.query(models.ContentPin).filter(
        models.ContentPin.target_type == target_type,
        models.ContentPin.target_id == target_id,
        models.ContentPin.is_active == True,
    ).first()
    if not pin:
        return None
    pin.is_active = False
    pin.updated_at = datetime.utcnow()
    db.commit()
    return pin


def content_pin_to_spotlight(pin: models.ContentPin):
    return {
        "id": pin.id,
        "title": pin.title,
        "description": pin.description,
        "image_url": pin.image_url,
        "target_route": PIN_ROUTE_BY_TARGET.get(pin.target_type),
        "target_id": pin.target_id,
        "is_active": pin.is_active,
        "updated_at": pin.updated_at,
        "is_pinned": True,
        "pin_priority": pin.priority,
        "target_type": pin.target_type,
    }


def calculate_scorecard_scores(question_map: Dict[str, models.ScorecardQuestion], responses: List[schemas.ScorecardResponseInput]):
    category_values: Dict[str, List[int]] = {category: [] for category in SCORECARD_CATEGORIES}
    all_values: List[int] = []

    for response in responses:
        question = question_map.get(response.question_id)
        if not question or question.question_type != "likert":
            continue
        value = response.answer_numeric
        if value is None:
            continue
        if value < 1 or value > 5:
            raise HTTPException(status_code=400, detail="Likert responses must be between 1 and 5")
        if question.category:
            category_values.setdefault(question.category, []).append(value)
        all_values.append(value)

    category_scores = {}
    for category, values in category_values.items():
        if values:
            category_scores[category] = round((sum(values) / (len(values) * 5)) * 100, 2)

    coexistence_index = round((sum(all_values) / (len(all_values) * 5)) * 100, 2) if all_values else 0.0
    return category_scores, coexistence_index


def find_or_create_scorecard_participant(db: Session, event_id: str, profile: schemas.ScorecardParticipantProfile):
    if not profile.consent:
        raise HTTPException(status_code=400, detail="Consent is required before submitting the scorecard")
    if not ((profile.full_name or "").strip() or (profile.anonymous_code or "").strip()):
        raise HTTPException(status_code=400, detail="Provide a full name or anonymous participant code")

    query = db.query(models.ScorecardParticipant).filter(models.ScorecardParticipant.event_id == event_id)
    participant = None
    if profile.anonymous_code:
        participant = query.filter(models.ScorecardParticipant.anonymous_code == profile.anonymous_code.strip()).first()
    if not participant and profile.phone_number:
        participant = query.filter(models.ScorecardParticipant.phone_number == profile.phone_number.strip()).first()
    if not participant and profile.full_name:
        participant = query.filter(
            models.ScorecardParticipant.full_name == profile.full_name.strip(),
            models.ScorecardParticipant.community_location == profile.community_location.strip(),
        ).first()

    if not participant:
        participant = models.ScorecardParticipant(id=str(uuid.uuid4()), event_id=event_id)
        db.add(participant)

    participant.full_name = (profile.full_name or "").strip() or None
    participant.anonymous_code = (profile.anonymous_code or "").strip() or None
    participant.phone_number = (profile.phone_number or "").strip() or None
    participant.county = profile.county.strip()
    participant.community_location = profile.community_location.strip()
    participant.user_type = profile.user_type
    participant.participation_type = profile.participation_type
    participant.consent = profile.consent
    participant.updated_at = datetime.utcnow()
    db.flush()
    return participant


def participant_score_pair(db: Session, event_id: str, participant_id: str):
    baseline = db.query(models.ScorecardSurvey).filter(
        models.ScorecardSurvey.event_id == event_id,
        models.ScorecardSurvey.participant_id == participant_id,
        models.ScorecardSurvey.survey_type == "baseline",
    ).order_by(models.ScorecardSurvey.created_at.desc()).first()
    followup = db.query(models.ScorecardSurvey).filter(
        models.ScorecardSurvey.event_id == event_id,
        models.ScorecardSurvey.participant_id == participant_id,
        models.ScorecardSurvey.survey_type == "followup",
    ).order_by(models.ScorecardSurvey.created_at.desc()).first()
    baseline_score = baseline.coexistence_index if baseline else None
    followup_score = followup.coexistence_index if followup else None
    change = None
    if baseline_score is not None and followup_score is not None:
        change = round(followup_score - baseline_score, 2)
    return baseline_score, followup_score, change


def get_or_create_reporting_export(db: Session, event_id: str, admin: Optional[models.User] = None):
    report = db.query(models.ScorecardReportingExport).filter(
        models.ScorecardReportingExport.event_id == event_id
    ).order_by(models.ScorecardReportingExport.updated_at.desc()).first()
    if not report:
        report = models.ScorecardReportingExport(
            id=str(uuid.uuid4()),
            event_id=event_id,
            fields=DEFAULT_REPORTING_FIELDS.copy(),
            created_by_id=admin.id if admin else None,
        )
        db.add(report)
        db.flush()
    return report


def scorecard_dashboard_payload(db: Session, event_id: str):
    event = db.query(models.Event).filter(models.Event.id == event_id).first()
    if not event:
        raise HTTPException(status_code=404, detail="Event not found")

    participants = db.query(models.ScorecardParticipant).filter(models.ScorecardParticipant.event_id == event_id).all()
    surveys = db.query(models.ScorecardSurvey).filter(models.ScorecardSurvey.event_id == event_id).all()
    baseline_count = sum(1 for survey in surveys if survey.survey_type == "baseline")
    followup_count = sum(1 for survey in surveys if survey.survey_type == "followup")
    avg_index = round(sum(s.coexistence_index or 0 for s in surveys) / len(surveys), 2) if surveys else 0.0

    changes = []
    for participant in participants:
        _, _, change = participant_score_pair(db, event_id, participant.id)
        if change is not None:
            changes.append(change)
    avg_change = round(sum(changes) / len(changes), 2) if changes else 0.0

    def count_by(attr):
        result = {}
        for participant in participants:
            key = getattr(participant, attr) or "Unknown"
            result[key] = result.get(key, 0) + 1
        return result

    participation_counts = count_by("participation_type")
    latest_category_scores: Dict[str, List[float]] = {}
    for survey in surveys:
        for category, score in (survey.category_scores or {}).items():
            latest_category_scores.setdefault(category, []).append(float(score or 0))

    category_averages = {
        category: round(sum(values) / len(values), 2)
        for category, values in latest_category_scores.items()
        if values
    }

    evidence = db.query(models.ScorecardEvidence).filter(
        models.ScorecardEvidence.event_id == event_id
    ).order_by(models.ScorecardEvidence.created_at.desc()).all()
    report = db.query(models.ScorecardReportingExport).filter(
        models.ScorecardReportingExport.event_id == event_id
    ).order_by(models.ScorecardReportingExport.updated_at.desc()).first()

    return {
        "event": {
            "id": event.id,
            "title": event.title,
            "start_time": str(event.start_time),
            "follow_up_requested_at": str(event.follow_up_requested_at) if event.follow_up_requested_at else None,
        },
        "total_participants": len(participants),
        "baseline_surveys_completed": baseline_count,
        "followup_surveys_completed": followup_count,
        "average_coexistence_index": avg_index,
        "average_change_from_baseline_to_followup": avg_change,
        "participants_by_county": count_by("county"),
        "participants_by_user_type": count_by("user_type"),
        "participation_type_counts": participation_counts,
        "story_labs_attended": participation_counts.get("story lab", 0),
        "listening_circles_attended": participation_counts.get("listening circle", 0),
        "podcast_listeners": participation_counts.get("podcast listener", 0),
        "category_averages": category_averages,
        "evidence": [
            {
                "id": item.id,
                "evidence_type": item.evidence_type,
                "url": item.url,
                "notes": item.notes,
                "created_at": str(item.created_at),
            }
            for item in evidence
        ],
        "reporting_fields": {**DEFAULT_REPORTING_FIELDS, **((report.fields or {}) if report else {})},
    }

def get_rate_for_currency(currency: Optional[str]) -> Optional[float]:
    normalized = (currency or "KES").strip().upper()
    live_rates = exchange_rates_cache.get("rates") if "exchange_rates_cache" in globals() else {}
    rates = {**FALLBACK_EXCHANGE_RATES, **(live_rates or {})}
    try:
        return float(rates.get(normalized))
    except (TypeError, ValueError):
        return None

def convert_amount(amount: float, from_currency: Optional[str], to_currency: Optional[str]) -> Optional[float]:
    amount = float(amount or 0)
    from_rate = get_rate_for_currency(from_currency)
    to_rate = get_rate_for_currency(to_currency)
    if not from_rate or not to_rate:
        return None
    return (amount / from_rate) * to_rate

def minimum_listing_price_for_currency(currency: Optional[str]) -> float:
    converted = convert_amount(MIN_MARKETPLACE_LISTING_PRICE_KES, "KES", currency or "KES")
    return round(converted if converted is not None else MIN_MARKETPLACE_LISTING_PRICE_KES, 2)

def validate_marketplace_base_price(base_price: float, currency: Optional[str]) -> float:
    currency_code = (currency or "KES").strip().upper()
    if base_price is None or float(base_price or 0) <= 0:
        raise HTTPException(status_code=400, detail="Enter a valid listing price")

    final_price = round(float(base_price) * MARKETPLACE_PRICE_MULTIPLIER, 2)
    final_price_kes = convert_amount(final_price, currency_code, "KES")
    if final_price_kes is None:
        raise HTTPException(status_code=400, detail=f"Currency {currency_code} is not supported for marketplace pricing")

    if final_price_kes + 0.01 < MIN_MARKETPLACE_LISTING_PRICE_KES:
        minimum_final = minimum_listing_price_for_currency(currency_code)
        minimum_base = round(minimum_final / MARKETPLACE_PRICE_MULTIPLIER, 2)
        raise HTTPException(
            status_code=400,
            detail=(
                f"Minimum allowed final listing price is KES {MIN_MARKETPLACE_LISTING_PRICE_KES:,.0f} "
                f"or about {currency_code} {minimum_final:,.2f}. "
                f"Enter at least {currency_code} {minimum_base:,.2f} before marketplace mark-up."
            )
        )

    return final_price

def calculate_karma_reward(amount: float) -> int:
    """1 point per KES 100, with a small minimum and abuse-safe cap."""
    amount = float(amount or 0)
    if amount <= 0:
        return 0
    return min(
        KARMA_MAX_REWARD_PER_ORDER,
        max(KARMA_MIN_REWARD_PER_ORDER, int(amount // KARMA_REWARD_AMOUNT_STEP))
    )

def calculate_karma_redemption(user: models.User, order_total: float, requested_points: int) -> tuple[int, float]:
    requested_points = int(requested_points or 0)
    if requested_points <= 0:
        return 0, 0.0

    available_karma = int(user.available_karma or 0)
    if available_karma < KARMA_REDEMPTION_TARGET:
        raise HTTPException(
            status_code=400,
            detail=f"You need at least {KARMA_REDEMPTION_TARGET} points before redeeming a discount."
        )
    if requested_points < KARMA_REDEMPTION_TARGET:
        raise HTTPException(
            status_code=400,
            detail=f"Redeem at least {KARMA_REDEMPTION_TARGET} points."
        )

    max_discount_amount = max(float(order_total or 0) * KARMA_MAX_ORDER_DISCOUNT_RATE, 0)
    max_points_for_order = int(max_discount_amount // KARMA_POINT_VALUE)
    points_to_redeem = min(requested_points, available_karma, max_points_for_order)
    if points_to_redeem < KARMA_REDEMPTION_TARGET:
        raise HTTPException(
            status_code=400,
            detail=f"This order can only use discounts from {KARMA_REDEMPTION_TARGET} points or more."
        )

    discount_amount = round(points_to_redeem * KARMA_POINT_VALUE, 2)
    return points_to_redeem, discount_amount

def add_notification(db: Session, user_id: str, title: str, message: str, type: str = "info", commit: bool = True):
    new_notif = models.Notification(
        id=str(uuid.uuid4()),
        user_id=user_id,
        title=title,
        message=message,
        type=type
    )
    db.add(new_notif)
    if commit:
        db.commit()
    else:
        db.flush()
    return new_notif

def award_karma(db: Session, user_id: str, amount: int, category: str, description: str = None, commit: bool = True):
    """Helper to award karma and track the transaction"""
    amount = int(amount or 0)
    if amount <= 0:
        return None

    user = db.query(models.User).filter(models.User.id == user_id).first()
    if user:
        user.karma_points = int(user.karma_points or 0) + amount
        user.available_karma = int(user.available_karma or 0) + amount
        transaction = models.KarmaTransaction(
            id=str(uuid.uuid4()),
            user_id=user_id,
            amount=amount,
            category=category,
            description=description
        )
        db.add(transaction)
        if commit:
            db.commit()
        else:
            db.flush()
        return transaction
    return None


app = FastAPI(title="Lovedogs 360 API")

# Initialize database tables on startup
try:
    models.Base.metadata.create_all(bind=database.engine)
    logger.info("Successfully initialized database tables")
    bootstrap_admin_user()
    
    # Run OAuth migrations
    db = database.SessionLocal()
    try:
        migration_statements = [
            """ALTER TABLE users
               ADD COLUMN IF NOT EXISTS auth_provider VARCHAR(50) DEFAULT 'email';""",
            """ALTER TABLE users
               ADD COLUMN IF NOT EXISTS google_id VARCHAR(255) UNIQUE;""",
            """ALTER TABLE users
               ADD COLUMN IF NOT EXISTS payment_method VARCHAR;""",
            """ALTER TABLE users
               ALTER COLUMN hashed_password DROP NOT NULL;""",
            """ALTER TABLE users
               ALTER COLUMN full_name DROP NOT NULL;""",
            """ALTER TABLE services
               ADD COLUMN IF NOT EXISTS item_type VARCHAR DEFAULT 'services';""",
            """ALTER TABLE services
               ADD COLUMN IF NOT EXISTS image_url VARCHAR;""",
            """ALTER TABLE services
               ADD COLUMN IF NOT EXISTS latitude DOUBLE PRECISION;""",
            """ALTER TABLE services
               ADD COLUMN IF NOT EXISTS longitude DOUBLE PRECISION;""",
            """ALTER TABLE services
               ADD COLUMN IF NOT EXISTS address VARCHAR;""",
            """ALTER TABLE services
               ADD COLUMN IF NOT EXISTS location_landmark VARCHAR;""",
            """ALTER TABLE services
               ADD COLUMN IF NOT EXISTS is_published BOOLEAN DEFAULT TRUE;""",
            """ALTER TABLE services
               ADD COLUMN IF NOT EXISTS currency VARCHAR DEFAULT 'KES';""",
            """ALTER TABLE services
               ADD COLUMN IF NOT EXISTS stock_count INTEGER DEFAULT 0;""",
            """ALTER TABLE services
               ADD COLUMN IF NOT EXISTS slots_available INTEGER DEFAULT 0;""",
            """ALTER TABLE services
               ADD COLUMN IF NOT EXISTS is_busy BOOLEAN DEFAULT FALSE;""",
            """ALTER TABLE services
               ADD COLUMN IF NOT EXISTS images JSON;""",
            """ALTER TABLE services
               ADD COLUMN IF NOT EXISTS admin_approved BOOLEAN DEFAULT FALSE;""",
            """ALTER TABLE services
               ADD COLUMN IF NOT EXISTS rejection_reason VARCHAR;""",
            """ALTER TABLE direct_messages
               ADD COLUMN IF NOT EXISTS read_at TIMESTAMP;""",
            """ALTER TABLE orders
               ADD COLUMN IF NOT EXISTS discount_amount DOUBLE PRECISION DEFAULT 0;""",
            """ALTER TABLE orders
               ADD COLUMN IF NOT EXISTS karma_points_redeemed INTEGER DEFAULT 0;""",
            """ALTER TABLE transactions
               ADD COLUMN IF NOT EXISTS payout_method VARCHAR;""",
            """ALTER TABLE transactions
               ADD COLUMN IF NOT EXISTS destination VARCHAR;""",
            """ALTER TABLE transactions
               ADD COLUMN IF NOT EXISTS created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP;""",
            """ALTER TABLE transactions
               ADD COLUMN IF NOT EXISTS processed_at TIMESTAMP;""",
            """ALTER TABLE events
               ADD COLUMN IF NOT EXISTS admin_created BOOLEAN DEFAULT FALSE;""",
            """ALTER TABLE events
               ADD COLUMN IF NOT EXISTS scorecard_enabled BOOLEAN DEFAULT TRUE;""",
            """ALTER TABLE events
               ADD COLUMN IF NOT EXISTS follow_up_requested_at TIMESTAMP;""",
            """ALTER TABLE events
               ADD COLUMN IF NOT EXISTS poster_url VARCHAR;""",
            """ALTER TABLE events
               ADD COLUMN IF NOT EXISTS images JSON;""",
            """ALTER TABLE events
               ADD COLUMN IF NOT EXISTS ticket_price DOUBLE PRECISION DEFAULT 0;""",
            """ALTER TABLE events
               ADD COLUMN IF NOT EXISTS currency VARCHAR DEFAULT 'KES';""",
            """ALTER TABLE events
               ADD COLUMN IF NOT EXISTS ticket_tiers JSON;""",
            """ALTER TABLE events
               ADD COLUMN IF NOT EXISTS attendee_type_question VARCHAR;""",
            """ALTER TABLE registrations
               ADD COLUMN IF NOT EXISTS amount DOUBLE PRECISION DEFAULT 0;""",
            """ALTER TABLE registrations
               ADD COLUMN IF NOT EXISTS currency VARCHAR DEFAULT 'KES';""",
            """ALTER TABLE registrations
               ADD COLUMN IF NOT EXISTS payment_status VARCHAR DEFAULT 'free';""",
            """ALTER TABLE registrations
               ADD COLUMN IF NOT EXISTS ticket_tier_id VARCHAR;""",
            """ALTER TABLE registrations
               ADD COLUMN IF NOT EXISTS ticket_tier_label VARCHAR;""",
            """ALTER TABLE registrations
               ADD COLUMN IF NOT EXISTS attendee_type_justification VARCHAR;""",
            """ALTER TABLE registrations
               ADD COLUMN IF NOT EXISTS pesapal_tracking_id VARCHAR;""",
            """ALTER TABLE registrations
               ADD COLUMN IF NOT EXISTS pesapal_merchant_reference VARCHAR;""",
            """ALTER TABLE registrations
               ADD COLUMN IF NOT EXISTS paid_at TIMESTAMP;""",
            """CREATE INDEX IF NOT EXISTS idx_users_google_id ON users(google_id);""",
            """CREATE INDEX IF NOT EXISTS idx_users_auth_provider ON users(auth_provider);""",
            """CREATE INDEX IF NOT EXISTS idx_users_email ON users(email);""",
            """CREATE INDEX IF NOT EXISTS idx_services_marketplace ON services(item_type, is_published, admin_approved);""",
            """CREATE INDEX IF NOT EXISTS idx_direct_messages_users ON direct_messages(sender_id, receiver_id, created_at);""",
            """CREATE INDEX IF NOT EXISTS idx_content_pins_target ON content_pins(target_type, target_id, is_active);""",
            """CREATE INDEX IF NOT EXISTS idx_scorecard_participants_event ON scorecard_participants(event_id);""",
            """CREATE INDEX IF NOT EXISTS idx_scorecard_surveys_event_type ON scorecard_surveys(event_id, survey_type);""",
            """CREATE INDEX IF NOT EXISTS idx_registrations_payment ON registrations(payment_status, pesapal_tracking_id);""",
        ]
        
        for i, stmt in enumerate(migration_statements, 1):
            try:
                db.execute(text(stmt))
                logger.info(f"✅ OAuth migration step {i} completed")
            except Exception as e:
                error_msg = str(e).lower()
                if any(skip in error_msg for skip in ["already exists", "duplicate", "constraint"]):
                    logger.info(f"⚠️  OAuth migration step {i} skipped (already exists)")
                else:
                    logger.warning(f"⚠️  OAuth migration step {i}: {str(e)[:100]}")
        
        db.commit()
        logger.info("✅ Google OAuth database migrations completed")
    except Exception as e:
        logger.warning(f"⚠️  OAuth migrations warning: {str(e)[:150]}")
    finally:
        db.close()
    seed_scorecard_questions()
        
except Exception as e:
    logger.error(f"Error initializing database tables: {e}")

@app.get("/")
def read_root():
    return {"status": "healthy", "message": "Lovedogs 360 API is running"}

@app.get("/health")
def health_check():
    return {"status": "healthy"}

@app.get("/health/db")
def database_health_check(db: Session = Depends(database.get_db)):
    try:
        db.execute(text("SELECT 1"))
        return {"status": "healthy", "database": "connected"}
    except Exception as e:
        logger.error(f"Database health check failed: {e}")
        raise HTTPException(status_code=503, detail="Database connection failed")

app.add_middleware(
    CORSMiddleware,
    allow_origins=[
        "https://lovedogs360.com",
        "https://admin.lovedogs360.com",
        "https://hunter-k9lr.vercel.app",
        "http://localhost:3000",
        "http://localhost:8081",
        "http://localhost:19006",
        "http://localhost:8082",
    ],
    allow_origin_regex=r"https://.*\.vercel\.app",
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

oauth2_scheme = OAuth2PasswordBearer(tokenUrl="token")

# Initialize AI engine
ai_engine_instance = ai_engine.DogIDEngine()

@app.post("/register", response_model=schemas.UserResponse)
def register(user: schemas.UserCreate, db: Session = Depends(database.get_db)):
    db_user = db.query(models.User).filter(models.User.email == user.email).first()
    if db_user:
        raise HTTPException(status_code=400, detail="Email already registered")
    
    hashed_password = auth.get_password_hash(user.password)
    new_user = models.User(
        id=str(uuid.uuid4()),
        email=user.email,
        full_name=user.full_name,
        hashed_password=hashed_password,
        role=user.role,
        phone_number=user.phone_number,
        country=user.country,
        language=user.language,
        latitude=user.latitude,
        longitude=user.longitude,
        address=user.address,
        bio=user.bio
    )
    try:
        db.add(new_user)
        db.commit()
        db.refresh(new_user)
        logger.info(f"Successfully registered user {new_user.email}")
        return new_user
    except Exception as e:
        db.rollback()
        logger.error(f"Error registering user {user.email}: {e}")
        from fastapi import HTTPException
        raise HTTPException(status_code=400, detail=str(e))

@app.post("/password/forgot")
def forgot_password(request: schemas.PasswordResetRequest, db: Session = Depends(database.get_db)):
    if not is_smtp_configured():
        raise HTTPException(
            status_code=503,
            detail="Password reset email is not configured yet. Please contact support."
        )

    email = request.email.lower()
    user = db.query(models.User).filter(models.User.email == email).first()
    response = {"message": "If this email exists, password reset instructions will be sent shortly."}

    if not user:
        return response

    token = secrets.token_urlsafe(32)
    expires_at = datetime.utcnow() + timedelta(minutes=30)
    reset_record = models.PasswordResetToken(
        id=str(uuid.uuid4()),
        user_id=user.id,
        token_hash=hash_reset_token(token),
        expires_at=expires_at,
    )
    db.add(reset_record)
    db.commit()

    try:
        send_password_reset_email(user.email, token)
        logger.info(f"Password reset email sent to {user.email}")
    except Exception as e:
        db.delete(reset_record)
        db.commit()
        logger.error(f"Failed to send password reset email to {user.email}: {e}")
        raise HTTPException(status_code=503, detail="Could not send password reset email. Please try again later.")

    logger.info(f"Password reset requested for {user.email}. Token expires at {expires_at.isoformat()}Z")
    return response

@app.post("/password/reset")
def reset_password(request: schemas.PasswordResetConfirm, db: Session = Depends(database.get_db)):
    token_hash = hash_reset_token(request.token)
    token_data = db.query(models.PasswordResetToken).filter(
        models.PasswordResetToken.token_hash == token_hash,
        models.PasswordResetToken.used_at.is_(None),
    ).first()

    if not token_data or token_data.expires_at < datetime.utcnow():
        raise HTTPException(status_code=400, detail="Invalid or expired reset token")

    if len(request.new_password) < 8:
        raise HTTPException(status_code=400, detail="Password must be at least 8 characters")

    user = db.query(models.User).filter(models.User.id == token_data.user_id).first()
    if not user:
        raise HTTPException(status_code=400, detail="Invalid or expired reset token")

    user.hashed_password = auth.get_password_hash(request.new_password)
    token_data.used_at = datetime.utcnow()
    db.commit()
    return {"message": "Password reset successful. You can now log in."}

@app.post("/token", response_model=schemas.Token)
def login(form_data: OAuth2PasswordRequestForm = Depends(), db: Session = Depends(database.get_db)):
    user = db.query(models.User).filter(models.User.email == form_data.username).first()
    if not user or not auth.verify_password(form_data.password, user.hashed_password):
        raise HTTPException(
            status_code=status.HTTP_401_UNAUTHORIZED,
            detail="Incorrect username or password",
            headers={"WWW-Authenticate": "Bearer"},
        )
    access_token = auth.create_access_token(data={"sub": user.email, "role": user.role})
    return {"access_token": access_token, "token_type": "bearer"}

@app.post("/auth/google", response_model=schemas.GoogleLoginResponse)
async def google_auth(request: schemas.GoogleLoginRequest, db: Session = Depends(database.get_db)):
    try:
        id_info = auth.verify_google_token(request.id_token)
    except auth.GoogleVerificationUnavailable:
        raise HTTPException(
            status_code=503,
            detail="Google login is temporarily unavailable. Please try again shortly."
        )
    if not id_info:
        raise HTTPException(status_code=400, detail="Invalid Google token")
    
    email = (id_info.get("email") or "").strip().lower()
    full_name = (id_info.get("name") or "").strip()
    google_id = id_info.get("sub")
    email_verified = id_info.get("email_verified")
    
    if not email or not google_id:
        raise HTTPException(status_code=400, detail="Invalid Google token - missing required profile details")

    if email_verified is False:
        raise HTTPException(status_code=400, detail="Google account email is not verified")
    
    user = db.query(models.User).filter(models.User.google_id == google_id).first()
    if not user:
        user = db.query(models.User).filter(models.User.email == email).first()
        if user and user.google_id and user.google_id != google_id:
            raise HTTPException(status_code=400, detail="This email is already linked to a different Google account")

    if not user:
        user = models.User(
            id=str(uuid.uuid4()),
            email=email,
            full_name=full_name or email.split("@")[0],
            hashed_password=None,
            role=models.UserRole.BUYER.value,
            auth_provider="google",
            google_id=google_id,
            language="en",
        )
        db.add(user)
    else:
        user.full_name = user.full_name or full_name or email.split("@")[0]
        user.google_id = user.google_id or google_id
        user.auth_provider = "google"

    try:
        db.commit()
        db.refresh(user)
    except IntegrityError:
        db.rollback()
        raise HTTPException(status_code=409, detail="This Google account is already linked to another user")
    
    access_token = auth.create_access_token(data={"sub": user.email, "role": user.role})
    
    return {
        "access_token": access_token,
        "token_type": "bearer",
        "user": user
    }

async def get_current_user(token: str = Depends(oauth2_scheme), db: Session = Depends(database.get_db)):
    credentials_exception = HTTPException(
        status_code=status.HTTP_401_UNAUTHORIZED,
        detail="Could not validate credentials",
        headers={"WWW-Authenticate": "Bearer"},
    )
    try:
        payload = auth.jwt.decode(token, auth.SECRET_KEY, algorithms=[auth.ALGORITHM])
        email: str = payload.get("sub")
        if email is None:
            raise credentials_exception
    except auth.JWTError:
        raise credentials_exception
    user = db.query(models.User).filter(models.User.email == email).first()
    if user is None:
        raise credentials_exception
    if user is None:
        raise credentials_exception
    return user

def require_admin(current_user: models.User = Depends(get_current_user)):
    if current_user.role not in {models.UserRole.ADMIN.value, models.UserRole.SUPER_ADMIN.value}:
        raise HTTPException(status_code=403, detail="Admin access required")
    return current_user

@app.get("/users/me", response_model=schemas.UserResponse)
async def read_users_me(current_user: models.User = Depends(get_current_user)):
    avg_rating = 0.0
    total = len(current_user.ratings_received)
    if total > 0:
        avg_rating = sum(r.score for r in current_user.ratings_received) / total
    
    # Create a response object with computed fields
    response = schemas.UserResponse.from_orm(current_user)
    response.average_rating = avg_rating
    response.total_ratings = total
    return response

@app.put("/users/me", response_model=schemas.UserResponse)
async def update_user_me(
    user_update: schemas.UserUpdate,
    db: Session = Depends(database.get_db),
    current_user: models.User = Depends(get_current_user)
):
    user = db.query(models.User).filter(models.User.id == current_user.id).first()
    if user_update.full_name is not None:
        user.full_name = user_update.full_name
    if user_update.phone_number is not None:
        user.phone_number = user_update.phone_number
    if user_update.bio is not None:
        user.bio = user_update.bio
    if user_update.profile_image is not None:
        user.profile_image = user_update.profile_image
    if user_update.country is not None:
        user.country = user_update.country
    if user_update.language is not None:
        user.language = user_update.language
    if user_update.mpesa_phone_number is not None:
        user.mpesa_phone_number = user_update.mpesa_phone_number
    if user_update.preferred_currency is not None:
        user.preferred_currency = user_update.preferred_currency
    if user_update.payment_method is not None:
        payment_method = user_update.payment_method.strip().lower() if user_update.payment_method else None
        if payment_method and payment_method not in {"mpesa", "card"}:
            raise HTTPException(status_code=400, detail="Payment method must be mpesa or card")
        user.payment_method = payment_method
    
    db.commit()
    db.refresh(user)
    return user

@app.get("/dogs/{dog_id}", response_model=schemas.DogResponse)
async def get_dog(
    dog_id: str,
    db: Session = Depends(database.get_db),
    current_user: models.User = Depends(get_current_user)
):
    dog = db.query(models.Dog).filter(models.Dog.id == dog_id).first()
    if not dog:
        raise HTTPException(status_code=404, detail="Dog not found")
    if dog.owner_id != current_user.id and current_user.role not in [models.UserRole.ADMIN, models.UserRole.PROVIDER]:
        raise HTTPException(status_code=403, detail="Not authorized")
    return dog

@app.put("/dogs/{dog_id}", response_model=schemas.DogResponse)
async def update_dog(
    dog_id: str,
    dog_update: schemas.DogUpdate,
    db: Session = Depends(database.get_db),
    current_user: models.User = Depends(get_current_user)
):
    dog = db.query(models.Dog).filter(models.Dog.id == dog_id).first()
    if not dog:
        raise HTTPException(status_code=404, detail="Dog not found")
    if dog.owner_id != current_user.id:
        raise HTTPException(status_code=403, detail="Not authorized")

    if dog_update.name is not None: dog.name = dog_update.name
    if dog_update.breed is not None: dog.breed = dog_update.breed
    if dog_update.color is not None: dog.color = dog_update.color
    if dog_update.height is not None: dog.height = dog_update.height
    if dog_update.weight is not None: dog.weight = dog_update.weight
    if dog_update.body_structure is not None: dog.body_structure = dog_update.body_structure
    if dog_update.bio is not None: dog.bio = dog_update.bio
    if dog_update.body_image is not None: dog.body_image = dog_update.body_image
    
    db.commit()
    db.refresh(dog)
    return dog

@app.post("/dogs", response_model=schemas.DogResponse)
async def create_dog(
    dog: schemas.DogCreate, 
    db: Session = Depends(database.get_db),
    current_user: models.User = Depends(get_current_user)
):
    new_dog = models.Dog(
        id=str(uuid.uuid4()),
        owner_id=current_user.id,
        **dog.dict()
    )
    db.add(new_dog)
    db.commit()
    db.refresh(new_dog)
    return new_dog

@app.post("/dogs/report-lost")
async def report_lost_dog(
    lost_info: schemas.DogCreate,
    db: Session = Depends(database.get_db),
    current_user: models.User = Depends(get_current_user)
):
    # Enhanced matching logic: Breed match + Similarity in Color, Age, and Weight
    query = db.query(models.Dog).filter(models.Dog.breed == lost_info.breed)
    
    potential_matches = []
    all_dogs = query.all()
    
    for d in all_dogs:
        score = 0
        if d.color == lost_info.color: score += 50
        
        # Age proximity (within 2 years)
        if d.age and lost_info.age:
            if abs(d.age - lost_info.age) <= 2: score += 25
            
        # Weight proximity (within 5kg)
        if d.weight and lost_info.weight:
            if abs(d.weight - lost_info.weight) <= 5: score += 25
            
        if score >= 50:
            potential_matches.append({"id": d.id, "name": d.name, "similarity": f"{score}%"})
            logger.info(f"Potential match found for dog {d.name} ({score}%). Notifying owner.")

    return {
        "message": f"Lost report processed. Found {len(potential_matches)} potential matches.",
        "matches": potential_matches
    }

@app.post("/dogs/identify")
async def identify_dog(
    file: UploadFile = File(...),
    db: Session = Depends(database.get_db)
):
    contents = await file.read()
    descriptor = ai_engine_instance.extract_descriptor(contents)
    if descriptor is None:
        return {"dog_id": None, "confidence": 0, "message": "Failed to extract nose print"}

    known_dogs = db.query(models.Dog).all()
    match = ai_engine_instance.identify_dog(descriptor, known_dogs)
    
    if match and match["confidence"] > 60: # Threshold
        return {
            "dog_id": match["dog_id"],
            "name": match["name"],
            "confidence": match["confidence"],
            "message": f"Best match: {match['name']}"
        }
    
    return {"dog_id": None, "confidence": match["confidence"] if match else 0, "message": "No strong match found"}

@app.get("/my-dogs", response_model=List[schemas.DogResponse])
def get_my_dogs(
    db: Session = Depends(database.get_db), 
    current_user: models.User = Depends(get_current_user)
):
    return db.query(models.Dog).filter(models.Dog.owner_id == current_user.id).all()

@app.get("/services", response_model=List[schemas.ServiceResponse])
def list_services(
    item_type: Optional[str] = None, 
    lat: Optional[float] = None, 
    lon: Optional[float] = None, 
    radius: Optional[float] = 50.0, # default 50km
    db: Session = Depends(database.get_db)
):
    query = db.query(models.Service).filter(models.Service.is_published == True, models.Service.admin_approved == True)
    if item_type:
        query = query.filter(models.Service.item_type == item_type)
    
    services = query.all()
    pin_map = get_active_pin_map(db, "service")
    apply_pin_metadata(services, pin_map)
    
    if lat is not None and lon is not None:
        # Filter and sort by distance
        filtered_services = []
        for s in services:
            if s.latitude is not None and s.longitude is not None:
                dist = calculate_distance(lat, lon, s.latitude, s.longitude)
                if dist <= radius:
                    # Injects distance into the response object temporarily
                    setattr(s, 'distance', round(dist, 2))
                    filtered_services.append(s)
            else:
                filtered_services.append(s)
        
        # Sort by distance
        filtered_services.sort(key=lambda x: (
            0 if getattr(x, 'is_pinned', False) else 1,
            -(getattr(x, 'pin_priority', 0) or 0),
            getattr(x, 'distance', 999999),
        ))
        return filtered_services
        
    return sort_items_with_pins(services, pin_map, secondary_key=lambda s: s.title or "")

@app.post("/services", response_model=schemas.ServiceResponse)
def create_service(
    service: schemas.ServiceCreate, 
    db: Session = Depends(database.get_db),
    current_user: models.User = Depends(get_current_user)
):
    # Apply 23.5% platform fee markup to the provider's price
    final_price = validate_marketplace_base_price(service.price, service.currency)

    new_service = models.Service(
        id=str(uuid.uuid4()),
        provider_id=current_user.id,
        title=service.title,
        description=service.description,
        price=final_price,
        category=service.category,
        item_type=service.item_type,
        image_url=service.image_url,
        latitude=service.latitude,
        longitude=service.longitude,
        address=service.address,
        location_landmark=service.location_landmark,
        is_published=service.is_published,
        currency=service.currency,
        stock_count=service.stock_count,
        slots_available=service.slots_available,
        is_busy=service.is_busy,
        images=service.images,
    )
    db.add(new_service)

    # Handle Nested Form Fields
    if service.form_fields:
        for f_idx, field_data in enumerate(service.form_fields):
            new_field = models.ServiceFormField(
                id=str(uuid.uuid4()),
                service_id=new_service.id,
                field_type=field_data.field_type,
                label=field_data.label,
                options=field_data.options,
                is_required=field_data.is_required,
                sort_order=field_data.sort_order or f_idx
            )
            db.add(new_field)

    db.commit()
    db.refresh(new_service)
    return new_service
    
@app.get("/services/{service_id}", response_model=schemas.ServiceResponse)
def get_service_detail(service_id: str, db: Session = Depends(database.get_db), current_user: models.User = Depends(get_current_user)):
    service = db.query(models.Service).filter(models.Service.id == service_id).first()
    if not service:
        raise HTTPException(status_code=404, detail="Service not found")
    
    # Security: If not approved, only provider or admin can view
    if not service.admin_approved and service.provider_id != current_user.id and current_user.role != models.UserRole.ADMIN:
        raise HTTPException(status_code=403, detail="Service pending admin approval")
        
    return service

@app.get("/services/{service_id}/form-fields", response_model=List[schemas.ServiceFormFieldResponse])
def get_service_form_fields(service_id: str, db: Session = Depends(database.get_db)):
    return db.query(models.ServiceFormField).filter(models.ServiceFormField.service_id == service_id).order_by(models.ServiceFormField.sort_order).all()

@app.post("/services/{service_id}/form-fields")
def save_service_form_fields(
    service_id: str,
    fields: List[schemas.ServiceFormFieldCreate],
    db: Session = Depends(database.get_db),
    current_user: models.User = Depends(get_current_user)
):
    service = db.query(models.Service).filter(models.Service.id == service_id).first()
    if not service:
        raise HTTPException(status_code=404, detail="Service not found")
    if service.provider_id != current_user.id and current_user.role != models.UserRole.ADMIN:
        raise HTTPException(status_code=403, detail="Not authorized")

    # Sync fields
    db.query(models.ServiceFormField).filter(models.ServiceFormField.service_id == service_id).delete()
    for f in fields:
        db.add(models.ServiceFormField(
            id=str(uuid.uuid4()), service_id=service_id, field_type=f.field_type,
            label=f.label, options=f.options, is_required=f.is_required, sort_order=f.sort_order
        ))
    db.commit()
    return {"status": "success"}

@app.get("/services/{service_id}/responses")
def get_service_responses(
    service_id: str,
    db: Session = Depends(database.get_db),
    current_user: models.User = Depends(get_current_user)
):
    service = db.query(models.Service).filter(models.Service.id == service_id).first()
    if not service:
        raise HTTPException(status_code=404, detail="Service not found")
    if service.provider_id != current_user.id and current_user.role != models.UserRole.ADMIN:
        raise HTTPException(status_code=403, detail="Not authorized")

    orders = db.query(models.Order).filter(models.Order.service_id == service_id).all()
    results = []
    for o in orders:
        buyer = db.query(models.User).filter(models.User.id == o.buyer_id).first()
        resps = []
        for r in o.responses:
            resps.append({
                "id": r.id, "field_id": r.field_id,
                "field_label": r.field.label if r.field else "Deleted Field",
                "answer_value": r.answer_value
            })
        results.append({
            "order_id": o.id, "status": o.status, "created_at": o.created_at,
            "buyer": {"full_name": buyer.full_name, "email": buyer.email, 
                      "phone": buyer.phone_number if (o.share_phone or current_user.role == "admin") else None},
            "responses": resps
        })
    return results

PAID_ORDER_STATES = {
    models.OrderStatus.PAID.value,
    models.OrderStatus.COMPLETED.value,
    models.OrderStatus.SETTLED.value,
}

def order_status_value(status):
    if hasattr(status, "value"):
        return status.value
    raw = str(status or "").strip()
    if "." in raw:
        raw = raw.rsplit(".", 1)[-1]
    return raw.lower()

def order_status_filter_values(*statuses):
    values = set()
    for status in statuses:
        value = order_status_value(status)
        if not value:
            continue
        values.add(value)
        values.add(value.upper())
        values.add(f"OrderStatus.{value.upper()}")
    return list(values)

PAID_ORDER_STATUS_VALUES = order_status_filter_values(
    models.OrderStatus.PAID,
    models.OrderStatus.COMPLETED,
    models.OrderStatus.SETTLED,
)
PENDING_ORDER_STATUS_VALUES = order_status_filter_values(models.OrderStatus.PENDING)

def support_status_key(status):
    raw = str(status or "open").strip().lower().replace("_", "-")
    if raw in {"in progress", "in-progress", "inprogress"}:
        return "in-progress"
    if raw == "resolved":
        return "resolved"
    return "open"

def support_status_label(status):
    labels = {
        "open": "Open",
        "in-progress": "In-Progress",
        "resolved": "Resolved",
    }
    return labels.get(support_status_key(status), "Open")

def is_order_paid(order: models.Order) -> bool:
    return order_status_value(order.status) in PAID_ORDER_STATES

def is_pesapal_payment_successful(status_res: dict) -> bool:
    if not isinstance(status_res, dict):
        return False
    status_code = status_res.get("payment_status_code") or status_res.get("status_code")
    status_text = (
        status_res.get("payment_status_description")
        or status_res.get("payment_status")
        or status_res.get("status")
        or ""
    )
    return str(status_code) == "1" or str(status_text).strip().lower() in {"completed", "paid", "success", "successful"}

def mark_order_paid(db: Session, order: models.Order) -> bool:
    current_status = order_status_value(order.status)
    if current_status in PAID_ORDER_STATES:
        return False

    if current_status != models.OrderStatus.PENDING.value:
        return False

    service = db.query(models.Service).filter(models.Service.id == order.service_id).first()
    if service:
        if service.item_type == "products" and service.stock_count is not None:
            service.stock_count = max(service.stock_count - 1, 0)
        elif service.item_type != "products" and service.slots_available is not None:
            service.slots_available = max(service.slots_available - 1, 0)

    order.status = models.OrderStatus.PAID.value
    buyer_points = calculate_karma_reward(order.amount)
    seller_points = calculate_karma_reward(order.payout)
    item_label = "product" if service and service.item_type == "products" else "service"
    item_title = service.title if service else "Marketplace item"

    if buyer_points:
        award_karma(
            db,
            order.buyer_id,
            buyer_points,
            "purchase",
            f"Purchased {item_title}",
            commit=False
        )
        add_notification(
            db,
            order.buyer_id,
            "Purchase confirmed",
            f"Your {item_label} purchase '{item_title}' is confirmed. You earned {buyer_points} points.",
            "purchase",
            commit=False
        )

    if service and service.provider_id and seller_points:
        award_karma(
            db,
            service.provider_id,
            seller_points,
            "sale",
            f"Sold {item_title}",
            commit=False
        )
        add_notification(
            db,
            service.provider_id,
            "New sale",
            f"Your {item_label} '{item_title}' was bought. You earned {seller_points} points and payout is now in escrow.",
            "sale",
            commit=False
        )

    return True


def is_event_paid(event: models.Event) -> bool:
    return float(getattr(event, "ticket_price", 0) or 0) > 0


def sanitize_ticket_tiers(tiers: Optional[List[Dict[str, Any]]], default_currency: str = "KES") -> List[Dict[str, Any]]:
    tiers = tiers or []
    if not isinstance(tiers, list):
        return []
    normalized = []
    for idx, tier in enumerate(tiers):
        if not isinstance(tier, dict):
            continue
        label = str(tier.get("label") or tier.get("name") or "").strip()
        if not label:
            continue
        tier_id = str(tier.get("id") or label.lower().replace(" ", "_") or f"tier_{idx + 1}").strip()
        try:
            price = max(float(tier.get("price", 0) or 0), 0)
        except (TypeError, ValueError):
            price = 0
        normalized.append({
            "id": tier_id,
            "label": label,
            "price": price,
            "currency": str(tier.get("currency") or default_currency or "KES").strip().upper(),
            "description": str(tier.get("description") or "").strip(),
            "requires_justification": bool(tier.get("requires_justification", True)),
        })
    return normalized


def normalize_event_ticket_tiers(event: models.Event) -> List[Dict[str, Any]]:
    return sanitize_ticket_tiers(getattr(event, "ticket_tiers", None), getattr(event, "currency", None) or "KES")


def resolve_event_ticket_tier(event: models.Event, ticket_tier_id: Optional[str]) -> Dict[str, Any]:
    tiers = normalize_event_ticket_tiers(event)
    if not tiers:
        return {
            "id": "standard",
            "label": "Standard",
            "price": max(float(getattr(event, "ticket_price", 0) or 0), 0),
            "currency": str(getattr(event, "currency", None) or "KES").strip().upper(),
            "description": "",
            "requires_justification": False,
        }

    selected_id = str(ticket_tier_id or "").strip()
    if not selected_id:
        raise HTTPException(status_code=400, detail="Choose a registration type")
    for tier in tiers:
        if tier["id"] == selected_id:
            return tier
    raise HTTPException(status_code=400, detail="Invalid registration type")


def is_registration_paid(registration: models.Registration) -> bool:
    return str(getattr(registration, "payment_status", "free") or "free").lower() in {"free", "paid"}


def mark_event_registration_paid(db: Session, registration: models.Registration, tracking_id: Optional[str] = None) -> bool:
    if not registration:
        return False
    if str(registration.payment_status or "").lower() == "paid" and registration.status == "registered":
        return False

    event = db.query(models.Event).filter(models.Event.id == registration.event_id).first()
    if event and event.capacity and event.capacity > 0:
        registered_count = db.query(models.Registration).filter(
            models.Registration.event_id == event.id,
            models.Registration.status.in_(["registered", "checked-in"]),
        ).count()
        if registered_count >= event.capacity and registration.status not in {"registered", "checked-in"}:
            registration.status = "waitlisted"
            registration.payment_status = "paid"
            registration.pesapal_tracking_id = tracking_id or registration.pesapal_tracking_id
            registration.paid_at = datetime.utcnow()
            return True

    registration.status = "registered"
    registration.payment_status = "paid"
    registration.pesapal_tracking_id = tracking_id or registration.pesapal_tracking_id
    registration.paid_at = datetime.utcnow()
    if not registration.ticket_token:
        registration.ticket_token = secrets.token_urlsafe(16)

    if event:
        add_notification(
            db,
            registration.user_id,
            "Event payment confirmed",
            f"Your payment for '{event.title}' is confirmed. Your ticket is ready.",
            "payment",
            commit=False,
        )
        reward = calculate_karma_reward(registration.amount or 0)
        if reward:
            award_karma(
                db,
                registration.user_id,
                reward,
                "event_registration",
                f"Registered for paid event: {event.title}",
                commit=False,
            )
    return True

@app.post("/orders", response_model=schemas.OrderResponse)
def create_order(
    order_data: schemas.OrderCreate, 
    db: Session = Depends(database.get_db),
    current_user: models.User = Depends(get_current_user)
):
    service = db.query(models.Service).filter(models.Service.id == order_data.service_id).first()
    if not service:
        raise HTTPException(status_code=404, detail="Service not found")

    if not service.is_published or not service.admin_approved:
        raise HTTPException(status_code=400, detail="This marketplace item is not available for purchase")

    if service.item_type == "products":
        if service.stock_count is not None and service.stock_count <= 0:
            raise HTTPException(status_code=400, detail="This product is out of stock")
    else:
        if service.is_busy:
            raise HTTPException(status_code=400, detail="This service is currently unavailable")
        if service.slots_available is not None and service.slots_available <= 0:
            raise HTTPException(status_code=400, detail="No slots are available for this service")
        
    # Validate required form fields
    form_fields = db.query(models.ServiceFormField).filter(models.ServiceFormField.service_id == order_data.service_id).all()
    provided = {r.field_id: r.answer_value for r in (order_data.form_responses or [])}
    for field in form_fields:
        if field.is_required and not provided.get(field.id):
            raise HTTPException(status_code=400, detail=f"Question '{field.label}' is required")

    points_redeemed, discount_amount = calculate_karma_redemption(
        current_user,
        service.price,
        order_data.karma_points_to_redeem or 0
    )
    amount_due = round(max(float(service.price or 0) - discount_amount, 1), 2)
    payout = round(amount_due / 1.235, 2)
    commission = round(amount_due - payout, 2)
    
    new_order = models.Order(
        id=str(uuid.uuid4()),
        buyer_id=current_user.id,
        service_id=order_data.service_id,
        amount=amount_due,
        commission=commission,
        payout=payout,
        discount_amount=discount_amount,
        karma_points_redeemed=points_redeemed,
        status=models.OrderStatus.PENDING.value,
        share_phone=order_data.share_phone
    )
    db.add(new_order)
    db.flush()

    if points_redeemed:
        current_user.available_karma = int(current_user.available_karma or 0) - points_redeemed
        db.add(models.KarmaTransaction(
            id=str(uuid.uuid4()),
            user_id=current_user.id,
            amount=-points_redeemed,
            category="purchase_discount",
            description=f"Redeemed {points_redeemed} points for order discount on {service.title}"
        ))
    
    if order_data.form_responses:
        for resp in order_data.form_responses:
            db.add(models.OrderFormResponse(
                id=str(uuid.uuid4()), order_id=new_order.id,
                field_id=resp.field_id, answer_value=resp.answer_value
            ))

    db.commit()
    db.refresh(new_order)
    return new_order

@app.post("/orders/{order_id}/pay")
def pay_order(
    order_id: str,
    db: Session = Depends(database.get_db),
    admin: models.User = Depends(require_admin)
):
    order = db.query(models.Order).filter(models.Order.id == order_id).first()
    if not order:
        raise HTTPException(status_code=404, detail="Order not found")

    if order_status_value(order.status) == models.OrderStatus.CANCELLED.value:
        raise HTTPException(status_code=400, detail="Cancelled orders cannot be marked as paid")

    if is_order_paid(order):
        return {"message": "Order already paid", "status": order.status}

    if not mark_order_paid(db, order):
        raise HTTPException(status_code=400, detail="Only pending orders can be marked as paid")

    db.commit()
    return {"message": "Order payment confirmed by admin", "status": order.status}

@app.post("/orders/{order_id}/cancel")
def cancel_order(
    order_id: str,
    db: Session = Depends(database.get_db),
    current_user: models.User = Depends(get_current_user)
):
    order = db.query(models.Order).filter(models.Order.id == order_id).first()
    if not order:
        raise HTTPException(status_code=404, detail="Order not found")

    is_admin = current_user.role in {models.UserRole.ADMIN.value, models.UserRole.SUPER_ADMIN.value}
    if order.buyer_id != current_user.id and not is_admin:
        raise HTTPException(status_code=403, detail="Not authorized")

    current_status = order_status_value(order.status)
    if current_status == models.OrderStatus.CANCELLED.value:
        return {"message": "Order is already cancelled", "status": order.status}

    if current_status != models.OrderStatus.PENDING.value:
        raise HTTPException(
            status_code=400,
            detail="Only unpaid pending orders can be cancelled. Paid orders require refund support."
        )

    service = db.query(models.Service).filter(models.Service.id == order.service_id).first()
    item_title = service.title if service else "Marketplace item"
    item_label = "product" if service and service.item_type == "products" else "service"

    order.status = models.OrderStatus.CANCELLED.value
    add_notification(
        db,
        order.buyer_id,
        "Order Cancelled",
        f"Your unpaid {item_label} order for '{item_title}' has been cancelled.",
        "order",
        commit=False,
    )
    if service and service.provider_id:
        add_notification(
            db,
            service.provider_id,
            "Pending Order Cancelled",
            f"The pending {item_label} order for '{item_title}' was cancelled before payment. No payout or inventory was affected.",
            "order",
            commit=False,
        )

    db.commit()
    return {"message": "Order cancelled successfully", "status": order.status}

@app.get("/orders/{order_id}/receipt")
def get_order_receipt(order_id: str, db: Session = Depends(database.get_db), current_user: models.User = Depends(get_current_user)):
    order = db.query(models.Order).filter(models.Order.id == order_id).first()
    if not order:
        raise HTTPException(status_code=404, detail="Order not found")
    
    if not is_order_paid(order):
        raise HTTPException(status_code=400, detail="Receipt only available for paid orders")

    # Allow buyer OR provider OR admin to get receipt
    service = db.query(models.Service).filter(models.Service.id == order.service_id).first()
    if not service:
         raise HTTPException(status_code=404, detail="Service not found")

    if current_user.id not in [order.buyer_id, service.provider_id] and current_user.role != models.UserRole.ADMIN:
        raise HTTPException(status_code=403, detail="Not authorized")

    buyer = db.query(models.User).filter(models.User.id == order.buyer_id).first()
    provider = db.query(models.User).filter(models.User.id == service.provider_id).first()

    pdf_content = generate_receipt_pdf(order, service, buyer, provider)
    
    return StreamingResponse(
        io.BytesIO(pdf_content),
        media_type="application/pdf",
        headers={"Content-Disposition": f"attachment; filename=receipt_{order_id}.pdf"}
    )

# --- Buyer Orders & Seller Earnings ---

@app.get("/my-orders")
def get_my_orders(db: Session = Depends(database.get_db), current_user: models.User = Depends(get_current_user)):
    """Buyer sees all their purchases with payment status."""
    orders = db.query(models.Order).filter(models.Order.buyer_id == current_user.id).order_by(models.Order.created_at.desc()).all()
    result = []
    for order in orders:
        service = db.query(models.Service).filter(models.Service.id == order.service_id).first()
        provider = None
        if service:
            provider = db.query(models.User).filter(models.User.id == service.provider_id).first()
        result.append({
            "id": order.id,
            "service_title": service.title if service else "Unknown",
            "service_image": service.image_url if service else None,
            "provider_name": provider.full_name if provider else "Unknown",
            "amount": order.amount,
            "discount_amount": getattr(order, "discount_amount", 0) or 0,
            "karma_points_redeemed": getattr(order, "karma_points_redeemed", 0) or 0,
            "status": order.status,
            "created_at": str(order.created_at) if order.created_at else None
        })
    return result

def get_payout_destination(user: models.User, method: Optional[str] = None) -> Optional[str]:
    payout_method = (method or getattr(user, "payment_method", None) or "").strip().lower()
    if payout_method == "mpesa":
        return user.mpesa_phone_number
    if payout_method == "card":
        return "Pesapal card/bank payout"
    return None

def get_provider_wallet_summary(db: Session, user: models.User):
    services = db.query(models.Service).filter(models.Service.provider_id == user.id).all()
    service_ids = [s.id for s in services]
    total_earned = 0.0
    in_escrow = 0.0
    available_gross = 0.0
    settled = 0.0

    if service_ids:
        orders = db.query(models.Order).filter(
            models.Order.service_id.in_(service_ids),
            models.Order.status.in_(PAID_ORDER_STATUS_VALUES)
        ).all()
        for order in orders:
            payout = float(order.payout or 0)
            status_lower = order_status_value(order.status)
            total_earned += payout
            if status_lower == models.OrderStatus.PAID.value:
                in_escrow += payout
            elif status_lower == models.OrderStatus.COMPLETED.value:
                available_gross += payout
            elif status_lower == models.OrderStatus.SETTLED.value:
                settled += payout

    pending_withdrawal = sum(float(t.amount or 0) for t in db.query(models.Transaction).filter(
        models.Transaction.user_id == user.id,
        models.Transaction.type == "withdrawal",
        models.Transaction.status == "pending",
    ).all())
    available = max(available_gross - pending_withdrawal, 0.0)
    payment_method = (getattr(user, "payment_method", None) or "").strip().lower() or None

    return {
        "currency": user.preferred_currency or "KES",
        "total_earned": round(total_earned, 2),
        "in_escrow": round(in_escrow, 2),
        "available": round(available, 2),
        "available_before_pending_withdrawals": round(available_gross, 2),
        "pending_withdrawal": round(pending_withdrawal, 2),
        "settled": round(settled, 2),
        "withdrawable": round(available, 2),
        "payment_method": payment_method,
        "payout_destination": get_payout_destination(user, payment_method),
    }

@app.get("/wallet/summary")
def wallet_summary(
    db: Session = Depends(database.get_db),
    current_user: models.User = Depends(get_current_user)
):
    return get_provider_wallet_summary(db, current_user)

@app.post("/withdrawals/request")
def request_withdrawal(
    req: schemas.WithdrawalRequest,
    db: Session = Depends(database.get_db),
    current_user: models.User = Depends(get_current_user)
):
    wallet = get_provider_wallet_summary(db, current_user)
    available = float(wallet["withdrawable"] or 0)
    pending_withdrawal = float(wallet.get("pending_withdrawal") or 0)
    if pending_withdrawal > 0:
        raise HTTPException(status_code=400, detail="You already have a pending withdrawal request")

    amount = round(float(req.amount or available), 2)
    if amount <= 0:
        raise HTTPException(status_code=400, detail="No available balance to withdraw")
    if amount > available:
        raise HTTPException(status_code=400, detail="Withdrawal amount exceeds available balance")
    if abs(amount - available) > 0.01:
        raise HTTPException(status_code=400, detail="Withdraw the full available balance for now")

    method = (req.method or current_user.payment_method or "").strip().lower()
    if method not in {"mpesa", "card"}:
        raise HTTPException(status_code=400, detail="Set a payout method first")

    destination = get_payout_destination(current_user, method)
    if method == "mpesa" and not destination:
        raise HTTPException(status_code=400, detail="Add your M-Pesa phone number before requesting withdrawal")

    withdrawal = models.Transaction(
        id=str(uuid.uuid4()),
        order_id=None,
        user_id=current_user.id,
        amount=amount,
        type="withdrawal",
        status="pending",
        payout_method=method,
        destination=destination,
        created_at=datetime.utcnow(),
        processed_at=None,
    )
    db.add(withdrawal)

    add_notification(
        db,
        current_user.id,
        "Withdrawal Requested",
        f"Your KES {amount:,.2f} withdrawal request is pending admin processing.",
        "payout",
        commit=False,
    )

    admins = db.query(models.User).filter(models.User.role.in_([
        models.UserRole.ADMIN.value,
        models.UserRole.SUPER_ADMIN.value,
    ])).all()
    destination_label = destination or "Pesapal card/bank payout"
    for admin_user in admins:
        add_notification(
            db,
            admin_user.id,
            "Seller Withdrawal Request",
            f"{current_user.full_name or current_user.email} requested KES {amount:,.2f} to {method.upper()} ({destination_label}).",
            "payout",
            commit=False,
        )

    db.commit()
    return {
        "message": "Withdrawal request submitted",
        "withdrawal_id": withdrawal.id,
        "amount": amount,
        "status": withdrawal.status,
        "withdrawal": serialize_withdrawal(db, withdrawal),
        "wallet": get_provider_wallet_summary(db, current_user),
    }

def serialize_withdrawal(db: Session, withdrawal: models.Transaction):
    seller = db.query(models.User).filter(models.User.id == withdrawal.user_id).first()
    return {
        "id": withdrawal.id,
        "seller_id": withdrawal.user_id,
        "seller_name": seller.full_name if seller else "Unknown",
        "seller_email": seller.email if seller else None,
        "amount": float(withdrawal.amount or 0),
        "status": withdrawal.status,
        "method": withdrawal.payout_method,
        "destination": withdrawal.destination,
        "created_at": str(withdrawal.created_at) if getattr(withdrawal, "created_at", None) else None,
        "processed_at": str(withdrawal.processed_at) if withdrawal.processed_at else None,
    }

def get_user_withdrawals(db: Session, user_id: str):
    withdrawals = db.query(models.Transaction).filter(
        models.Transaction.user_id == user_id,
        models.Transaction.type == "withdrawal",
    ).order_by(models.Transaction.created_at.desc().nullslast()).limit(25).all()
    return [serialize_withdrawal(db, withdrawal) for withdrawal in withdrawals]

@app.get("/withdrawals")
def list_my_withdrawals(
    db: Session = Depends(database.get_db),
    current_user: models.User = Depends(get_current_user)
):
    return get_user_withdrawals(db, current_user.id)

@app.get("/admin/withdrawals")
def admin_list_withdrawals(
    status_filter: Optional[str] = None,
    db: Session = Depends(database.get_db),
    admin: models.User = Depends(require_admin)
):
    query = db.query(models.Transaction).filter(models.Transaction.type == "withdrawal")
    if status_filter:
        query = query.filter(models.Transaction.status == status_filter.strip().lower())
    withdrawals = query.order_by(models.Transaction.processed_at.desc().nullslast()).all()
    return [serialize_withdrawal(db, withdrawal) for withdrawal in withdrawals]

@app.post("/admin/withdrawals/{withdrawal_id}/complete")
def admin_complete_withdrawal(
    withdrawal_id: str,
    db: Session = Depends(database.get_db),
    admin: models.User = Depends(require_admin)
):
    withdrawal = db.query(models.Transaction).filter(
        models.Transaction.id == withdrawal_id,
        models.Transaction.type == "withdrawal",
    ).first()
    if not withdrawal:
        raise HTTPException(status_code=404, detail="Withdrawal request not found")
    if withdrawal.status != "pending":
        raise HTTPException(status_code=400, detail=f"Withdrawal is already {withdrawal.status}")

    seller = db.query(models.User).filter(models.User.id == withdrawal.user_id).first()
    if not seller:
        raise HTTPException(status_code=404, detail="Seller not found")

    services = db.query(models.Service).filter(models.Service.provider_id == seller.id).all()
    service_ids = [service.id for service in services]
    completed_orders = []
    if service_ids:
        completed_orders = db.query(models.Order).filter(
            models.Order.service_id.in_(service_ids),
            models.Order.status.in_(order_status_filter_values(models.OrderStatus.COMPLETED)),
        ).order_by(models.Order.created_at.asc()).all()

    withdrawal_amount = round(float(withdrawal.amount or 0), 2)
    remaining = withdrawal_amount
    orders_to_settle = []
    for order in completed_orders:
        payout = round(float(order.payout or 0), 2)
        if payout <= remaining + 0.01:
            orders_to_settle.append(order)
            remaining = round(remaining - payout, 2)
        if abs(remaining) <= 0.01:
            break

    if abs(remaining) > 0.01:
        raise HTTPException(
            status_code=400,
            detail="Withdrawal amount does not match available completed payouts"
        )

    for order in orders_to_settle:
        order.status = models.OrderStatus.SETTLED.value

    withdrawal.status = "completed"
    withdrawal.processed_at = datetime.utcnow()
    destination = withdrawal.destination or get_payout_destination(seller, withdrawal.payout_method) or "configured payout destination"
    add_notification(
        db,
        seller.id,
        "Withdrawal Completed",
        f"Your KES {withdrawal_amount:,.2f} withdrawal to {destination} has been marked as paid.",
        "payout",
        commit=False,
    )
    db.commit()
    return {
        "message": f"Withdrawal of KES {withdrawal_amount:,.2f} marked as completed.",
        "withdrawal": serialize_withdrawal(db, withdrawal),
        "wallet": get_provider_wallet_summary(db, seller),
    }

@app.get("/my-earnings")
def get_my_earnings(db: Session = Depends(database.get_db), current_user: models.User = Depends(get_current_user)):
    """Seller sees all earnings from their services, with escrow/available status."""
    # Get all services owned by the current user
    my_services = db.query(models.Service).filter(models.Service.provider_id == current_user.id).all()
    service_ids = [s.id for s in my_services]

    if not service_ids:
        return {
            "wallet": get_provider_wallet_summary(db, current_user),
            "earnings": [],
            "withdrawals": get_user_withdrawals(db, current_user.id),
        }

    # Get all orders for those services that have been paid or beyond
    orders = db.query(models.Order).filter(
        models.Order.service_id.in_(service_ids),
        models.Order.status.in_(PAID_ORDER_STATUS_VALUES)
    ).order_by(models.Order.created_at.desc()).all()

    earnings = []
    total_earned = 0
    in_escrow = 0
    available = 0
    settled = 0

    for order in orders:
        service = db.query(models.Service).filter(models.Service.id == order.service_id).first()
        buyer = db.query(models.User).filter(models.User.id == order.buyer_id).first()
        payout = order.payout or 0
        status_lower = order_status_value(order.status)

        total_earned += payout

        # Determine escrow state:
        # paid = buyer paid but delivery not confirmed → in escrow (greyed out)
        # completed = delivery confirmed but payout not approved → available (highlighted)
        # settled = payout disbursed → settled (green checkmark)
        if status_lower == "paid":
            in_escrow += payout
            escrow_label = "in_escrow"
        elif status_lower == "completed":
            available += payout
            escrow_label = "available"
        elif status_lower == "settled":
            settled += payout
            escrow_label = "settled"
        else:
            escrow_label = "unknown"

        earnings.append({
            "id": order.id,
            "service_title": service.title if service else "Unknown",
            "buyer_name": buyer.full_name if buyer else "Unknown",
            "gross_amount": order.amount,
            "commission": order.commission,
            "payout": payout,
            "discount_amount": getattr(order, "discount_amount", 0) or 0,
            "karma_points_redeemed": getattr(order, "karma_points_redeemed", 0) or 0,
            "order_status": order.status,
            "escrow_status": escrow_label,
            "created_at": str(order.created_at) if order.created_at else None
        })

    return {
        "wallet": {
            **get_provider_wallet_summary(db, current_user),
        },
        "earnings": earnings,
        "withdrawals": get_user_withdrawals(db, current_user.id),
    }

# --- Ratings API ---

@app.post("/ratings", response_model=schemas.RatingResponse)
def submit_rating(rating: schemas.RatingCreate, db: Session = Depends(database.get_db), current_user: models.User = Depends(get_current_user)):
    # Verify order exists and belongs to rater
    order = db.query(models.Order).filter(models.Order.id == rating.order_id).first()
    if not order:
        raise HTTPException(status_code=404, detail="Order not found")
    if order.buyer_id != current_user.id:
        raise HTTPException(status_code=403, detail="Only the buyer can rate this service")
    if not is_order_paid(order):
        raise HTTPException(status_code=400, detail="Can only rate completed/paid services")
    
    # Check if already rated
    existing_rating = db.query(models.Rating).filter(models.Rating.order_id == rating.order_id).first()
    if existing_rating:
        raise HTTPException(status_code=400, detail="Order already rated")

    new_rating = models.Rating(
        id=str(uuid.uuid4()),
        order_id=rating.order_id,
        rater_id=current_user.id,
        rated_id=rating.rated_id,
        score=rating.score,
        comment=rating.comment
    )
    db.add(new_rating)

    # Update rated user's average_rating and total_ratings
    rated_user = db.query(models.User).filter(models.User.id == rating.rated_id).first()
    if rated_user:
        current_total = rated_user.total_ratings or 0
        current_avg = rated_user.average_rating or 0.0
        
        new_total = current_total + 1
        new_avg = ((current_avg * current_total) + rating.score) / new_total
        
        rated_user.total_ratings = new_total
        rated_user.average_rating = new_avg

    db.commit()
    db.refresh(new_rating)
    return new_rating

@app.get("/users/{user_id}/ratings")
def get_user_ratings(user_id: str, db: Session = Depends(database.get_db)):
    ratings = db.query(models.Rating).filter(models.Rating.rated_id == user_id).all()
    if not ratings:
        return {"average_score": 0, "count": 0, "ratings": []}
    
    avg_score = sum(r.score for r in ratings) / len(ratings)
    
    return {
        "average_score": round(avg_score, 1),
        "count": len(ratings),
        "ratings": [
            {
                "score": r.score,
                "comment": r.comment,
                "created_at": r.created_at,
                "rater_name": db.query(models.User.full_name).filter(models.User.id == r.rater_id).scalar()
            } for r in ratings
        ]
    }


# --- Health Records API ---

@app.post("/dogs/{dog_id}/health-records", response_model=schemas.HealthRecordResponse)
def create_health_record(
    dog_id: str,
    record: schemas.HealthRecordCreate,
    db: Session = Depends(database.get_db),
    current_user: models.User = Depends(get_current_user)
):
    dog = db.query(models.Dog).filter(models.Dog.id == dog_id).first()
    if not dog:
        raise HTTPException(status_code=404, detail="Dog not found")
        
    if dog.owner_id != current_user.id: 
        # Check permissions? Maybe provider/admin can add records too?
        # For now, owner + admin + provider
        if current_user.role not in [models.UserRole.ADMIN, models.UserRole.PROVIDER]:
             raise HTTPException(status_code=403, detail="Not authorized")

    new_record = models.HealthRecord(
        id=str(uuid.uuid4()),
        dog_id=dog_id,
        record_type=record.record_type,
        date=datetime.fromisoformat(record.date.replace('Z', '+00:00')),
        next_due_date=datetime.fromisoformat(record.next_due_date.replace('Z', '+00:00')) if record.next_due_date else None,
        notes=record.notes
    )
    db.add(new_record)
    db.commit()
    db.refresh(new_record)
    return new_record

@app.get("/dogs/{dog_id}/health-records", response_model=List[schemas.HealthRecordResponse])
def get_dog_health_records(
    dog_id: str,
    db: Session = Depends(database.get_db),
    current_user: models.User = Depends(get_current_user)
):
    dog = db.query(models.Dog).filter(models.Dog.id == dog_id).first()
    if not dog:
        raise HTTPException(status_code=404, detail="Dog not found")
        
    # Allow owner, admin, provider to see records (provider needs to see vaccination status)
    if dog.owner_id != current_user.id and current_user.role not in [models.UserRole.ADMIN, models.UserRole.PROVIDER]:
        pass # Maybe restrict? For now open to authenticated users who know the dog ID? 
        # Actually better to restrict to owner/admin/provider who has context. 
        # Let's start with strict:
        # raise HTTPException(status_code=403, detail="Not authorized")
    
    return db.query(models.HealthRecord).filter(models.HealthRecord.dog_id == dog_id).all()

# --- Events API ---

@app.post("/events", response_model=schemas.EventResponse)
def create_event(
    event: schemas.EventCreate, 
    db: Session = Depends(database.get_db),
    current_user: models.User = Depends(get_current_user)
):
    # Optional: Check if user is admin or allowed to create events
    if current_user.role not in [models.UserRole.ADMIN, models.UserRole.PROVIDER, models.UserRole.ADMIN.value, models.UserRole.PROVIDER.value, models.UserRole.SUPER_ADMIN.value]:
        raise HTTPException(status_code=403, detail="Not authorized to create events")
    admin_created = is_admin_user(current_user)

    new_event = models.Event(
        id=str(uuid.uuid4()),
        organizer_id=current_user.id,
        title=event.title,
        description=event.description,
        location=event.location,
        poster_url=event.poster_url,
        images=event.images,
        start_time=parse_datetime_value(event.start_time),
        end_time=parse_datetime_value(event.end_time),
        capacity=event.capacity,
        ticket_price=max(float(event.ticket_price or 0), 0),
        currency=(event.currency or "KES").strip().upper(),
        ticket_tiers=sanitize_ticket_tiers(event.ticket_tiers, event.currency or "KES"),
        attendee_type_question=event.attendee_type_question,
        category=event.category,
        is_public=event.is_public,
        admin_created=admin_created,
        scorecard_enabled=event.scorecard_enabled if event.scorecard_enabled is not None else True,
    )
    db.add(new_event)
    db.flush()
    if admin_created:
        ensure_content_pin(
            db,
            "event",
            new_event.id,
            current_user,
            title=new_event.title,
            description=new_event.description,
            image_url=new_event.poster_url,
            priority=150,
            commit=False,
        )
    db.commit()
    db.refresh(new_event)
    return new_event

@app.get("/events", response_model=List[schemas.EventResponse])
def list_events(skip: int = 0, limit: int = 100, db: Session = Depends(database.get_db)):
    pin_map = get_active_pin_map(db, "event")
    events = db.query(models.Event).filter(models.Event.is_public == 1).all()
    for event in events:
        event.registrant_count = db.query(models.Registration).filter(
            models.Registration.event_id == event.id,
            models.Registration.status.in_(["registered", "checked-in"])
        ).count()
    apply_pin_metadata(events, pin_map)
    events = sort_items_with_pins(events, pin_map, secondary_key=lambda e: e.start_time or datetime.max)
    return events[skip:skip + limit]

@app.get("/events/{event_id}", response_model=schemas.EventResponse)
def get_event(event_id: str, db: Session = Depends(database.get_db)):
    event = db.query(models.Event).filter(models.Event.id == event_id).first()
    if not event:
        raise HTTPException(status_code=404, detail="Event not found")
    
    event.registrant_count = db.query(models.Registration).filter(
        models.Registration.event_id == event.id,
        models.Registration.status.in_(["registered", "checked-in"])
    ).count()
    pin = get_active_pin_map(db, "event").get(event.id)
    event.is_pinned = pin is not None
    event.pin_priority = pin.priority if pin else None
    return event

@app.post("/events/{event_id}/register", response_model=schemas.RegistrationResponse)
def register_for_event(
    event_id: str,
    registration: schemas.RegistrationCreate,
    db: Session = Depends(database.get_db),
    current_user: models.User = Depends(get_current_user)
):
    event = db.query(models.Event).filter(models.Event.id == event_id).first()
    if not event:
        raise HTTPException(status_code=404, detail="Event not found")

    selected_tier = resolve_event_ticket_tier(event, registration.ticket_tier_id)
    event_has_tiers = len(normalize_event_ticket_tiers(event)) > 0
    attendee_justification = (registration.attendee_type_justification or "").strip()
    if event_has_tiers and selected_tier.get("requires_justification", True) and len(attendee_justification) < 3:
        raise HTTPException(status_code=400, detail="Tell us briefly why this registration type applies")

    # Check capacity
    paid_event = float(selected_tier["price"] or 0) > 0
    status = "pending_payment" if paid_event else "registered"
    if event.capacity > 0:
        count = db.query(models.Registration).filter(
            models.Registration.event_id == event_id,
            models.Registration.status.in_(["registered", "checked-in"]),
        ).count()
        if count >= event.capacity:
            if registration.join_waitlist:
                status = "waitlisted"
            else:
                raise HTTPException(status_code=400, detail="Event is full")
            
    # Check existing registration
    existing = db.query(models.Registration).filter(
        models.Registration.event_id == event_id,
        models.Registration.user_id == current_user.id
    ).first()
    if existing:
        raise HTTPException(status_code=400, detail="Already registered for this event")
        
    new_reg = models.Registration(
        id=str(uuid.uuid4()),
        event_id=event_id,
        user_id=current_user.id,
        dog_id=registration.dog_id,
        status=status,
        role=registration.role or "attendee",
        share_phone=registration.share_phone,
        ticket_token=secrets.token_urlsafe(16) if not paid_event and status == "registered" else None,
        amount=float(selected_tier["price"] or 0),
        currency=(selected_tier.get("currency") or event.currency or "KES").strip().upper(),
        payment_status="pending" if paid_event else "free",
        ticket_tier_id=selected_tier["id"],
        ticket_tier_label=selected_tier["label"],
        attendee_type_justification=attendee_justification or None,
        pesapal_merchant_reference=None,
    )
    new_reg.pesapal_merchant_reference = new_reg.id
    db.add(new_reg)
    db.flush()
    
    if registration.form_responses:
        for resp in registration.form_responses:
            new_resp = models.RegistrationResponse(
                id=str(uuid.uuid4()),
                registration_id=new_reg.id,
                field_id=resp.field_id,
                answer_value=resp.answer_value
            )
            db.add(new_resp)

    db.commit()
    db.refresh(new_reg)
    return new_reg


@app.post("/event-registrations/{registration_id}/payment/initiate")
async def initiate_event_registration_payment(
    registration_id: str,
    email: str = "",
    phone: str = "0700000000",
    db: Session = Depends(database.get_db),
    current_user: models.User = Depends(get_current_user),
):
    registration = db.query(models.Registration).filter(models.Registration.id == registration_id).first()
    if not registration:
        raise HTTPException(status_code=404, detail="Registration not found")
    if registration.user_id != current_user.id and not is_admin_user(current_user):
        raise HTTPException(status_code=403, detail="Not authorized")
    event = db.query(models.Event).filter(models.Event.id == registration.event_id).first()
    if not event:
        raise HTTPException(status_code=404, detail="Event not found")
    amount_due = max(float(registration.amount or 0), 0)
    if registration.status == "waitlisted":
        raise HTTPException(status_code=400, detail="Waitlisted registrations cannot be paid until a slot is available")
    if amount_due <= 0:
        if registration.status not in {"registered", "checked-in"}:
            registration.status = "registered"
        registration.payment_status = "free"
        if not registration.ticket_token:
            registration.ticket_token = secrets.token_urlsafe(16)
        db.commit()
        return {"message": "This event is free", "payment_success": True, "registration_status": registration.status}
    if str(registration.payment_status or "").lower() == "paid":
        return {"message": "Payment already confirmed", "payment_success": True, "registration_status": registration.status}

    ipn_url = os.getenv("PESAPAL_IPN_URL")
    callback_url = os.getenv("PESAPAL_CALLBACK_URL")
    if not ipn_url or not callback_url:
        raise HTTPException(status_code=500, detail="Pesapal checkout is not configured. Please set PESAPAL_IPN_URL and PESAPAL_CALLBACK_URL.")

    ipn_res = pesapal.register_ipn(ipn_url)
    ipn_id = ipn_res.get("ipn_id")
    if not ipn_id:
        detail = ipn_res.get("error") or ipn_res.get("message") or ipn_res
        raise HTTPException(status_code=502, detail=f"Failed to register IPN with Pesapal: {detail}")

    merchant_reference = registration.pesapal_merchant_reference or registration.id
    registration.pesapal_merchant_reference = merchant_reference
    registration.payment_status = "pending"
    registration.amount = amount_due
    registration.currency = (registration.currency or event.currency or "KES").strip().upper()
    db.commit()

    order_res = pesapal.submit_order(
        order_id=merchant_reference,
        amount=registration.amount,
        description=f"Lovedogs 360 - Event ticket: {event.title} ({registration.ticket_tier_label or 'Standard'})",
        email=email or current_user.email,
        phone=phone or current_user.phone_number or "0700000000",
        callback_url=callback_url,
        ipn_id=ipn_id,
        currency=registration.currency,
    )
    tracking_id = order_res.get("order_tracking_id") or order_res.get("OrderTrackingId")
    if tracking_id:
        registration.pesapal_tracking_id = tracking_id
        db.commit()
    if not order_res.get("redirect_url"):
        detail = order_res.get("error") or order_res.get("message") or order_res
        raise HTTPException(status_code=502, detail=f"Failed to start Pesapal checkout: {detail}")
    return order_res


@app.get("/event-registrations/{registration_id}/payment/status")
async def event_registration_payment_status(
    registration_id: str,
    tracking_id: Optional[str] = None,
    db: Session = Depends(database.get_db),
    current_user: models.User = Depends(get_current_user),
):
    registration = db.query(models.Registration).filter(models.Registration.id == registration_id).first()
    if not registration:
        raise HTTPException(status_code=404, detail="Registration not found")
    if registration.user_id != current_user.id and not is_admin_user(current_user):
        raise HTTPException(status_code=403, detail="Not authorized")

    status_res = None
    payment_success = is_registration_paid(registration)
    tracking = tracking_id or registration.pesapal_tracking_id
    if tracking and not payment_success:
        status_res = pesapal.get_transaction_status(tracking)
        if is_pesapal_payment_successful(status_res):
            if mark_event_registration_paid(db, registration, tracking):
                db.commit()
            payment_success = True
        else:
            payment_success = is_registration_paid(registration)

    return {
        "registration_id": registration.id,
        "registration_status": registration.status,
        "payment_status": registration.payment_status,
        "payment_success": payment_success,
        "ticket_token": registration.ticket_token,
        "payment_response": status_res,
    }


@app.post("/events/{event_id}/save")
def toggle_save_event(event_id: str, db: Session = Depends(database.get_db), current_user: models.User = Depends(get_current_user)):
    event = db.query(models.Event).filter(models.Event.id == event_id).first()
    if not event:
        raise HTTPException(status_code=404, detail="Event not found")
        
    existing = db.query(models.SavedEvent).filter(
        models.SavedEvent.event_id == event_id,
        models.SavedEvent.user_id == current_user.id
    ).first()
    
    if existing:
        db.delete(existing)
        db.commit()
        return {"status": "unsaved"}
    else:
        new_save = models.SavedEvent(
            id=str(uuid.uuid4()),
            event_id=event_id,
            user_id=current_user.id
        )
        db.add(new_save)
        db.commit()
        return {"status": "saved"}

@app.get("/saved-events", response_model=List[schemas.SavedEventResponse])
def get_saved_events(db: Session = Depends(database.get_db), current_user: models.User = Depends(get_current_user)):
    return db.query(models.SavedEvent).filter(models.SavedEvent.user_id == current_user.id).all()

@app.get("/events/{event_id}/form-fields", response_model=List[schemas.EventFormFieldResponse])
def get_event_form_fields(event_id: str, db: Session = Depends(database.get_db)):
    return db.query(models.EventFormField).filter(models.EventFormField.event_id == event_id).order_by(models.EventFormField.sort_order).all()

@app.post("/events/{event_id}/form-fields", response_model=List[schemas.EventFormFieldResponse])
def save_event_form_fields(
    event_id: str,
    fields: List[schemas.EventFormFieldCreate],
    db: Session = Depends(database.get_db),
    current_user: models.User = Depends(get_current_user)
):
    event = db.query(models.Event).filter(models.Event.id == event_id).first()
    if not event:
        raise HTTPException(status_code=404, detail="Event not found")
        
    if event.organizer_id != current_user.id and current_user.role != models.UserRole.ADMIN:
        raise HTTPException(status_code=403, detail="Not authorized to edit this event's form")
        
    db.query(models.EventFormField).filter(models.EventFormField.event_id == event_id).delete()
    
    new_fields = []
    for field in fields:
        new_field = models.EventFormField(
            id=str(uuid.uuid4()),
            event_id=event_id,
            field_type=field.field_type,
            label=field.label,
            options=field.options,
            is_required=field.is_required,
            sort_order=field.sort_order
        )
        db.add(new_field)
        new_fields.append(new_field)
        
    db.commit()
    for field in new_fields:
        db.refresh(field)
    return new_fields

@app.get("/events/{event_id}/responses", response_model=List[schemas.RegistrationWithResponses])
def get_event_responses(
    event_id: str,
    db: Session = Depends(database.get_db),
    current_user: models.User = Depends(get_current_user)
):
    event = db.query(models.Event).filter(models.Event.id == event_id).first()
    if not event:
        raise HTTPException(status_code=404, detail="Event not found")
        
    if event.organizer_id != current_user.id and current_user.role != models.UserRole.ADMIN:
        raise HTTPException(status_code=403, detail="Not authorized to view responses")
        
    registrations = db.query(models.Registration).filter(models.Registration.event_id == event_id).all()
    
    result = []
    for reg in registrations:
        reg_dict = {
            "id": reg.id,
            "event_id": reg.event_id,
            "user_id": reg.user_id,
            "status": reg.status,
            "role": reg.role,
            "share_phone": reg.share_phone,
            "amount": reg.amount,
            "currency": reg.currency,
            "payment_status": reg.payment_status,
            "ticket_tier_id": reg.ticket_tier_id,
            "ticket_tier_label": reg.ticket_tier_label,
            "attendee_type_justification": reg.attendee_type_justification,
            "pesapal_tracking_id": reg.pesapal_tracking_id,
            "paid_at": reg.paid_at,
            "created_at": reg.created_at,
            "responses": reg.responses
        }
        
        user = reg.user
        if user:
            reg_dict["user_name"] = user.full_name
            reg_dict["user_email"] = user.email
            if reg.share_phone or current_user.role == models.UserRole.ADMIN:
                reg_dict["user_phone"] = user.phone_number
                
        dog = reg.dog
        if dog:
            reg_dict["dog_name"] = dog.name
            
        result.append(reg_dict)

    return result


# --- Mbwa Rafiki Coexistence Scorecard API ---

@app.get("/scorecard/questions", response_model=List[schemas.ScorecardQuestionResponse])
def get_scorecard_questions(survey_type: Optional[str] = None, db: Session = Depends(database.get_db)):
    query = db.query(models.ScorecardQuestion).filter(models.ScorecardQuestion.is_active == True)
    if survey_type:
        normalized = survey_type.strip().lower()
        if normalized not in {"baseline", "followup"}:
            raise HTTPException(status_code=400, detail="survey_type must be baseline or followup")
        query = query.filter(models.ScorecardQuestion.survey_type == normalized)
    return query.order_by(models.ScorecardQuestion.survey_type, models.ScorecardQuestion.sort_order).all()


@app.post("/events/{event_id}/scorecard/surveys", response_model=schemas.ScorecardSurveyResult)
def submit_scorecard_survey(
    event_id: str,
    payload: schemas.ScorecardSurveyCreate,
    db: Session = Depends(database.get_db),
):
    survey_type = payload.survey_type.strip().lower()
    if survey_type not in {"baseline", "followup"}:
        raise HTTPException(status_code=400, detail="survey_type must be baseline or followup")

    event = db.query(models.Event).filter(models.Event.id == event_id).first()
    if not event:
        raise HTTPException(status_code=404, detail="Event not found")
    if event.scorecard_enabled is False:
        raise HTTPException(status_code=400, detail="Scorecard is not enabled for this event")

    questions = db.query(models.ScorecardQuestion).filter(
        models.ScorecardQuestion.survey_type == survey_type,
        models.ScorecardQuestion.is_active == True,
    ).order_by(models.ScorecardQuestion.sort_order).all()
    question_map = {q.id: q for q in questions}
    provided = {r.question_id: r for r in payload.responses}

    for question in questions:
        response = provided.get(question.id)
        if not response:
            raise HTTPException(status_code=400, detail=f"Missing response for: {question.prompt}")
        if question.question_type == "likert" and response.answer_numeric is None:
            raise HTTPException(status_code=400, detail=f"Select a 1-5 score for: {question.prompt}")
        if question.question_type == "open" and not (response.answer_text or "").strip():
            raise HTTPException(status_code=400, detail=f"Answer required for: {question.prompt}")

    category_scores, coexistence_index = calculate_scorecard_scores(question_map, payload.responses)
    participant = find_or_create_scorecard_participant(db, event_id, payload.participant)

    survey = models.ScorecardSurvey(
        id=str(uuid.uuid4()),
        event_id=event_id,
        participant_id=participant.id,
        survey_type=survey_type,
        category_scores=category_scores,
        coexistence_index=coexistence_index,
    )
    db.add(survey)
    db.flush()

    for response in payload.responses:
        question = question_map.get(response.question_id)
        if not question:
            continue
        db.add(models.ScorecardResponse(
            id=str(uuid.uuid4()),
            survey_id=survey.id,
            question_id=response.question_id,
            answer_numeric=response.answer_numeric if question.question_type == "likert" else None,
            answer_text=(response.answer_text or "").strip() if question.question_type == "open" else None,
        ))

    db.commit()
    db.refresh(survey)
    baseline_score, followup_score, change = participant_score_pair(db, event_id, participant.id)
    return {
        "id": survey.id,
        "event_id": survey.event_id,
        "participant_id": survey.participant_id,
        "survey_type": survey.survey_type,
        "category_scores": survey.category_scores or {},
        "coexistence_index": survey.coexistence_index or 0.0,
        "baseline_score": baseline_score,
        "followup_score": followup_score,
        "percentage_change": change,
        "created_at": survey.created_at,
    }


@app.get("/admin/scorecard/events")
def admin_scorecard_events(db: Session = Depends(database.get_db), admin: models.User = Depends(require_admin)):
    events = db.query(models.Event).order_by(models.Event.start_time.desc()).all()
    result = []
    for event in events:
        dashboard = scorecard_dashboard_payload(db, event.id)
        result.append({
            "id": event.id,
            "title": event.title,
            "start_time": str(event.start_time),
            "location": event.location,
            "scorecard_enabled": event.scorecard_enabled,
            "admin_created": event.admin_created,
            "total_participants": dashboard["total_participants"],
            "baseline_surveys_completed": dashboard["baseline_surveys_completed"],
            "followup_surveys_completed": dashboard["followup_surveys_completed"],
        })
    return result


@app.get("/admin/scorecard/{event_id}/dashboard")
def admin_scorecard_dashboard(event_id: str, db: Session = Depends(database.get_db), admin: models.User = Depends(require_admin)):
    return scorecard_dashboard_payload(db, event_id)


@app.get("/admin/scorecard/{event_id}/surveys")
def admin_scorecard_raw_surveys(event_id: str, db: Session = Depends(database.get_db), admin: models.User = Depends(require_admin)):
    participants = db.query(models.ScorecardParticipant).filter(models.ScorecardParticipant.event_id == event_id).all()
    result = []
    for participant in participants:
        for survey in sorted(participant.surveys, key=lambda s: s.created_at, reverse=True):
            result.append({
                "participant": {
                    "id": participant.id,
                    "full_name": participant.full_name,
                    "anonymous_code": participant.anonymous_code,
                    "phone_number": participant.phone_number,
                    "county": participant.county,
                    "community_location": participant.community_location,
                    "user_type": participant.user_type,
                    "participation_type": participant.participation_type,
                    "consent": participant.consent,
                },
                "survey": {
                    "id": survey.id,
                    "survey_type": survey.survey_type,
                    "category_scores": survey.category_scores or {},
                    "coexistence_index": survey.coexistence_index,
                    "created_at": str(survey.created_at),
                },
                "responses": [
                    {
                        "question": response.question.prompt if response.question else None,
                        "category": response.question.category if response.question else None,
                        "question_type": response.question.question_type if response.question else None,
                        "answer_numeric": response.answer_numeric,
                        "answer_text": response.answer_text,
                    }
                    for response in survey.responses
                ],
            })
    return result


@app.post("/admin/scorecard/{event_id}/prompt-followup")
def admin_prompt_scorecard_followup(event_id: str, db: Session = Depends(database.get_db), admin: models.User = Depends(require_admin)):
    event = db.query(models.Event).filter(models.Event.id == event_id).first()
    if not event:
        raise HTTPException(status_code=404, detail="Event not found")
    event.follow_up_requested_at = datetime.utcnow()

    registrations = db.query(models.Registration).filter(models.Registration.event_id == event_id).all()
    notified = 0
    for registration in registrations:
        if registration.user_id:
            add_notification(
                db,
                registration.user_id,
                "Mbwa Rafiki follow-up",
                f"Please complete the follow-up Coexistence Scorecard for {event.title}.",
                "info",
                commit=False,
            )
            notified += 1

    participants_with_phone = db.query(models.ScorecardParticipant).filter(
        models.ScorecardParticipant.event_id == event_id,
        models.ScorecardParticipant.phone_number.isnot(None),
    ).count()
    db.commit()
    return {
        "message": "Follow-up prompt recorded",
        "notified_registrants": notified,
        "participants_with_phone": participants_with_phone,
        "follow_up_requested_at": str(event.follow_up_requested_at),
    }


@app.post("/admin/scorecard/{event_id}/evidence")
def admin_add_scorecard_evidence(
    event_id: str,
    evidence_in: schemas.ScorecardEvidenceCreate,
    db: Session = Depends(database.get_db),
    admin: models.User = Depends(require_admin),
):
    event = db.query(models.Event).filter(models.Event.id == event_id).first()
    if not event:
        raise HTTPException(status_code=404, detail="Event not found")
    evidence = models.ScorecardEvidence(
        id=str(uuid.uuid4()),
        event_id=event_id,
        evidence_type=evidence_in.evidence_type,
        url=evidence_in.url,
        notes=evidence_in.notes,
        created_by_id=admin.id,
    )
    db.add(evidence)
    db.commit()
    db.refresh(evidence)
    return {
        "id": evidence.id,
        "evidence_type": evidence.evidence_type,
        "url": evidence.url,
        "notes": evidence.notes,
        "created_at": str(evidence.created_at),
    }


@app.post("/admin/scorecard/{event_id}/reporting")
def admin_save_scorecard_reporting(
    event_id: str,
    fields: schemas.ScorecardReportingFields,
    db: Session = Depends(database.get_db),
    admin: models.User = Depends(require_admin),
):
    event = db.query(models.Event).filter(models.Event.id == event_id).first()
    if not event:
        raise HTTPException(status_code=404, detail="Event not found")
    report = get_or_create_reporting_export(db, event_id, admin)
    field_values = fields.model_dump() if hasattr(fields, "model_dump") else fields.dict()
    report.fields = {**DEFAULT_REPORTING_FIELDS, **field_values}
    report.created_by_id = report.created_by_id or admin.id
    report.updated_at = datetime.utcnow()
    db.commit()
    return {"message": "Reporting fields saved", "reporting_fields": report.fields}

# --- Health & Wellness Hub API ---

@app.get("/health/wellness-score")
def get_wellness_score(
    db: Session = Depends(database.get_db),
    current_user: models.User = Depends(get_current_user)
):
    """Calculate aggregated wellness scores across all check-ins and health records"""
    # Fetch most recent check-in
    latest_checkin = db.query(models.CheckInData).filter(
        models.CheckInData.user_id == current_user.id
    ).order_by(models.CheckInData.created_at.desc()).first()
    
    if not latest_checkin:
        return {
            "overall_score": 0,
            "who5_score": 0,
            "pss_score": 0,
            "relationship_score": 0,
            "welfare_score": 0,
            "has_data": False
        }
        
    who5 = wellness_utils.calculate_who5_score(latest_checkin.who5_answers or {})
    pss = wellness_utils.calculate_pss10_score(latest_checkin.pss10_answers or {})
    rel = wellness_utils.calculate_relationship_score(latest_checkin.relationship_answers or {})
    wel = wellness_utils.calculate_dog_welfare_score(latest_checkin.welfare_snapshot or {})
    
    # Overall score (weighted average)
    # Give higher weight to WHO5 and Welfare
    overall = int((who5 * 0.3) + (wel * 0.3) + (rel * 0.2) + ((100 - pss) * 0.2))
    
    return {
        "overall_score": overall,
        "who5_score": who5,
        "pss_score": pss,
        "relationship_score": rel,
        "welfare_score": wel,
        "has_data": True,
        "last_checkin": latest_checkin.created_at
    }

@app.get("/health/summary")
def get_health_summary(
    db: Session = Depends(database.get_db),
    current_user: models.User = Depends(get_current_user)
):
    """Aggregate health data across all user dogs for dashboard display"""
    # 1. Get recent wellness checkin info
    latest_checkin = db.query(models.CheckInData).filter(
        models.CheckInData.user_id == current_user.id
    ).order_by(models.CheckInData.created_at.desc()).first()
    
    score = 0
    if latest_checkin:
        who5 = wellness_utils.calculate_who5_score(latest_checkin.who5_answers or {})
        wel = wellness_utils.calculate_dog_welfare_score(latest_checkin.welfare_snapshot or {})
        rel = wellness_utils.calculate_relationship_score(latest_checkin.relationship_answers or {})
        pss = wellness_utils.calculate_pss10_score(latest_checkin.pss10_answers or {})
        score = int((who5 * 0.3) + (wel * 0.3) + (rel * 0.2) + ((100 - pss) * 0.2))

    # 2. Get upcoming health records across all dogs
    user_dogs = db.query(models.Dog).filter(models.Dog.owner_id == current_user.id).all()
    dog_ids = [d.id for d in user_dogs]
    
    now = datetime.utcnow()
    upcoming_record = None
    if dog_ids:
        upcoming_record = db.query(models.HealthRecord).filter(
            models.HealthRecord.dog_id.in_(dog_ids),
            models.HealthRecord.next_due_date > now
        ).order_by(models.HealthRecord.next_due_date.asc()).first()
        
    alert_text = None
    if upcoming_record:
        dog = next((d for d in user_dogs if d.id == upcoming_record.dog_id), None)
        dog_name = dog.name if dog else "Your dog"
        days_until = (upcoming_record.next_due_date.replace(tzinfo=None) - now).days
        if days_until == 0:
            alert_text = f"{dog_name}'s {upcoming_record.record_type} is due TODAY!"
        elif days_until < 14:
            alert_text = f"{dog_name}'s {upcoming_record.record_type} is due in {days_until} days."
        else:
            alert_text = f"Upcoming: {dog_name}'s {upcoming_record.record_type} on {upcoming_record.next_due_date.strftime('%b %d')}."
            
    return {
        "overall_score": score,
        "upcoming_alert": alert_text,
        "has_data": len(user_dogs) > 0
    }


@app.get("/health/advisor/{dog_id}")
def get_health_advisor_insights(
    dog_id: str,
    db: Session = Depends(database.get_db),
    current_user: models.User = Depends(get_current_user)
):
    """Provide AI-driven health insights with heuristic fallback"""
    dog = db.query(models.Dog).filter(models.Dog.id == dog_id).first()
    if not dog:
        raise HTTPException(status_code=404, detail="Dog not found")
        
    # Gather Context for AI
    # 1. Recent Health Records
    records = db.query(models.HealthRecord).filter(
        models.HealthRecord.dog_id == dog_id
    ).order_by(models.HealthRecord.date.desc()).limit(5).all()
    
    # 2. Latest Wellness Stats
    latest_checkin = db.query(models.CheckInData).filter(
        models.CheckInData.user_id == current_user.id,
        models.CheckInData.dog_id == dog_id
    ).order_by(models.CheckInData.created_at.desc()).first()
    
    wellness_stats = {}
    if latest_checkin:
        wellness_stats = {
            "who5_score": wellness_utils.calculate_who5_score(latest_checkin.who5_answers or {}),
            "pss_score": wellness_utils.calculate_pss10_score(latest_checkin.pss10_answers or {}),
            "relationship_score": wellness_utils.calculate_relationship_score(latest_checkin.relationship_answers or {}),
            "welfare_score": wellness_utils.calculate_dog_welfare_score(latest_checkin.welfare_snapshot or {}),
        }
        wellness_stats["overall_score"] = int((wellness_stats["who5_score"] * 0.3) + (wellness_stats["welfare_score"] * 0.3) + (wellness_stats["relationship_score"] * 0.2) + ((100 - wellness_stats["pss_score"]) * 0.2))

    # Try Gemini AI First
    ai_response = gemini_advisor.generate_health_insights(
        dog_data={"name": dog.name, "breed": dog.breed, "age": dog.age, "weight": dog.weight},
        records=[{"record_type": r.record_type, "notes": r.notes, "date": str(r.date)} for r in records],
        wellness_stats=wellness_stats
    )
    
    if ai_response:
        return {
            "dog_name": dog.name,
            "breed": dog.breed,
            "insights": ai_response.get("insights", []),
            "pro_tip": ai_response.get("pro_tip", "Stay consistent with health checks!"),
            "engine": "Gemini AI"
        }

    # Fallback to Heuristics if AI fails
    insights = []
    breed_key = (dog.breed or "").lower()
    if "gsd" in breed_key or "german shepherd" in breed_key:
        insights.append("German Shepherds are prone to hip dysplasia. Ensure regular joint-focused exercise.")
    elif "rot" in breed_key or "rottweiler" in breed_key:
        insights.append("Rottweilers have a higher risk of heart issues. Maintain a lean weight.")
        
    latest_vax = next((r for r in records if r.record_type == "vaccination"), None)
    if not latest_vax:
        insights.append("We haven't recorded any vaccinations yet. Please update medical history.")
        
    if not insights:
        insights.append("Continue tracking daily activity and nutrition.")
        
    return {
        "dog_name": dog.name,
        "breed": dog.breed,
        "insights": insights,
        "pro_tip": "Regular check-ins improve accuracy by 40%.",
        "engine": "Heuristic Engine"
    }


@app.get("/my-registrations", response_model=List[schemas.RegistrationResponse])
def get_my_registrations(db: Session = Depends(database.get_db), current_user: models.User = Depends(get_current_user)):
    return db.query(models.Registration).filter(models.Registration.user_id == current_user.id).all()


@app.put("/services/{service_id}", response_model=schemas.ServiceResponse)
def update_service(
    service_id: str,
    service_update: schemas.ServiceBase,
    db: Session = Depends(database.get_db),
    current_user: models.User = Depends(get_current_user)
):
    service = db.query(models.Service).filter(models.Service.id == service_id).first()
    if not service:
        raise HTTPException(status_code=404, detail="Service not found")
    
    if service.provider_id != current_user.id and current_user.role != models.UserRole.ADMIN:
        raise HTTPException(status_code=403, detail="Not authorized")
    
    update_data = service_update.dict(exclude_unset=True)
    # Apply the same marketplace markup used during creation.
    if "price" in update_data:
        update_data["price"] = validate_marketplace_base_price(
            update_data["price"],
            update_data.get("currency", service.currency),
        )

    for key, value in update_data.items():
        setattr(service, key, value)
    
    db.commit()
    db.refresh(service)
    return service

@app.delete("/services/{service_id}")
def delete_service(
    service_id: str,
    db: Session = Depends(database.get_db),
    current_user: models.User = Depends(get_current_user)
):
    service = db.query(models.Service).filter(models.Service.id == service_id).first()
    if not service:
        raise HTTPException(status_code=404, detail="Service not found")
    
    if service.provider_id != current_user.id and current_user.role != models.UserRole.ADMIN:
        raise HTTPException(status_code=403, detail="Not authorized")
    
    db.delete(service)
    db.commit()
    return {"message": "Service deleted successfully"}

# ... (Orders API) ...

import openpyxl
from openpyxl.styles import Font

@app.post("/registrations/{registration_id}/checkin", response_model=schemas.RegistrationResponse)
def check_in(
    registration_id: str,
    db: Session = Depends(database.get_db),
    current_user: models.User = Depends(get_current_user)
):
    # Verify registration exists and belongs to user (or admin)
    reg = db.query(models.Registration).filter(models.Registration.id == registration_id).first()
    if not reg:
        raise HTTPException(status_code=404, detail="Registration not found")
        
    # Allow self-checkin (if event allows) or Admin/Provider checkin
    if reg.user_id != current_user.id and current_user.role not in [models.UserRole.ADMIN, models.UserRole.PROVIDER]:
        raise HTTPException(status_code=403, detail="Not authorized")
    if not is_registration_paid(reg):
        raise HTTPException(status_code=400, detail="Payment is not confirmed for this registration")

    reg.status = "checked-in"
    reg.check_in_time = datetime.utcnow()
    db.commit()
    db.refresh(reg)
    return reg

@app.get("/admin/export")
def export_data(
    type: str, 
    event_id: Optional[str] = None, 
    date_from: Optional[str] = None,
    location: Optional[str] = None,
    db: Session = Depends(database.get_db),
    current_user: models.User = Depends(get_current_user)
):
    # Expanded Access: Allow Admins AND Partners/Providers (if authorized)
    if current_user.role not in [models.UserRole.ADMIN.value, models.UserRole.SUPER_ADMIN.value, models.UserRole.PROVIDER.value]:
        raise HTTPException(status_code=403, detail="Not authorized")
    valid_export_types = {"registrations", "events", "users", "orders", "dogs", "cases", "community", "support", "scorecard"}
    if type not in valid_export_types:
        raise HTTPException(status_code=400, detail=f"Unsupported export type: {type}")
    if type == "scorecard" and not is_admin_user(current_user):
        raise HTTPException(status_code=403, detail="Only admins can export raw scorecard data")
        
    # Create Workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"{type.capitalize()} Data"
    
    # Headers Style
    header_font = Font(bold=True)
    
    if type == "registrations":
        query = db.query(models.Registration)
        if event_id:
            query = query.filter(models.Registration.event_id == event_id)
        
        registrations = query.all()
        headers = ["Registration ID", "Event ID", "Event Title", "User ID", "User Name", "User Email", "Dog ID", "Dog Name", "Status", "Role", "Registration Type", "Type Justification", "Payment Status", "Amount", "Currency", "Pesapal Tracking ID", "Paid At", "Check-in Time", "Created At"]
        ws.append(headers)
        for cell in ws[1]: cell.font = header_font
        
        for reg in registrations:
            event = db.query(models.Event).filter(models.Event.id == reg.event_id).first()
            user = db.query(models.User).filter(models.User.id == reg.user_id).first()
            dog = db.query(models.Dog).filter(models.Dog.id == reg.dog_id).first() if reg.dog_id else None
            ws.append([
                reg.id, reg.event_id, event.title if event else None,
                reg.user_id, user.full_name if user else None, user.email if user else None,
                reg.dog_id, dog.name if dog else None,
                reg.status, reg.role, reg.ticket_tier_label, reg.attendee_type_justification,
                reg.payment_status, reg.amount, reg.currency,
                reg.pesapal_tracking_id, str(reg.paid_at), str(reg.check_in_time), str(reg.created_at)
            ])
            
    elif type == "events":
        events = db.query(models.Event).all()
        headers = ["Event ID", "Title", "Date", "Location", "Poster URL", "Ticket Price", "Currency", "Ticket Tiers", "Attendee Type Question", "Organizer ID", "Organizer Name", "Organizer Email", "Category", "Registrations", "Paid Registrations", "Event Revenue", "Check-ins"]
        ws.append(headers)
        for cell in ws[1]: cell.font = header_font
        for ev in events:
            organizer = db.query(models.User).filter(models.User.id == ev.organizer_id).first()
            registrations = db.query(models.Registration).filter(models.Registration.event_id == ev.id).count()
            paid_registrations = db.query(models.Registration).filter(
                models.Registration.event_id == ev.id,
                models.Registration.payment_status == "paid",
            ).count()
            event_revenue = db.query(func.coalesce(func.sum(models.Registration.amount), 0)).filter(
                models.Registration.event_id == ev.id,
                models.Registration.payment_status == "paid",
            ).scalar()
            checkins = db.query(models.Registration).filter(models.Registration.event_id == ev.id, models.Registration.status == "checked-in").count()
            ws.append([
                ev.id, ev.title, str(ev.start_time), ev.location, ev.poster_url,
                ev.ticket_price, ev.currency, json.dumps(ev.ticket_tiers or []), ev.attendee_type_question,
                ev.organizer_id,
                organizer.full_name if organizer else None, organizer.email if organizer else None,
                ev.category, registrations, paid_registrations, float(event_revenue or 0), checkins
            ])

    elif type == "users":
        users = db.query(models.User).all()
        headers = ["User ID", "Full Name", "Email", "Role", "Phone", "Country", "Dogs", "Listings", "Orders", "Paid Orders", "Average Rating", "Total Ratings", "Created At"]
        ws.append(headers)
        for cell in ws[1]: cell.font = header_font
        for u in users:
            dog_count = db.query(models.Dog).filter(models.Dog.owner_id == u.id).count()
            listing_count = db.query(models.Service).filter(models.Service.provider_id == u.id).count()
            order_count = db.query(models.Order).filter(models.Order.buyer_id == u.id).count()
            paid_order_count = db.query(models.Order).filter(
                models.Order.buyer_id == u.id,
                models.Order.status.in_(PAID_ORDER_STATUS_VALUES),
            ).count()
            ws.append([
                u.id, u.full_name, u.email, u.role, u.phone_number, u.country,
                dog_count, listing_count, order_count, paid_order_count,
                u.average_rating or 0, u.total_ratings or 0, str(u.created_at)
            ])

    elif type == "orders":
        orders = db.query(models.Order).all()
        headers = ["Order ID", "Buyer ID", "Buyer Name", "Buyer Email", "Service ID", "Service Title", "Provider ID", "Provider Name", "Amount", "Discount", "Points Redeemed", "Paid Amount", "Commission", "Paid Commission", "Payout", "Paid Payout", "Status", "Is Paid", "Created At"]
        ws.append(headers)
        for cell in ws[1]: cell.font = header_font
        for o in orders:
            buyer = db.query(models.User).filter(models.User.id == o.buyer_id).first()
            service = db.query(models.Service).filter(models.Service.id == o.service_id).first()
            provider = db.query(models.User).filter(models.User.id == service.provider_id).first() if service else None
            status_value = order_status_value(o.status)
            paid = status_value in PAID_ORDER_STATES
            amount = float(o.amount or 0)
            commission = float(o.commission or 0)
            payout = float(o.payout or 0)
            discount_amount = float(getattr(o, "discount_amount", 0) or 0)
            karma_points_redeemed = int(getattr(o, "karma_points_redeemed", 0) or 0)
            ws.append([
                o.id, o.buyer_id, buyer.full_name if buyer else None, buyer.email if buyer else None,
                o.service_id, service.title if service else None, service.provider_id if service else None,
                provider.full_name if provider else None, amount, discount_amount, karma_points_redeemed, amount if paid else 0,
                commission, commission if paid else 0, payout, payout if paid else 0,
                status_value, "Yes" if paid else "No", str(o.created_at)
            ])

    elif type == "dogs":
        dogs = db.query(models.Dog).all()
        headers = ["Dog ID", "Name", "Breed", "Owner ID", "Owner Name", "Owner Email", "Age", "Health Records", "Nose-PID", "Created At"]
        ws.append(headers)
        for cell in ws[1]: cell.font = header_font
        for d in dogs:
            owner = db.query(models.User).filter(models.User.id == d.owner_id).first()
            health_records = db.query(models.HealthRecord).filter(models.HealthRecord.dog_id == d.id).count()
            ws.append([
                d.id, d.name, d.breed, d.owner_id, owner.full_name if owner else None,
                owner.email if owner else None, d.age, health_records,
                "Yes" if d.nose_print_descriptor or d.nose_print_image else "No",
                str(d.created_at) if hasattr(d, 'created_at') else "N/A"
            ])

    elif type == "cases":
        cases = db.query(models.CaseReport).all()
        headers = ["Case ID", "Title", "Type", "Status", "Author ID", "Author Name", "Author Email", "Approved", "Rejection Reason", "Created At"]
        ws.append(headers)
        for cell in ws[1]: cell.font = header_font
        for c in cases:
            author = db.query(models.User).filter(models.User.id == c.author_id).first()
            ws.append([
                c.id, c.title, c.case_type, c.status, c.author_id,
                author.full_name if author else None, author.email if author else None,
                "Yes" if c.is_approved else "No", c.rejection_reason, str(c.created_at)
            ])

    elif type == "community":
        posts = db.query(models.CommunityMessage).all()
        headers = ["Post ID", "Author ID", "Author Name", "Author Email", "Content", "Reactions", "Flags", "Hidden", "Created At"]
        ws.append(headers)
        for cell in ws[1]: cell.font = header_font
        for p in posts:
            author = db.query(models.User).filter(models.User.id == p.author_id).first()
            reaction_count = db.query(models.ChatReaction).filter(models.ChatReaction.message_id == p.id).count()
            ws.append([
                p.id, p.author_id, author.full_name if author else None, author.email if author else None,
                p.content, reaction_count, p.flag_count, "Yes" if p.is_hidden else "No", str(p.created_at)
            ])

    elif type == "support":
        tickets = db.query(models.SupportTicket).all()
        headers = ["Ticket ID", "User ID", "User Name", "User Email", "Subject", "Status", "Admin Reply", "Created At", "Updated At"]
        ws.append(headers)
        for cell in ws[1]: cell.font = header_font
        for t in tickets:
            user = db.query(models.User).filter(models.User.id == t.user_id).first()
            ws.append([
                t.id, t.user_id, user.full_name if user else None, user.email if user else None,
                t.subject, support_status_label(t.status), t.admin_reply, str(t.created_at), str(t.updated_at)
            ])

    elif type == "scorecard":
        survey_query = db.query(models.ScorecardSurvey)
        if event_id:
            survey_query = survey_query.filter(models.ScorecardSurvey.event_id == event_id)
        surveys = survey_query.order_by(models.ScorecardSurvey.created_at.desc()).all()
        headers = [
            "Event ID", "Event Title", "Survey ID", "Survey Type", "Participant ID",
            "Full Name", "Anonymous Code", "Phone Number", "County", "Community/Location",
            "User Type", "Participation Type", "Consent", "Coexistence Index",
            "Baseline Score", "Follow-up Score", "Percentage Point Change",
            "Category Scores", "Responses JSON",
            "Community members engaged", "Trainings/story labs conducted",
            "Animals indirectly benefiting", "Materials/tools produced",
            "Human wellbeing outcome notes", "Animal welfare outcome notes",
            "Environmental benefit notes", "Social cohesion notes",
            "Evidence links or uploaded files",
        ]
        ws.append(headers)
        for cell in ws[1]: cell.font = header_font

        for survey in surveys:
            event = db.query(models.Event).filter(models.Event.id == survey.event_id).first()
            participant = db.query(models.ScorecardParticipant).filter(models.ScorecardParticipant.id == survey.participant_id).first()
            baseline_score, followup_score, change = participant_score_pair(db, survey.event_id, survey.participant_id)
            report = db.query(models.ScorecardReportingExport).filter(
                models.ScorecardReportingExport.event_id == survey.event_id
            ).order_by(models.ScorecardReportingExport.updated_at.desc()).first()
            reporting_fields = {**DEFAULT_REPORTING_FIELDS, **((report.fields or {}) if report else {})}
            evidence_links = "; ".join([
                e.url for e in db.query(models.ScorecardEvidence).filter(models.ScorecardEvidence.event_id == survey.event_id).all()
            ])
            response_payload = [
                {
                    "question": response.question.prompt if response.question else None,
                    "category": response.question.category if response.question else None,
                    "type": response.question.question_type if response.question else None,
                    "answer_numeric": response.answer_numeric,
                    "answer_text": response.answer_text,
                }
                for response in survey.responses
            ]
            ws.append([
                survey.event_id,
                event.title if event else None,
                survey.id,
                survey.survey_type,
                participant.id if participant else None,
                participant.full_name if participant else None,
                participant.anonymous_code if participant else None,
                participant.phone_number if participant else None,
                participant.county if participant else None,
                participant.community_location if participant else None,
                participant.user_type if participant else None,
                participant.participation_type if participant else None,
                "Yes" if participant and participant.consent else "No",
                survey.coexistence_index,
                baseline_score,
                followup_score,
                change,
                json.dumps(survey.category_scores or {}),
                json.dumps(response_payload),
                reporting_fields.get("community_members_engaged"),
                reporting_fields.get("trainings_story_labs_conducted"),
                reporting_fields.get("animals_indirectly_benefiting"),
                reporting_fields.get("materials_tools_produced"),
                reporting_fields.get("human_wellbeing_outcome_notes"),
                reporting_fields.get("animal_welfare_outcome_notes"),
                reporting_fields.get("environmental_benefit_notes"),
                reporting_fields.get("social_cohesion_notes"),
                reporting_fields.get("evidence_links_or_uploaded_files") or evidence_links,
            ])

            
    # Save to buffer
    stream = io.BytesIO()
    wb.save(stream)
    stream.seek(0)
            
    response = StreamingResponse(iter([stream.getvalue()]), media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    response.headers["Content-Disposition"] = f"attachment; filename=export_{type}_{datetime.utcnow().timestamp()}.xlsx"
    return response


# --- Lovedogs 360 Event Specific API ---

@app.get("/events/{event_id}/journey", response_model=schemas.ProgramJourneyResponse)
def get_program_journey(
    event_id: str,
    db: Session = Depends(database.get_db),
    current_user: models.User = Depends(get_current_user)
):
    journey = db.query(models.ProgramJourney).filter(
        models.ProgramJourney.event_id == event_id,
        models.ProgramJourney.user_id == current_user.id
    ).first()
    
    if not journey:
        # Create a baseline journey if none exists (Auto-start T1)
        # In a real app we might only do this after they register, but for UX ease:
        journey = models.ProgramJourney(
            id=str(uuid.uuid4()),
            event_id=event_id,
            user_id=current_user.id,
            progress_percentage=0.0,
            current_timepoint="T1"
        )
        db.add(journey)
        db.commit()
        db.refresh(journey)

    return journey

@app.post("/events/{event_id}/sync")
def bulk_sync_event_data(
    event_id: str,
    payload: schemas.BulkSyncPayload,
    db: Session = Depends(database.get_db),
    current_user: models.User = Depends(get_current_user)
):
    added_checkins = 0
    added_observations = 0

    # Sync Check-ins 
    for c in payload.checkins:
        # Prevent duplicates based on timepoint per user per event
        existing = db.query(models.CheckInData).filter(
            models.CheckInData.event_id == event_id,
            models.CheckInData.user_id == c.user_id,
            models.CheckInData.timepoint == c.timepoint
        ).first()

        if not existing:
            new_checkin = models.CheckInData(
                id=str(uuid.uuid4()),
                event_id=event_id,
                user_id=c.user_id,
                dog_id=c.dog_id,
                timepoint=c.timepoint,
                who5_answers=c.who5_answers,
                pss10_answers=c.pss10_answers,
                relationship_answers=c.relationship_answers,
                welfare_snapshot=c.welfare_snapshot
            )
            db.add(new_checkin)
            added_checkins += 1

            # Update Journey Progress automatically
            journey = db.query(models.ProgramJourney).filter(
                models.ProgramJourney.event_id == event_id,
                models.ProgramJourney.user_id == c.user_id
            ).first()
            if journey:
                # Basic Logic for progress bumping
                if c.timepoint == "T1": 
                    journey.current_timepoint = "T2"
                    journey.progress_percentage = max(journey.progress_percentage, 25.0)
                elif c.timepoint == "T2":
                    journey.current_timepoint = "T3"
                    journey.progress_percentage = max(journey.progress_percentage, 50.0)
                elif c.timepoint == "T3":
                    journey.current_timepoint = "T4"
                    journey.progress_percentage = max(journey.progress_percentage, 75.0)
                elif c.timepoint == "T4":
                    journey.progress_percentage = 100.0

    # Sync Observations (Facilitators / Vets)
    for obs in payload.observations:
        new_obs = models.LiveObservation(
            id=str(uuid.uuid4()),
            event_id=event_id,
            observer_id=current_user.id,
            participant_id=obs.participant_id,
            dog_id=obs.dog_id,
            behavior=obs.behavior,
            intensity=obs.intensity,
            notes=obs.notes,
            timestamp=obs.timestamp.replace(tzinfo=None) if obs.timestamp else datetime.utcnow(),
            is_offline_sync=obs.is_offline_sync,
            synced_at=datetime.utcnow()
        )
        db.add(new_obs)
        added_observations += 1

    db.commit()
    return {"message": "Sync successful", "checkins_synced": added_checkins, "observations_synced": added_observations}

@app.post("/events/{event_id}/live-log", response_model=schemas.LiveObservationResponse)
def live_log_observation(
    event_id: str,
    observation: schemas.LiveObservationCreate,
    db: Session = Depends(database.get_db),
    current_user: models.User = Depends(get_current_user)
):
    if current_user.role not in [models.UserRole.ADMIN, models.UserRole.PROVIDER]:
        raise HTTPException(status_code=403, detail="Not authorized to log observations")

    new_obs = models.LiveObservation(
        id=str(uuid.uuid4()),
        event_id=event_id,
        observer_id=current_user.id,
        participant_id=observation.participant_id,
        dog_id=observation.dog_id,
        behavior=observation.behavior,
        intensity=observation.intensity,
        notes=observation.notes,
        timestamp=observation.timestamp.replace(tzinfo=None) if observation.timestamp else datetime.utcnow(),
        is_offline_sync=False,
        synced_at=datetime.utcnow()
    )
    db.add(new_obs)
    db.commit()
    db.refresh(new_obs)
    return new_obs


# --- Admin Dashboard API ---

def normalize_app_platform(platform: Optional[str]) -> str:
    normalized = (platform or "all").strip().lower()
    if normalized not in {"android", "ios", "all"}:
        raise HTTPException(status_code=400, detail="Platform must be android, ios, or all")
    return normalized

@app.get("/app/version/latest", response_model=Optional[schemas.AppVersionResponse])
def get_latest_app_version(
    platform: Optional[str] = None,
    db: Session = Depends(database.get_db)
):
    normalized_platform = normalize_app_platform(platform)
    eligible_platforms = ["all"] if normalized_platform == "all" else [normalized_platform, "all"]
    return (
        db.query(models.AppVersion)
        .filter(
            models.AppVersion.is_active == True,
            models.AppVersion.platform.in_(eligible_platforms),
        )
        .order_by(models.AppVersion.created_at.desc())
        .first()
    )

@app.post("/app/version", response_model=schemas.AppVersionResponse)
def create_app_version(
    version: schemas.AppVersionCreate,
    db: Session = Depends(database.get_db),
    admin: models.User = Depends(require_admin)
):
    app_version = models.AppVersion(
        id=str(uuid.uuid4()),
        version=version.version.strip(),
        platform=normalize_app_platform(version.platform),
        release_notes=version.release_notes,
        download_url=version.download_url,
        is_required=version.is_required,
    )
    db.add(app_version)
    db.commit()
    db.refresh(app_version)
    return app_version

@app.put("/app/version/{version_id}", response_model=schemas.AppVersionResponse)
def update_app_version(
    version_id: str,
    version_update: schemas.AppVersionUpdate,
    db: Session = Depends(database.get_db),
    admin: models.User = Depends(require_admin)
):
    app_version = db.query(models.AppVersion).filter(models.AppVersion.id == version_id).first()
    if not app_version:
        raise HTTPException(status_code=404, detail="App version not found")

    if version_update.release_notes is not None:
        app_version.release_notes = version_update.release_notes
    if version_update.download_url is not None:
        app_version.download_url = version_update.download_url
    if version_update.is_required is not None:
        app_version.is_required = version_update.is_required
    if version_update.is_active is not None:
        app_version.is_active = version_update.is_active
    app_version.updated_at = datetime.utcnow()

    db.commit()
    db.refresh(app_version)
    return app_version

@app.get("/admin/stats")
def admin_stats(db: Session = Depends(database.get_db), admin: models.User = Depends(require_admin)):
    total_users = db.query(models.User).count()
    total_services = db.query(models.Service).count()
    total_orders = db.query(models.Order).count()
    total_events = db.query(models.Event).count()
    
    from sqlalchemy import func
    total_revenue = db.query(func.coalesce(func.sum(models.Order.amount), 0)).filter(
        models.Order.status.in_(PAID_ORDER_STATUS_VALUES)
    ).scalar()
    total_commission = db.query(func.coalesce(func.sum(models.Order.commission), 0)).filter(
        models.Order.status.in_(PAID_ORDER_STATUS_VALUES)
    ).scalar()
    
    return {
        "total_users": total_users,
        "total_services": total_services,
        "total_orders": total_orders,
        "total_events": total_events,
        "total_revenue": round(float(total_revenue), 2),
        "total_commission": round(float(total_commission), 2)
    }

@app.get("/admin/users")
def admin_list_users(db: Session = Depends(database.get_db), admin: models.User = Depends(require_admin)):
    users = db.query(models.User).order_by(models.User.created_at.desc()).all()
    result = []
    for u in users:
        dog_count = db.query(models.Dog).filter(models.Dog.owner_id == u.id).count()
        listing_count = db.query(models.Service).filter(models.Service.provider_id == u.id).count()
        order_count = db.query(models.Order).filter(models.Order.buyer_id == u.id).count()
        paid_order_count = db.query(models.Order).filter(
            models.Order.buyer_id == u.id,
            models.Order.status.in_(PAID_ORDER_STATUS_VALUES),
        ).count()
        result.append({
            "id": u.id,
            "full_name": u.full_name,
            "email": u.email,
            "phone_number": u.phone_number,
            "role": u.role,
            "bio": u.bio,
            "country": u.country,
            "preferred_currency": u.preferred_currency,
            "average_rating": u.average_rating or 0,
            "total_ratings": u.total_ratings or 0,
            "dog_count": dog_count,
            "listing_count": listing_count,
            "order_count": order_count,
            "paid_order_count": paid_order_count,
            "created_at": str(u.created_at) if u.created_at else None,
        })
    return result

@app.get("/admin/orders")
def admin_list_orders(db: Session = Depends(database.get_db), admin: models.User = Depends(require_admin)):
    orders = db.query(models.Order).order_by(models.Order.created_at.desc()).all()
    result = []
    for order in orders:
        service = db.query(models.Service).filter(models.Service.id == order.service_id).first()
        buyer = db.query(models.User).filter(models.User.id == order.buyer_id).first()
        provider = None
        if service:
            provider = db.query(models.User).filter(models.User.id == service.provider_id).first()
        status_value = order_status_value(order.status)
        is_paid = status_value in PAID_ORDER_STATES
        amount = float(order.amount or 0)
        commission = float(order.commission or 0)
        payout = float(order.payout or 0)
        discount_amount = float(getattr(order, "discount_amount", 0) or 0)
        karma_points_redeemed = int(getattr(order, "karma_points_redeemed", 0) or 0)
        result.append({
            "id": order.id,
            "buyer_name": buyer.full_name if buyer else "Unknown",
            "buyer_email": buyer.email if buyer else "",
            "buyer_phone": buyer.phone_number if buyer else None,
            "buyer_id": order.buyer_id,
            "provider_name": provider.full_name if provider else "Unknown",
            "provider_id": service.provider_id if service else None,
            "service_title": service.title if service else "Unknown",
            "service_id": order.service_id,
            "item_type": service.item_type if service else None,
            "amount": amount,
            "commission": commission,
            "payout": payout,
            "discount_amount": discount_amount,
            "karma_points_redeemed": karma_points_redeemed,
            "paid_amount": amount if is_paid else 0,
            "paid_commission": commission if is_paid else 0,
            "paid_payout": payout if is_paid else 0,
            "status": status_value,
            "is_paid": is_paid,
            "share_phone": order.share_phone,
            "service_stock_count": service.stock_count if service else None,
            "service_slots_available": service.slots_available if service else None,
            "form_responses": [
                {
                    "label": r.field.label if r.field else "Question",
                    "answer": r.answer_value
                } for r in (order.responses or [])
            ],
            "created_at": str(order.created_at) if order.created_at else None
        })
    return result

@app.post("/admin/orders/{order_id}/complete")
def admin_complete_order(
    order_id: str,
    db: Session = Depends(database.get_db),
    admin: models.User = Depends(require_admin)
):
    """Admin confirms service delivery — transitions order from PAID to COMPLETED."""
    order = db.query(models.Order).filter(models.Order.id == order_id).first()
    if not order:
        raise HTTPException(status_code=404, detail="Order not found")
    current_status = order_status_value(order.status)
    if current_status in {models.OrderStatus.COMPLETED.value, models.OrderStatus.SETTLED.value}:
        return {
            "message": "Delivery was already confirmed for this order.",
            "status": order.status
        }
    if current_status != models.OrderStatus.PAID.value:
        raise HTTPException(
            status_code=400,
            detail=f"Order must be in 'paid' status to mark as completed. Current status: {order.status}"
        )

    order.status = models.OrderStatus.COMPLETED.value
    service = db.query(models.Service).filter(models.Service.id == order.service_id).first()
    item_title = service.title if service else "Marketplace item"
    add_notification(
        db,
        order.buyer_id,
        "Delivery confirmed",
        f"Delivery has been confirmed for '{item_title}'. You can now leave a rating.",
        "delivery",
        commit=False
    )
    if service and service.provider_id:
        add_notification(
            db,
            service.provider_id,
            "Delivery confirmed",
            f"Delivery for '{item_title}' has been confirmed. Your payout is now available for withdrawal.",
            "delivery",
            commit=False
        )
    db.commit()
    return {"message": "Order marked as completed. Seller payout is now available for withdrawal.", "status": order.status}

@app.post("/admin/orders/{order_id}/settle")
def admin_settle_order(
    order_id: str,
    db: Session = Depends(database.get_db),
    admin: models.User = Depends(require_admin)
):
    """Admin approves seller payout — transitions order from COMPLETED to SETTLED and credits the seller's wallet."""
    order = db.query(models.Order).filter(models.Order.id == order_id).first()
    if not order:
        raise HTTPException(status_code=404, detail="Order not found")
    if order_status_value(order.status) != models.OrderStatus.COMPLETED.value:
        raise HTTPException(
            status_code=400,
            detail=f"Order must be in 'completed' status to settle. Current status: {order.status}"
        )

    # Find the service provider
    service = db.query(models.Service).filter(models.Service.id == order.service_id).first()
    if not service:
        raise HTTPException(status_code=404, detail="Service not found for this order")

    provider_id = service.provider_id
    provider = db.query(models.User).filter(models.User.id == provider_id).first()
    payout_amount = order.payout or 0

    # Credit the provider's wallet (create wallet if it doesn't exist)
    wallet = db.query(models.Transaction).filter(
        models.Transaction.order_id == order_id,
        models.Transaction.type == "payout"
    ).first()

    if wallet:
        raise HTTPException(status_code=400, detail="Payout transaction already exists for this order")

    # Create payout transaction record
    tx = models.Transaction(
        id=str(uuid.uuid4()),
        order_id=order_id,
        user_id=provider_id,
        amount=payout_amount,
        type="payout",
        status="completed",
        payout_method=getattr(provider, "payment_method", None),
        destination=get_payout_destination(provider) if provider else None
    )
    db.add(tx)

    # Update order status to SETTLED
    order.status = models.OrderStatus.SETTLED.value
    db.commit()

    # Send notification to provider
    try:
        destination = get_payout_destination(provider) if provider else None
        destination_text = f" to {destination}" if destination else ""
        notification = models.Notification(
            id=str(uuid.uuid4()),
            user_id=provider_id,
            title="Payout Approved! 💰",
            message=f"Your payout of KES {payout_amount:,.2f} for '{service.title}' has been approved and settled{destination_text}.",
            type="payout"
        )
        db.add(notification)
        db.commit()
    except Exception:
        pass  # Non-critical — don't fail the settlement

    return {
        "message": f"Payout of KES {payout_amount:,.2f} approved and settled for provider.",
        "status": order.status,
        "payout_amount": payout_amount
    }

@app.get("/admin/services")
def admin_list_services(db: Session = Depends(database.get_db), admin: models.User = Depends(require_admin)):
    from sqlalchemy import func
    services = db.query(models.Service).order_by(models.Service.title.asc()).all()
    result = []
    for s in services:
        provider = db.query(models.User).filter(models.User.id == s.provider_id).first()
        paid_order_count = db.query(models.Order).filter(
            models.Order.service_id == s.id,
            models.Order.status.in_(PAID_ORDER_STATUS_VALUES),
        ).count()
        pending_order_count = db.query(models.Order).filter(
            models.Order.service_id == s.id,
            models.Order.status.in_(PENDING_ORDER_STATUS_VALUES),
        ).count()
        paid_revenue = db.query(func.coalesce(func.sum(models.Order.amount), 0)).filter(
            models.Order.service_id == s.id,
            models.Order.status.in_(PAID_ORDER_STATUS_VALUES),
        ).scalar()
        result.append({
            "id": s.id,
            "title": s.title,
            "description": s.description,
            "price": s.price,
            "category": s.category,
            "item_type": s.item_type,
            "is_published": s.is_published,
            "admin_approved": s.admin_approved,
            "rejection_reason": s.rejection_reason,
            "stock_count": s.stock_count,
            "slots_available": s.slots_available,
            "is_busy": s.is_busy,
            "provider_name": provider.full_name if provider else "Unknown",
            "provider_id": s.provider_id,
            "provider_email": provider.email if provider else None,
            "image_url": s.image_url,
            "paid_order_count": paid_order_count,
            "pending_order_count": pending_order_count,
            "paid_revenue": round(float(paid_revenue or 0), 2),
        })
    return result


class ApprovalRequest(BaseModel):
    is_approved: bool
    rejection_reason: Optional[str] = None

@app.get("/admin/pending-approvals")
def admin_list_pending(db: Session = Depends(database.get_db), admin: models.User = Depends(require_admin)):
    pending_services = db.query(models.Service).filter(
        models.Service.admin_approved == False,
        models.Service.rejection_reason.is_(None),
    ).order_by(models.Service.title.asc()).all()
    pending_reports = db.query(models.CaseReport).filter(
        models.CaseReport.is_approved == False,
        models.CaseReport.rejection_reason.is_(None),
    ).order_by(models.CaseReport.created_at.desc()).all()
    
    services_out = []
    for s in pending_services:
        provider = db.query(models.User).filter(models.User.id == s.provider_id).first()
        services_out.append({
            "id": s.id, "title": s.title, "description": s.description, "price": s.price,
            "category": s.category, "item_type": s.item_type,
            "stock_count": s.stock_count, "slots_available": s.slots_available,
            "is_published": s.is_published, "admin_approved": s.admin_approved,
            "provider_name": provider.full_name if provider else "Unknown",
            "provider_id": s.provider_id,
            "provider_email": provider.email if provider else None,
        })
    
    reports_out = []
    for r in pending_reports:
        author = db.query(models.User).filter(models.User.id == r.author_id).first()
        reports_out.append({
            "id": r.id, "title": r.title, "description": r.description,
            "case_type": r.case_type, "location": r.location, "status": r.status,
            "author_name": author.full_name if author else "Unknown",
            "author_id": r.author_id,
            "created_at": str(r.created_at)
        })
        
    return {
        "pending_services": services_out,
        "pending_reports": reports_out
    }

@app.post("/admin/approve/{item_type}/{item_id}")
def admin_approve_item(
    item_type: str, 
    item_id: str, 
    req: ApprovalRequest,
    db: Session = Depends(database.get_db), 
    admin: models.User = Depends(require_admin)
):
    if item_type == "service":
        item = db.query(models.Service).filter(models.Service.id == item_id).first()
        if not item: raise HTTPException(status_code=404, detail="Service not found")
        item.admin_approved = req.is_approved
        item.rejection_reason = req.rejection_reason
        if req.is_approved:
            item.is_published = True
            item.rejection_reason = None
        else:
            item.is_published = False
    elif item_type == "report":
        item = db.query(models.CaseReport).filter(models.CaseReport.id == item_id).first()
        if not item: raise HTTPException(status_code=404, detail="Report not found")
        item.is_approved = req.is_approved
        item.rejection_reason = req.rejection_reason
    else:
        raise HTTPException(status_code=400, detail="Invalid item type")

    if not req.is_approved and not item.rejection_reason:
        item.rejection_reason = "Rejected by admin"
        
    db.commit()
    
    # Notify User
    if req.is_approved:
        title = f"{item_type.capitalize()} Approved"
        msg = f"Your {item_type} '{item.title}' has been approved and is now visible to the public."
        notif_type = "approval"
    else:
        title = f"{item_type.capitalize()} Rejected"
        msg = f"Your {item_type} '{item.title}' was rejected."
        if req.rejection_reason:
            msg += f" Reason: {req.rejection_reason}"
        notif_type = "rejection"
    
    recipient_id = item.provider_id if item_type == "service" else item.author_id
    create_notification(db, recipient_id, title, msg, notif_type)

    return {"message": "Success", "is_approved": req.is_approved}

@app.get("/spotlight", response_model=List[schemas.SpotlightResponse])
def get_spotlight(db: Session = Depends(database.get_db)):
    # Return active content pins first, followed by legacy spotlight items.
    pins = db.query(models.ContentPin).filter(*active_pin_filter()).order_by(
        models.ContentPin.priority.desc(),
        models.ContentPin.updated_at.desc(),
    ).all()
    spotlights = db.query(models.Spotlight).filter(models.Spotlight.is_active == True).order_by(models.Spotlight.updated_at.desc()).all()
    return [content_pin_to_spotlight(pin) for pin in pins] + spotlights

@app.get("/admin/spotlight", response_model=List[schemas.SpotlightResponse])
def get_admin_spotlight(db: Session = Depends(database.get_db), admin: models.User = Depends(require_admin)):
    spotlights = db.query(models.Spotlight).order_by(models.Spotlight.updated_at.desc()).all()
    return spotlights

@app.post("/admin/spotlight", response_model=schemas.SpotlightResponse)
def create_admin_spotlight(
    spotlight_in: schemas.SpotlightBase,
    db: Session = Depends(database.get_db),
    admin: models.User = Depends(require_admin)
):
    spotlight = models.Spotlight(
        title=spotlight_in.title,
        description=spotlight_in.description,
        image_url=spotlight_in.image_url,
        target_route=spotlight_in.target_route,
        target_id=spotlight_in.target_id,
        is_active=spotlight_in.is_active
    )
    db.add(spotlight)
    db.commit()
    db.refresh(spotlight)
    return spotlight

@app.delete("/admin/spotlight/{spotlight_id}")
def delete_admin_spotlight(
    spotlight_id: int,
    db: Session = Depends(database.get_db),
    admin: models.User = Depends(require_admin)
):
    spotlight = db.query(models.Spotlight).filter(models.Spotlight.id == spotlight_id).first()
    if not spotlight:
        raise HTTPException(status_code=404, detail="Spotlight not found")
    db.delete(spotlight)
    db.commit()
    return {"message": "Spotlight removed"}


@app.get("/admin/pins", response_model=List[schemas.ContentPinResponse])
def admin_list_pins(db: Session = Depends(database.get_db), admin: models.User = Depends(require_admin)):
    return db.query(models.ContentPin).order_by(
        models.ContentPin.is_active.desc(),
        models.ContentPin.priority.desc(),
        models.ContentPin.updated_at.desc(),
    ).all()


@app.post("/admin/pins", response_model=schemas.ContentPinResponse)
def admin_pin_content(
    pin_in: schemas.ContentPinCreate,
    db: Session = Depends(database.get_db),
    admin: models.User = Depends(require_admin),
):
    return ensure_content_pin(
        db,
        pin_in.target_type,
        pin_in.target_id,
        admin,
        title=pin_in.title,
        description=pin_in.description,
        image_url=pin_in.image_url,
        priority=pin_in.priority or 100,
        expires_at=pin_in.expires_at,
    )


@app.delete("/admin/pins/{target_type}/{target_id}")
def admin_unpin_content(
    target_type: str,
    target_id: str,
    db: Session = Depends(database.get_db),
    admin: models.User = Depends(require_admin),
):
    pin = deactivate_content_pin(db, target_type, target_id)
    if not pin:
        return {"message": "Content was not pinned", "is_pinned": False}
    return {"message": "Content unpinned", "is_pinned": False}


@app.get("/admin/pinnable-content")
def admin_pinnable_content(db: Session = Depends(database.get_db), admin: models.User = Depends(require_admin)):
    def pin_meta(target_type, item_id):
        pin = get_active_pin_map(db, target_type).get(str(item_id))
        return {"is_pinned": pin is not None, "pin_priority": pin.priority if pin else None}

    events = db.query(models.Event).order_by(models.Event.start_time.desc()).limit(50).all()
    services = db.query(models.Service).filter(
        models.Service.is_published == True,
        models.Service.admin_approved == True,
    ).order_by(models.Service.title.asc()).limit(50).all()
    cases = db.query(models.CaseReport).filter(
        models.CaseReport.is_approved == True,
    ).order_by(models.CaseReport.created_at.desc()).limit(50).all()
    community = db.query(models.CommunityMessage).filter(
        models.CommunityMessage.is_hidden == False,
    ).order_by(models.CommunityMessage.created_at.desc()).limit(50).all()

    return {
        "events": [
            {
                "id": item.id,
                "title": item.title,
                "description": item.description,
                "meta": item.location,
                "created_at": str(item.created_at),
                **pin_meta("event", item.id),
            }
            for item in events
        ],
        "services": [
            {
                "id": item.id,
                "title": item.title,
                "description": item.description,
                "meta": f"{item.item_type} | {item.category}",
                "created_at": str(item.provider_id),
                **pin_meta("service", item.id),
            }
            for item in services
        ],
        "cases": [
            {
                "id": item.id,
                "title": item.title,
                "description": item.description,
                "meta": item.case_type,
                "created_at": str(item.created_at),
                **pin_meta("case", item.id),
            }
            for item in cases
        ],
        "community": [
            {
                "id": item.id,
                "title": (item.content or "Community post")[:80],
                "description": item.content,
                "meta": "community post",
                "created_at": str(item.created_at),
                **pin_meta("community", item.id),
            }
            for item in community
        ],
    }

# =====================================================
# Admin Analytics & Enhanced Management Endpoints
# =====================================================

@app.get("/admin/analytics")
def admin_analytics(db: Session = Depends(database.get_db), admin: models.User = Depends(require_admin)):
    from sqlalchemy import func
    now = datetime.utcnow()
    thirty_days_ago = now - timedelta(days=30)
    sixty_days_ago = now - timedelta(days=60)

    # --- Growth metrics (current 30d vs previous 30d) ---
    new_users_30d = db.query(models.User).filter(models.User.created_at >= thirty_days_ago).count()
    new_users_prev_30d = db.query(models.User).filter(
        models.User.created_at >= sixty_days_ago, models.User.created_at < thirty_days_ago
    ).count()
    new_orders_30d = db.query(models.Order).filter(models.Order.created_at >= thirty_days_ago).count()
    new_orders_prev_30d = db.query(models.Order).filter(
        models.Order.created_at >= sixty_days_ago, models.Order.created_at < thirty_days_ago
    ).count()
    new_paid_orders_30d = db.query(models.Order).filter(
        models.Order.status.in_(PAID_ORDER_STATUS_VALUES),
        models.Order.created_at >= thirty_days_ago,
    ).count()
    new_paid_orders_prev_30d = db.query(models.Order).filter(
        models.Order.status.in_(PAID_ORDER_STATUS_VALUES),
        models.Order.created_at >= sixty_days_ago,
        models.Order.created_at < thirty_days_ago,
    ).count()

    revenue_30d = float(db.query(func.coalesce(func.sum(models.Order.amount), 0)).filter(
        models.Order.status.in_(PAID_ORDER_STATUS_VALUES), models.Order.created_at >= thirty_days_ago
    ).scalar())
    revenue_prev_30d = float(db.query(func.coalesce(func.sum(models.Order.amount), 0)).filter(
        models.Order.status.in_(PAID_ORDER_STATUS_VALUES),
        models.Order.created_at >= sixty_days_ago, models.Order.created_at < thirty_days_ago
    ).scalar())

    # --- Users by role ---
    role_counts = db.query(models.User.role, func.count(models.User.id)).group_by(models.User.role).all()
    users_by_role = {role: count for role, count in role_counts}

    # --- Orders by status ---
    status_counts = db.query(models.Order.status, func.count(models.Order.id)).group_by(models.Order.status).all()
    orders_by_status = {}
    for status, count in status_counts:
        key = order_status_value(status)
        orders_by_status[key] = orders_by_status.get(key, 0) + count

    # --- Top services by paid order count/revenue ---
    top_services_q = db.query(
        models.Service.title,
        func.count(models.Order.id).label("order_count"),
        func.coalesce(func.sum(models.Order.amount), 0).label("revenue")
    ).join(models.Order, models.Order.service_id == models.Service.id).filter(
        models.Order.status.in_(PAID_ORDER_STATUS_VALUES)
    ).group_by(
        models.Service.id, models.Service.title
    ).order_by(func.count(models.Order.id).desc()).limit(5).all()
    top_services = [{"title": t, "order_count": c, "revenue": round(float(r), 2)} for t, c, r in top_services_q]

    # --- Recent activity feed ---
    recent_activity = []
    recent_users = db.query(models.User).order_by(models.User.created_at.desc()).limit(5).all()
    for u in recent_users:
        recent_activity.append({
            "type": "registration", "icon": "person-add",
            "description": f"{u.full_name} joined as {u.role}",
            "time": str(u.created_at)
        })
    recent_orders = db.query(models.Order).order_by(models.Order.created_at.desc()).limit(5).all()
    for o in recent_orders:
        svc = db.query(models.Service.title).filter(models.Service.id == o.service_id).scalar()
        status_label = order_status_value(o.status).replace("_", " ").title()
        recent_activity.append({
            "type": "order", "icon": "cart",
            "description": f"{status_label} order for {svc or 'Unknown'}",
            "time": str(o.created_at)
        })
    recent_cases = db.query(models.CaseReport).order_by(models.CaseReport.created_at.desc()).limit(3).all()
    for c in recent_cases:
        recent_activity.append({
            "type": "case", "icon": "alert-circle",
            "description": f"Case reported: {c.title}",
            "time": str(c.created_at)
        })
    recent_activity.sort(key=lambda x: x["time"], reverse=True)
    recent_activity = recent_activity[:10]

    # --- Community stats ---
    total_cases = db.query(models.CaseReport).count()
    open_cases = db.query(models.CaseReport).filter(models.CaseReport.status == "open").count()
    total_dogs = db.query(models.Dog).count()
    total_community_posts = db.query(models.CommunityMessage).count()
    flagged_posts = db.query(models.CommunityMessage).filter(models.CommunityMessage.flag_count > 0).count()

    # --- Support stats ---
    all_ticket_statuses = db.query(models.SupportTicket.status).all()
    open_tickets = sum(1 for (ticket_status,) in all_ticket_statuses if support_status_key(ticket_status) != "resolved")
    total_tickets = len(all_ticket_statuses)

    # --- Pending approvals count ---
    pending_services = db.query(models.Service).filter(
        models.Service.admin_approved == False,
        models.Service.rejection_reason.is_(None),
    ).count()
    pending_reports = db.query(models.CaseReport).filter(
        models.CaseReport.is_approved == False,
        models.CaseReport.rejection_reason.is_(None),
    ).count()

    # --- Monthly revenue (last 6 months) ---
    monthly_revenue = []
    month_names = ["Jan", "Feb", "Mar", "Apr", "May", "Jun", "Jul", "Aug", "Sep", "Oct", "Nov", "Dec"]
    for i in range(5, -1, -1):
        month_index = (now.year * 12 + now.month - 1) - i
        month_year = month_index // 12
        month_number = month_index % 12 + 1
        month_start = now.replace(year=month_year, month=month_number, day=1, hour=0, minute=0, second=0, microsecond=0)
        if month_start.month == 12:
            month_end = month_start.replace(year=month_start.year + 1, month=1)
        else:
            month_end = month_start.replace(month=month_start.month + 1)
        rev = float(db.query(func.coalesce(func.sum(models.Order.amount), 0)).filter(
            models.Order.status.in_(PAID_ORDER_STATUS_VALUES),
            models.Order.created_at >= month_start, models.Order.created_at < month_end
        ).scalar())
        comm = float(db.query(func.coalesce(func.sum(models.Order.commission), 0)).filter(
            models.Order.status.in_(PAID_ORDER_STATUS_VALUES),
            models.Order.created_at >= month_start, models.Order.created_at < month_end
        ).scalar())
        monthly_revenue.append({
            "month": month_names[month_start.month - 1],
            "revenue": round(rev, 2),
            "commission": round(comm, 2)
        })

    # --- Event stats ---
    total_events = db.query(models.Event).count()
    upcoming_events = db.query(models.Event).filter(models.Event.start_time > now).count()
    total_registrations = db.query(models.Registration).count()
    scorecard_participants = db.query(models.ScorecardParticipant).count()
    scorecard_baselines = db.query(models.ScorecardSurvey).filter(models.ScorecardSurvey.survey_type == "baseline").count()
    scorecard_followups = db.query(models.ScorecardSurvey).filter(models.ScorecardSurvey.survey_type == "followup").count()

    return {
        "total_users": db.query(models.User).count(),
        "total_services": db.query(models.Service).count(),
        "total_orders": db.query(models.Order).count(),
        "total_paid_orders": db.query(models.Order).filter(models.Order.status.in_(PAID_ORDER_STATUS_VALUES)).count(),
        "pending_orders": db.query(models.Order).filter(models.Order.status.in_(PENDING_ORDER_STATUS_VALUES)).count(),
        "total_events": total_events,
        "total_revenue": round(float(db.query(func.coalesce(func.sum(models.Order.amount), 0)).filter(
            models.Order.status.in_(PAID_ORDER_STATUS_VALUES)).scalar()), 2),
        "total_commission": round(float(db.query(func.coalesce(func.sum(models.Order.commission), 0)).filter(
            models.Order.status.in_(PAID_ORDER_STATUS_VALUES)).scalar()), 2),
        "new_users_30d": new_users_30d,
        "new_users_prev_30d": new_users_prev_30d,
        "new_orders_30d": new_orders_30d,
        "new_orders_prev_30d": new_orders_prev_30d,
        "new_paid_orders_30d": new_paid_orders_30d,
        "new_paid_orders_prev_30d": new_paid_orders_prev_30d,
        "revenue_30d": round(revenue_30d, 2),
        "revenue_prev_30d": round(revenue_prev_30d, 2),
        "users_by_role": users_by_role,
        "orders_by_status": orders_by_status,
        "top_services": top_services,
        "recent_activity": recent_activity,
        "total_cases": total_cases,
        "open_cases": open_cases,
        "total_dogs": total_dogs,
        "total_community_posts": total_community_posts,
        "flagged_posts": flagged_posts,
        "open_tickets": open_tickets,
        "total_tickets": total_tickets,
        "pending_services": pending_services,
        "pending_reports": pending_reports,
        "monthly_revenue": monthly_revenue,
        "upcoming_events": upcoming_events,
        "total_registrations": total_registrations,
        "scorecard_participants": scorecard_participants,
        "scorecard_baselines": scorecard_baselines,
        "scorecard_followups": scorecard_followups,
    }


@app.get("/admin/events")
def admin_list_events(db: Session = Depends(database.get_db), admin: models.User = Depends(require_admin)):
    events = db.query(models.Event).order_by(models.Event.start_time.desc()).all()
    pin_map = get_active_pin_map(db, "event")
    result = []
    for e in events:
        organizer = db.query(models.User).filter(models.User.id == e.organizer_id).first()
        all_reg_count = db.query(models.Registration).filter(models.Registration.event_id == e.id).count()
        reg_count = db.query(models.Registration).filter(
            models.Registration.event_id == e.id,
            models.Registration.status.in_(["registered", "checked-in"]),
        ).count()
        paid_count = db.query(models.Registration).filter(
            models.Registration.event_id == e.id,
            models.Registration.payment_status == "paid",
        ).count()
        pending_payment_count = db.query(models.Registration).filter(
            models.Registration.event_id == e.id,
            models.Registration.payment_status == "pending",
        ).count()
        event_revenue = db.query(func.coalesce(func.sum(models.Registration.amount), 0)).filter(
            models.Registration.event_id == e.id,
            models.Registration.payment_status == "paid",
        ).scalar()
        checkin_count = db.query(models.Registration).filter(
            models.Registration.event_id == e.id, models.Registration.status == "checked-in"
        ).count()
        pin = pin_map.get(e.id)
        result.append({
            "id": e.id, "title": e.title, "description": e.description,
            "poster_url": e.poster_url, "images": e.images or [],
            "location": e.location, "start_time": str(e.start_time), "end_time": str(e.end_time),
            "capacity": e.capacity, "ticket_price": e.ticket_price, "currency": e.currency,
            "ticket_tiers": normalize_event_ticket_tiers(e),
            "attendee_type_question": e.attendee_type_question,
            "category": e.category, "is_public": e.is_public,
            "admin_created": e.admin_created, "scorecard_enabled": e.scorecard_enabled,
            "follow_up_requested_at": str(e.follow_up_requested_at) if e.follow_up_requested_at else None,
            "is_pinned": pin is not None, "pin_priority": pin.priority if pin else None,
            "organizer_name": organizer.full_name if organizer else "Unknown",
            "registration_count": reg_count, "all_registration_count": all_reg_count,
            "paid_registration_count": paid_count,
            "pending_payment_count": pending_payment_count,
            "event_revenue": round(float(event_revenue or 0), 2),
            "checkin_count": checkin_count,
            "created_at": str(e.created_at)
        })
    return result


@app.delete("/admin/events/{event_id}")
def admin_delete_event(event_id: str, db: Session = Depends(database.get_db), admin: models.User = Depends(require_admin)):
    event = db.query(models.Event).filter(models.Event.id == event_id).first()
    if not event:
        raise HTTPException(status_code=404, detail="Event not found")
    # Delete related registrations first
    db.query(models.ContentPin).filter(models.ContentPin.target_type == "event", models.ContentPin.target_id == event_id).delete()
    db.query(models.ScorecardEvidence).filter(models.ScorecardEvidence.event_id == event_id).delete()
    db.query(models.ScorecardReportingExport).filter(models.ScorecardReportingExport.event_id == event_id).delete()
    participant_ids = [row[0] for row in db.query(models.ScorecardParticipant.id).filter(models.ScorecardParticipant.event_id == event_id).all()]
    if participant_ids:
        survey_ids = [row[0] for row in db.query(models.ScorecardSurvey.id).filter(models.ScorecardSurvey.participant_id.in_(participant_ids)).all()]
        if survey_ids:
            db.query(models.ScorecardResponse).filter(models.ScorecardResponse.survey_id.in_(survey_ids)).delete(synchronize_session=False)
        db.query(models.ScorecardSurvey).filter(models.ScorecardSurvey.participant_id.in_(participant_ids)).delete(synchronize_session=False)
        db.query(models.ScorecardParticipant).filter(models.ScorecardParticipant.id.in_(participant_ids)).delete(synchronize_session=False)
    db.query(models.Registration).filter(models.Registration.event_id == event_id).delete()
    db.query(models.EventFormField).filter(models.EventFormField.event_id == event_id).delete()
    db.delete(event)
    db.commit()
    return {"message": "Event deleted"}


@app.put("/admin/events/{event_id}/ticketing", response_model=schemas.EventResponse)
def admin_update_event_ticketing(
    event_id: str,
    ticketing: schemas.EventTicketingUpdate,
    db: Session = Depends(database.get_db),
    admin: models.User = Depends(require_admin)
):
    event = db.query(models.Event).filter(models.Event.id == event_id).first()
    if not event:
        raise HTTPException(status_code=404, detail="Event not found")

    currency = (ticketing.currency or event.currency or "KES").strip().upper()
    tiers = sanitize_ticket_tiers(ticketing.ticket_tiers, currency)
    ticket_price = max(float(ticketing.ticket_price or 0), 0)
    if tiers:
        ticket_price = max([ticket_price] + [float(tier.get("price") or 0) for tier in tiers])

    event.currency = currency
    event.ticket_price = ticket_price
    event.ticket_tiers = tiers
    event.attendee_type_question = ticketing.attendee_type_question if tiers else None

    db.commit()
    db.refresh(event)
    event.registrant_count = db.query(models.Registration).filter(
        models.Registration.event_id == event.id,
        models.Registration.status.in_(["registered", "checked-in"])
    ).count()
    pin = get_active_pin_map(db, "event").get(event.id)
    event.is_pinned = pin is not None
    event.pin_priority = pin.priority if pin else None
    return event


@app.get("/admin/community")
def admin_list_community(db: Session = Depends(database.get_db), admin: models.User = Depends(require_admin)):
    posts = db.query(models.CommunityMessage).order_by(models.CommunityMessage.created_at.desc()).limit(50).all()
    pin_map = get_active_pin_map(db, "community")
    result = []
    for p in posts:
        author = db.query(models.User).filter(models.User.id == p.author_id).first()
        reaction_count = db.query(models.ChatReaction).filter(models.ChatReaction.message_id == p.id).count()
        pin = pin_map.get(p.id)
        result.append({
            "id": p.id, "content": p.content, "is_poll": p.is_poll,
            "flag_count": p.flag_count, "is_hidden": p.is_hidden,
            "is_pinned": pin is not None, "pin_priority": pin.priority if pin else None,
            "hashtags": p.hashtags or [], "reaction_count": reaction_count,
            "author_name": author.full_name if author else "Unknown",
            "author_id": p.author_id,
            "created_at": str(p.created_at)
        })
    return result


@app.post("/admin/community/{post_id}/hide")
def admin_hide_post(post_id: str, db: Session = Depends(database.get_db), admin: models.User = Depends(require_admin)):
    post = db.query(models.CommunityMessage).filter(models.CommunityMessage.id == post_id).first()
    if not post:
        raise HTTPException(status_code=404, detail="Post not found")
    post.is_hidden = not post.is_hidden
    db.commit()
    return {"message": f"Post {'hidden' if post.is_hidden else 'restored'}", "is_hidden": post.is_hidden}


@app.delete("/admin/community/{post_id}")
def admin_delete_post(post_id: str, db: Session = Depends(database.get_db), admin: models.User = Depends(require_admin)):
    post = db.query(models.CommunityMessage).filter(models.CommunityMessage.id == post_id).first()
    if not post:
        raise HTTPException(status_code=404, detail="Post not found")
    db.query(models.ContentPin).filter(models.ContentPin.target_type == "community", models.ContentPin.target_id == post_id).delete()
    db.query(models.ChatReaction).filter(models.ChatReaction.message_id == post_id).delete()
    db.query(models.CommunityPollVote).filter(models.CommunityPollVote.message_id == post_id).delete()
    db.delete(post)
    db.commit()
    return {"message": "Post deleted"}


class AnnouncementCreate(BaseModel):
    title: str
    message: str
    target_audience: Optional[str] = "all"

@app.get("/admin/announcements")
def admin_list_announcements(db: Session = Depends(database.get_db), admin: models.User = Depends(require_admin)):
    announcements = db.query(models.Announcement).order_by(models.Announcement.created_at.desc()).all()
    return [{"id": a.id, "title": a.title, "message": a.message, "target_audience": a.target_audience,
             "created_at": str(a.created_at)} for a in announcements]

@app.post("/admin/announcements")
def admin_create_announcement(data: AnnouncementCreate, db: Session = Depends(database.get_db), admin: models.User = Depends(require_admin)):
    ann = models.Announcement(id=str(uuid.uuid4()), title=data.title, message=data.message, target_audience=data.target_audience)
    db.add(ann)
    db.commit()
    db.refresh(ann)
    return {"id": ann.id, "title": ann.title, "message": ann.message, "target_audience": ann.target_audience, "created_at": str(ann.created_at)}

@app.delete("/admin/announcements/{announcement_id}")
def admin_delete_announcement(announcement_id: str, db: Session = Depends(database.get_db), admin: models.User = Depends(require_admin)):
    ann = db.query(models.Announcement).filter(models.Announcement.id == announcement_id).first()
    if not ann:
        raise HTTPException(status_code=404, detail="Announcement not found")
    db.delete(ann)
    db.commit()
    return {"message": "Announcement deleted"}

@app.get("/admin/dogs")
def admin_list_dogs(db: Session = Depends(database.get_db), admin: models.User = Depends(require_admin)):
    from sqlalchemy import func
    dogs = db.query(models.Dog).all()
    breed_counts = db.query(models.Dog.breed, func.count(models.Dog.id)).group_by(models.Dog.breed).all()
    result = []
    for d in dogs:
        owner = db.query(models.User).filter(models.User.id == d.owner_id).first()
        health_count = db.query(models.HealthRecord).filter(models.HealthRecord.dog_id == d.id).count()
        result.append({
            "id": d.id, "name": d.name, "breed": d.breed, "color": d.color,
            "age": d.age, "weight": d.weight, "pet_type": d.pet_type,
            "owner_name": owner.full_name if owner else "Unknown",
            "health_records": health_count,
            "has_nose_print": d.nose_print_descriptor is not None
        })
    return {
        "dogs": result,
        "breed_distribution": {breed: count for breed, count in breed_counts},
        "total": len(dogs)
    }

# =====================================================
# Case Reporting API — Social Dog Case Reporting
# =====================================================

@app.post("/cases", response_model=schemas.CaseReportResponse)
def create_case_report(
    report: schemas.CaseReportCreate,
    db: Session = Depends(database.get_db),
    current_user: models.User = Depends(get_current_user)
):
    new_report = models.CaseReport(
        id=str(uuid.uuid4()),
        author_id=current_user.id,
        case_type=report.case_type,
        title=report.title,
        description=report.description,
        image_url=report.image_url,
        breed=report.breed,
        color=report.color,
        location=report.location,
        latitude=report.latitude,
        longitude=report.longitude,
        images=report.images,
    )
    db.add(new_report)
    db.commit()
    db.refresh(new_report)
    
    # Award points for useful community safety reporting.
    award_karma(
        db,
        current_user.id,
        KARMA_CASE_REPORT_REWARD,
        "case_report",
        f"Reported case: {new_report.title}",
        commit=False,
    )
    add_notification(
        db,
        current_user.id,
        "Points earned",
        f"You earned {KARMA_CASE_REPORT_REWARD} points for reporting '{new_report.title}'.",
        "reward",
        commit=True,
    )

    return {
        **{c.name: getattr(new_report, c.name) for c in new_report.__table__.columns},
        "author": {"id": current_user.id, "full_name": current_user.full_name, "profile_image": current_user.profile_image},
        "like_count": 0,
        "comment_count": 0,
        "is_liked": False,
    }


@app.get("/cases", response_model=List[schemas.CaseReportResponse])
def list_case_reports(
    skip: int = 0,
    limit: int = 50,
    db: Session = Depends(database.get_db),
    current_user: models.User = Depends(get_current_user)
):
    from sqlalchemy import func

    from sqlalchemy import or_
    
    # Filter: Show approved reports OR reports owned by the current user
    query = db.query(models.CaseReport).filter(
        or_(
            models.CaseReport.is_approved == True,
            models.CaseReport.author_id == current_user.id
        )
    )
    
    reports = query.order_by(models.CaseReport.created_at.desc()).all()
    pin_map = get_active_pin_map(db, "case")
    apply_pin_metadata(reports, pin_map)
    reports = sort_items_with_pins(
        reports,
        pin_map,
        secondary_key=lambda r: r.created_at or datetime.min,
        reverse_secondary=True,
    )[skip:skip + limit]

    results = []
    for r in reports:
        like_count = db.query(models.CaseLike).filter(models.CaseLike.report_id == r.id).count()
        comment_count = db.query(models.CaseComment).filter(models.CaseComment.report_id == r.id).count()
        is_liked = db.query(models.CaseLike).filter(
            models.CaseLike.report_id == r.id,
            models.CaseLike.user_id == current_user.id
        ).first() is not None

        author = db.query(models.User).filter(models.User.id == r.author_id).first()

        results.append({
            **{c.name: getattr(r, c.name) for c in r.__table__.columns},
            "author": {"id": author.id, "full_name": author.full_name, "profile_image": author.profile_image} if author else None,
            "like_count": like_count,
            "comment_count": comment_count,
            "is_liked": is_liked,
            "is_pinned": getattr(r, "is_pinned", False),
            "pin_priority": getattr(r, "pin_priority", None),
        })

    return results


@app.get("/cases/{report_id}", response_model=schemas.CaseReportResponse)
def get_case_report(
    report_id: str,
    db: Session = Depends(database.get_db),
    current_user: models.User = Depends(get_current_user)
):
    r = db.query(models.CaseReport).filter(models.CaseReport.id == report_id).first()
    if not r:
        raise HTTPException(status_code=404, detail="Case report not found")
        
    # Security: If not approved, only author or admin can view
    if not r.is_approved and r.author_id != current_user.id and current_user.role != models.UserRole.ADMIN:
        raise HTTPException(status_code=403, detail="Report pending moderation")

    like_count = db.query(models.CaseLike).filter(models.CaseLike.report_id == r.id).count()
    comment_count = db.query(models.CaseComment).filter(models.CaseComment.report_id == r.id).count()
    is_liked = db.query(models.CaseLike).filter(
        models.CaseLike.report_id == r.id,
        models.CaseLike.user_id == current_user.id
    ).first() is not None
    author = db.query(models.User).filter(models.User.id == r.author_id).first()
    pin = get_active_pin_map(db, "case").get(r.id)

    return {
        **{c.name: getattr(r, c.name) for c in r.__table__.columns},
        "author": {"id": author.id, "full_name": author.full_name, "profile_image": author.profile_image} if author else None,
        "like_count": like_count,
        "comment_count": comment_count,
        "is_liked": is_liked,
        "is_pinned": pin is not None,
        "pin_priority": pin.priority if pin else None,
    }


@app.post("/cases/{report_id}/comments", response_model=schemas.CaseCommentResponse)
def add_case_comment(
    report_id: str,
    comment: schemas.CaseCommentCreate,
    db: Session = Depends(database.get_db),
    current_user: models.User = Depends(get_current_user)
):
    report = db.query(models.CaseReport).filter(models.CaseReport.id == report_id).first()
    if not report:
        raise HTTPException(status_code=404, detail="Case report not found")

    new_comment = models.CaseComment(
        id=str(uuid.uuid4()),
        report_id=report_id,
        author_id=current_user.id,
        content=comment.content,
        tagged_users=comment.tagged_users,
    )
    db.add(new_comment)
    db.commit()
    db.refresh(new_comment)
    
    award_karma(
        db,
        current_user.id,
        KARMA_CASE_COMMENT_REWARD,
        "comment",
        f"Commented on case: {report.title}",
        commit=False,
    )
    add_notification(
        db,
        current_user.id,
        "Points earned",
        f"You earned {KARMA_CASE_COMMENT_REWARD} points for commenting on '{report.title}'.",
        "reward",
        commit=True,
    )

    return {
        **{c.name: getattr(new_comment, c.name) for c in new_comment.__table__.columns},
        "author": {"id": current_user.id, "full_name": current_user.full_name, "profile_image": current_user.profile_image},
    }


@app.get("/cases/{report_id}/comments", response_model=List[schemas.CaseCommentResponse])
def list_case_comments(
    report_id: str,
    db: Session = Depends(database.get_db),
):
    comments = db.query(models.CaseComment).filter(
        models.CaseComment.report_id == report_id
    ).order_by(models.CaseComment.created_at.asc()).all()

    results = []
    for c in comments:
        author = db.query(models.User).filter(models.User.id == c.author_id).first()
        results.append({
            **{col.name: getattr(c, col.name) for col in c.__table__.columns},
            "author": {"id": author.id, "full_name": author.full_name, "profile_image": author.profile_image} if author else None,
        })
    return results


@app.post("/cases/{report_id}/like")
def toggle_case_like(
    report_id: str,
    db: Session = Depends(database.get_db),
    current_user: models.User = Depends(get_current_user)
):
    report = db.query(models.CaseReport).filter(models.CaseReport.id == report_id).first()
    if not report:
        raise HTTPException(status_code=404, detail="Case report not found")

    existing = db.query(models.CaseLike).filter(
        models.CaseLike.report_id == report_id,
        models.CaseLike.user_id == current_user.id
    ).first()

    if existing:
        db.delete(existing)
        db.commit()
        return {"liked": False}
    else:
        new_like = models.CaseLike(
            id=str(uuid.uuid4()),
            report_id=report_id,
            user_id=current_user.id,
        )
        db.add(new_like)
        db.commit()
        return {"liked": True}


@app.get("/users/search")
def search_users(
    q: str,
    db: Session = Depends(database.get_db),
    current_user: models.User = Depends(get_current_user)
):
    """Search users by name for @mention tagging"""
    users = db.query(models.User).filter(
        models.User.full_name.ilike(f"%{q}%")
    ).limit(10).all()
    return [
        {"id": u.id, "full_name": u.full_name, "profile_image": u.profile_image}
        for u in users
    ]

@app.post("/payments/initiate")
async def initiate_payment(
    order_id: str,
    amount: Optional[float] = None,
    email: str = "",
    phone: str = "0700000000",
    db: Session = Depends(database.get_db),
    current_user: models.User = Depends(get_current_user),
):
    order = db.query(models.Order).filter(models.Order.id == order_id).first()
    if not order:
        raise HTTPException(status_code=404, detail="Order not found")
    if order.buyer_id != current_user.id:
        raise HTTPException(status_code=403, detail="Not authorized")

    ipn_url = os.getenv("PESAPAL_IPN_URL")
    callback_url = os.getenv("PESAPAL_CALLBACK_URL")
    if not ipn_url or not callback_url:
        raise HTTPException(status_code=500, detail="Pesapal checkout is not configured. Please set PESAPAL_IPN_URL and PESAPAL_CALLBACK_URL.")

    # 1. Register IPN
    ipn_res = pesapal.register_ipn(ipn_url)
    ipn_id = ipn_res.get("ipn_id")
    
    if not ipn_id:
        detail = ipn_res.get("error") or ipn_res.get("message") or ipn_res
        raise HTTPException(status_code=502, detail=f"Failed to register IPN with Pesapal: {detail}")
        
    # 2. Submit Order
    order_res = pesapal.submit_order(
        order_id=order_id,
        amount=order.amount,
        description=f"Lovedogs 360 - Order {order_id}",
        email=email,
        phone=phone,
        callback_url=callback_url,
        ipn_id=ipn_id,
        currency=(db.query(models.Service.currency).filter(models.Service.id == order.service_id).scalar() or "KES"),
    )
    if not order_res.get("redirect_url"):
        detail = order_res.get("error") or order_res.get("message") or order_res
        raise HTTPException(status_code=502, detail=f"Failed to start Pesapal checkout: {detail}")
    
    return order_res

@app.get("/pesapal/callback")
async def pesapal_callback(OrderTrackingId: str, OrderMerchantReference: str, db: Session = Depends(database.get_db)):
    status_res = pesapal.get_transaction_status(OrderTrackingId)
    order = db.query(models.Order).filter(models.Order.id == OrderMerchantReference).first()
    if order and is_pesapal_payment_successful(status_res):
        if mark_order_paid(db, order):
            db.commit()
        return {"status": "processed", "type": "order", "order_status": order.status, "data": status_res}

    registration = db.query(models.Registration).filter(
        (models.Registration.id == OrderMerchantReference) |
        (models.Registration.pesapal_merchant_reference == OrderMerchantReference)
    ).first()
    if registration and is_pesapal_payment_successful(status_res):
        if mark_event_registration_paid(db, registration, OrderTrackingId):
            db.commit()
    return {
        "status": "processed",
        "type": "event_registration" if registration else None,
        "registration_status": registration.status if registration else None,
        "order_status": order.status if order else None,
        "data": status_res,
    }

@app.get("/payments/status/{order_id}")
async def payment_status(
    order_id: str,
    tracking_id: Optional[str] = None,
    db: Session = Depends(database.get_db),
    current_user: models.User = Depends(get_current_user),
):
    order = db.query(models.Order).filter(models.Order.id == order_id).first()
    if not order:
        raise HTTPException(status_code=404, detail="Order not found")
    if order.buyer_id != current_user.id and current_user.role != models.UserRole.ADMIN:
        raise HTTPException(status_code=403, detail="Not authorized")

    status_res = None
    payment_success = is_order_paid(order)
    if tracking_id and not payment_success:
        status_res = pesapal.get_transaction_status(tracking_id)
        pesapal_success = is_pesapal_payment_successful(status_res)
        if pesapal_success and mark_order_paid(db, order):
            db.commit()
            payment_success = True
        else:
            payment_success = is_order_paid(order)

    return {
        "order_id": order.id,
        "order_status": order.status,
        "payment_success": payment_success,
        "payment_status": status_res,
        "buyer_reward_points": calculate_karma_reward(order.amount) if payment_success else 0,
        "seller_reward_points": calculate_karma_reward(order.payout) if payment_success else 0,
        "discount_amount": getattr(order, "discount_amount", 0) or 0,
        "karma_points_redeemed": getattr(order, "karma_points_redeemed", 0) or 0,
    }

@app.get("/pesapal/ipn")
async def pesapal_ipn(OrderTrackingId: str, OrderMerchantReference: str, OrderNotificationType: str, db: Session = Depends(database.get_db)):
    logger.info(f"IPN Received: {OrderTrackingId} for Order {OrderMerchantReference}")
    status_res = pesapal.get_transaction_status(OrderTrackingId)
    order = db.query(models.Order).filter(models.Order.id == OrderMerchantReference).first()
    if order and is_pesapal_payment_successful(status_res):
        if mark_order_paid(db, order):
            db.commit()
        return {"status": "acknowledged", "type": "order"}
    registration = db.query(models.Registration).filter(
        (models.Registration.id == OrderMerchantReference) |
        (models.Registration.pesapal_merchant_reference == OrderMerchantReference)
    ).first()
    if registration and is_pesapal_payment_successful(status_res):
        if mark_event_registration_paid(db, registration, OrderTrackingId):
            db.commit()
        return {"status": "acknowledged", "type": "event_registration"}
    return {"status": "acknowledged"}
# =====================================================
# Community Hub & Social Endpoints
# =====================================================

import re
from collections import Counter
from sqlalchemy import String

@app.get("/chat/global", response_model=List[schemas.CommunityMessageResponse])
def get_global_chat(tag: Optional[str] = None, db: Session = Depends(database.get_db), current_user: models.User = Depends(get_current_user)):
    query = db.query(models.CommunityMessage).filter(
        models.CommunityMessage.is_global == True,
        models.CommunityMessage.is_hidden == False
    )
    if tag:
        query = query.filter(models.CommunityMessage.hashtags.cast(String).ilike(f'%"{tag}"%'))
        
    pin_map = get_active_pin_map(db, "community")
    messages = query.order_by(models.CommunityMessage.created_at.desc()).limit(100).all()
    apply_pin_metadata(messages, pin_map)
    messages = sort_items_with_pins(
        messages,
        pin_map,
        secondary_key=lambda m: m.created_at or datetime.min,
        reverse_secondary=True,
    )[:50]
    for msg in messages:
        if msg.is_poll and msg.poll_options:
            votes = db.query(models.CommunityPollVote).filter(models.CommunityPollVote.message_id == msg.id).all()
            results = {str(opt["id"]): 0 for opt in msg.poll_options}
            current_user_vote = None
            for v in votes:
                results[str(v.option_id)] = results.get(str(v.option_id), 0) + 1
                if v.user_id == current_user.id:
                    current_user_vote = v.option_id
            msg.poll_results = results
            msg.has_voted = current_user_vote
    return messages

@app.get("/chat/nearby", response_model=List[schemas.CommunityMessageResponse])
def get_nearby_chat(radius: float = 20.0, db: Session = Depends(database.get_db), current_user: models.User = Depends(get_current_user)):
    # User's registered location
    u_lat, u_lon = current_user.latitude, current_user.longitude
    if u_lat is None or u_lon is None:
        return []

    # Get all messages with lat/lng and filter manually (or use PostGIS if available, but Haversine is fallback)
    pin_map = get_active_pin_map(db, "community")
    messages = db.query(models.CommunityMessage).filter(
        models.CommunityMessage.latitude != None,
        models.CommunityMessage.is_hidden == False,
    ).all()
    apply_pin_metadata(messages, pin_map)
    nearby = []
    for msg in messages:
        dist = calculate_distance(u_lat, u_lon, msg.latitude, msg.longitude)
        if dist <= radius:
            nearby.append(msg)
    
    return sort_items_with_pins(
        nearby,
        pin_map,
        secondary_key=lambda x: x.created_at or datetime.min,
        reverse_secondary=True,
    )[:50]

@app.post("/chat/message", response_model=schemas.CommunityMessageResponse)
def post_community_message(msg: schemas.CommunityMessageCreate, db: Session = Depends(database.get_db), current_user: models.User = Depends(get_current_user)):
    hashtags = list(set([word.lower() for word in re.findall(r'#(\\w+)', msg.content)]))
    if msg.hashtags: # Also allow explicitly passed tags
        hashtags = list(set(hashtags + [h.lower().strip('#') for h in msg.hashtags]))
        
    new_msg = models.CommunityMessage(
        id=str(uuid.uuid4()),
        author_id=current_user.id,
        content=msg.content,
        latitude=current_user.latitude,
        longitude=current_user.longitude,
        is_global=msg.is_global,
        reshare_id=msg.reshare_id,
        hashtags=hashtags,
        is_poll=msg.is_poll,
        poll_options=msg.poll_options
    )
    db.add(new_msg)
    
    # Award Karma for participating
    award_karma(db, current_user.id, 1, "chat", "Sent a community message")
    
    db.commit()
    db.refresh(new_msg)
    
    if new_msg.is_poll:
        new_msg.poll_results = {str(opt["id"]): 0 for opt in (new_msg.poll_options or [])}
        new_msg.has_voted = None
        
    return new_msg

@app.get("/chat/trending-tags", response_model=List[schemas.TrendingTagResponse])
def get_trending_tags(db: Session = Depends(database.get_db)):
    seven_days_ago = datetime.utcnow() - timedelta(days=7)
    msgs = db.query(models.CommunityMessage).filter(
        models.CommunityMessage.created_at >= seven_days_ago,
        models.CommunityMessage.is_hidden == False
    ).all()
    
    all_tags = []
    for m in msgs:
        if m.hashtags:
            all_tags.extend(m.hashtags)
            
    counter = Counter(all_tags)
    return [{"tag": tag, "count": count} for tag, count in counter.most_common(10)]

@app.post("/chat/messages/{message_id}/vote")
def vote_poll(message_id: str, vote: schemas.PollVoteCreate, db: Session = Depends(database.get_db), current_user: models.User = Depends(get_current_user)):
    msg = db.query(models.CommunityMessage).filter(models.CommunityMessage.id == message_id).first()
    if not msg or not msg.is_poll:
        raise HTTPException(status_code=404, detail="Poll not found")
        
    existing = db.query(models.CommunityPollVote).filter(
        models.CommunityPollVote.message_id == message_id,
        models.CommunityPollVote.user_id == current_user.id
    ).first()
    
    if existing:
        existing.option_id = vote.option_id
    else:
        new_vote = models.CommunityPollVote(
            id=str(uuid.uuid4()),
            message_id=message_id,
            user_id=current_user.id,
            option_id=vote.option_id
        )
        db.add(new_vote)
    db.commit()
    return {"status": "success"}

@app.post("/chat/messages/{message_id}/flag")
def flag_message(message_id: str, db: Session = Depends(database.get_db), current_user: models.User = Depends(get_current_user)):
    msg = db.query(models.CommunityMessage).filter(models.CommunityMessage.id == message_id).first()
    if not msg:
        raise HTTPException(status_code=404, detail="Message not found")
        
    msg.flag_count += 1
    if msg.flag_count >= 3:
        msg.is_hidden = True
        
    db.commit()
    return {"status": "success", "is_hidden": msg.is_hidden}

@app.post("/chat/messages/{message_id}/react")
def react_to_message(message_id: str, reaction: schemas.ChatReactionCreate, db: Session = Depends(database.get_db), current_user: models.User = Depends(get_current_user)):
    # Check if existing reaction
    existing = db.query(models.ChatReaction).filter(
        models.ChatReaction.message_id == message_id,
        models.ChatReaction.user_id == current_user.id
    ).first()
    
    if existing:
        existing.reaction_type = reaction.reaction_type
    else:
        new_react = models.ChatReaction(
            id=str(uuid.uuid4()),
            message_id=message_id,
            user_id=current_user.id,
            reaction_type=reaction.reaction_type
        )
        db.add(new_react)
        
    db.commit()
    return {"status": "success"}

@app.get("/chat/dms", response_model=List[schemas.DirectMessageResponse])
def get_my_dms(db: Session = Depends(database.get_db), current_user: models.User = Depends(get_current_user)):
    return db.query(models.DirectMessage).filter(
        (models.DirectMessage.sender_id == current_user.id) | 
        (models.DirectMessage.receiver_id == current_user.id)
    ).order_by(models.DirectMessage.created_at.desc()).all()

@app.post("/chat/dm", response_model=schemas.DirectMessageResponse)
def send_direct_message(dm: schemas.DirectMessageCreate, db: Session = Depends(database.get_db), current_user: models.User = Depends(get_current_user)):
    receiver = db.query(models.User).filter(models.User.id == dm.receiver_id).first()
    if not receiver:
        raise HTTPException(status_code=404, detail="Receiver not found")
    if receiver.id == current_user.id:
        raise HTTPException(status_code=400, detail="You cannot message yourself")

    new_dm = models.DirectMessage(
        id=str(uuid.uuid4()),
        sender_id=current_user.id,
        receiver_id=dm.receiver_id,
        content=dm.content
    )
    db.add(new_dm)
    db.commit()
    db.refresh(new_dm)
    return new_dm

@app.post("/chat/dms/{message_id}/read")
def mark_direct_message_read(message_id: str, db: Session = Depends(database.get_db), current_user: models.User = Depends(get_current_user)):
    dm = db.query(models.DirectMessage).filter(
        models.DirectMessage.id == message_id,
        models.DirectMessage.receiver_id == current_user.id
    ).first()
    if not dm:
        raise HTTPException(status_code=404, detail="Direct message not found")

    if not dm.read_at:
        dm.read_at = datetime.utcnow()
        db.commit()

    return {"message": "Success"}

# =====================================================
# Karma & Status Endpoints
# =====================================================

@app.post("/users/status/heartbeat")
def user_heartbeat(db: Session = Depends(database.get_db), current_user: models.User = Depends(get_current_user)):
    current_user.is_online = True
    current_user.last_seen = datetime.utcnow()
    db.commit()
    return {"status": "online"}

@app.get("/users/online", response_model=List[schemas.UserResponse])
def get_online_users(db: Session = Depends(database.get_db)):
    # Simple logic: seen in the last 5 minutes
    five_mins_ago = datetime.utcnow() - timedelta(minutes=5)
    return db.query(models.User).filter(models.User.last_seen >= five_mins_ago).all()

@app.post("/karma/redeem")
def redeem_karma(req: schemas.KarmaRedeemRequest, db: Session = Depends(database.get_db), current_user: models.User = Depends(get_current_user)):
    if req.amount_to_redeem < KARMA_REDEMPTION_TARGET:
        raise HTTPException(status_code=400, detail=f"Redeem at least {KARMA_REDEMPTION_TARGET} points")
    if current_user.available_karma < KARMA_REDEMPTION_TARGET:
        raise HTTPException(status_code=400, detail=f"You need at least {KARMA_REDEMPTION_TARGET} points before redeeming")
    if current_user.available_karma < req.amount_to_redeem:
        raise HTTPException(status_code=400, detail="Insufficient karma points")
    
    current_user.available_karma -= req.amount_to_redeem
    
    # Track redemption
    transaction = models.KarmaTransaction(
        id=str(uuid.uuid4()),
        user_id=current_user.id,
        amount=-req.amount_to_redeem,
        category="redemption",
        description=f"Redeemed {req.amount_to_redeem} points for marketplace credit"
    )
    db.add(transaction)
    db.commit()
    
    return {"status": "success", "new_balance": current_user.available_karma}

# =====================================================
# Support Tickets & Announcements (Admin)
# =====================================================



@app.get("/admin/support-tickets")
def get_all_support_tickets(
    db: Session = Depends(database.get_db),
    admin: models.User = Depends(require_admin)
):
    tickets = db.query(models.SupportTicket).order_by(models.SupportTicket.created_at.desc()).all()
    results = []
    for t in tickets:
        u = db.query(models.User).filter(models.User.id == t.user_id).first()
        results.append({
            "id": t.id,
            "subject": t.subject,
            "message": t.message,
            "status": support_status_label(t.status),
            "status_key": support_status_key(t.status),
            "admin_reply": t.admin_reply,
            "images": t.images or [],
            "user_id": t.user_id,
            "user_name": u.full_name if u else "Unknown",
            "user_email": u.email if u else None,
            "created_at": str(t.created_at),
            "updated_at": str(t.updated_at) if t.updated_at else None,
        })
    return results

class AdminSupportReplyMsg(BaseModel):
    message: str

@app.post("/admin/support-tickets/{ticket_id}/reply")
def reply_support_ticket(
    ticket_id: str,
    req: AdminSupportReplyMsg,
    db: Session = Depends(database.get_db),
    admin: models.User = Depends(require_admin)
):
    ticket = db.query(models.SupportTicket).filter(models.SupportTicket.id == ticket_id).first()
    if not ticket:
        raise HTTPException(status_code=404, detail="Ticket not found")
    
    ticket.admin_reply = req.message
    ticket.status = "in-progress"
    ticket.updated_at = datetime.utcnow()
    db.commit()
    return {"message": "Reply sent"}

@app.post("/admin/support-tickets/{ticket_id}/resolve")
def resolve_support_ticket(
    ticket_id: str,
    db: Session = Depends(database.get_db),
    admin: models.User = Depends(require_admin)
):
    ticket = db.query(models.SupportTicket).filter(models.SupportTicket.id == ticket_id).first()
    if not ticket:
        raise HTTPException(status_code=404, detail="Ticket not found")
    
    ticket.status = "resolved"
    ticket.updated_at = datetime.utcnow()
    db.commit()
    return {"message": "Ticket resolved"}

@app.post("/admin/announcements", response_model=schemas.AnnouncementResponse)
def create_announcement(
    req: schemas.AnnouncementCreate,
    db: Session = Depends(database.get_db),
    admin: models.User = Depends(require_admin)
):
    announcement = models.Announcement(
        id=str(uuid.uuid4()),
        title=req.title,
        message=req.message,
        target_audience=req.target_audience
    )
    db.add(announcement)
    db.commit()
    db.refresh(announcement)
    return announcement

@app.get("/announcements", response_model=List[schemas.AnnouncementResponse])
def get_announcements(
    db: Session = Depends(database.get_db),
    current_user: models.User = Depends(get_current_user)
):
    # Depending on user role, filter target_audience
    if current_user.role == "buyer":
        return db.query(models.Announcement).filter(models.Announcement.target_audience.in_(["all", "buyers"])).order_by(models.Announcement.created_at.desc()).all()
    elif current_user.role == "provider":
        return db.query(models.Announcement).filter(models.Announcement.target_audience.in_(["all", "providers"])).order_by(models.Announcement.created_at.desc()).all()
    else:
        return db.query(models.Announcement).order_by(models.Announcement.created_at.desc()).all()

# --- Personalized Notifications ---

def create_notification(db: Session, user_id: str, title: str, message: str, type: str = "info"):
    return add_notification(db, user_id, title, message, type, commit=True)

@app.get("/notifications", response_model=List[schemas.NotificationResponse])
def get_notifications(
    db: Session = Depends(database.get_db),
    current_user: models.User = Depends(get_current_user)
):
    return db.query(models.Notification).filter(
        models.Notification.user_id == current_user.id
    ).order_by(models.Notification.created_at.desc()).limit(50).all()

@app.post("/notifications/{notification_id}/read")
def mark_notification_read(
    notification_id: str,
    db: Session = Depends(database.get_db),
    current_user: models.User = Depends(get_current_user)
):
    notif = db.query(models.Notification).filter(
        models.Notification.id == notification_id,
        models.Notification.user_id == current_user.id
    ).first()
    if notif:
        notif.is_read = True
        db.commit()
    return {"message": "Success"}

# =====================================================
# QR Code Ticket Verification (Admin)
# =====================================================

@app.get("/admin/verify-ticket")
def verify_event_ticket(
    token: str,
    db: Session = Depends(database.get_db),
    admin: models.User = Depends(require_admin)
):
    registration = db.query(models.Registration).filter(models.Registration.ticket_token == token).first()
    if not registration:
        raise HTTPException(status_code=404, detail="Ticket not found or invalid token")
    
    user = db.query(models.User).filter(models.User.id == registration.user_id).first()
    event = db.query(models.Event).filter(models.Event.id == registration.event_id).first()
    
    if not user or not event:
        raise HTTPException(status_code=404, detail="Data inconsistency found for ticket")
        
    return {
        "valid": True,
        "checked_in": registration.check_in_time is not None,
        "check_in_time": registration.check_in_time,
        "registration_status": registration.status,
        "user_name": user.full_name,
        "user_email": user.email,
        "event_title": event.title,
        "role": registration.role
    }

@app.post("/admin/check-in-ticket")
def check_in_ticket(
    token: str,
    db: Session = Depends(database.get_db),
    admin: models.User = Depends(require_admin)
):
    registration = db.query(models.Registration).filter(models.Registration.ticket_token == token).first()
    if not registration:
        raise HTTPException(status_code=404, detail="Ticket not found or invalid token")
    
    if registration.check_in_time is not None:
        raise HTTPException(status_code=400, detail="Ticket has already been used")
        
    registration.check_in_time = datetime.utcnow()
    registration.status = "checked-in"
    db.commit()
    
    return {"message": "Success", "checked_in": True, "time": registration.check_in_time}

# =====================================================
# Admin User Management
# =====================================================

@app.post("/admin/users", response_model=schemas.UserResponse)
def admin_create_user(
    req: schemas.UserCreate,
    db: Session = Depends(database.get_db),
    admin: models.User = Depends(require_admin)
):
    existing = db.query(models.User).filter(models.User.email == req.email).first()
    if existing:
        raise HTTPException(status_code=400, detail="Email already registered")
        
    user = models.User(
        id=str(uuid.uuid4()),
        email=req.email,
        full_name=req.full_name,
        role=req.role or "buyer",
        phone_number=req.phone_number,
        hashed_password=auth.get_password_hash(req.password)
    )
    db.add(user)
    db.commit()
    db.refresh(user)
    return user

@app.post("/admin/users/{user_id}/suspend")
def admin_suspend_user(
    user_id: str,
    db: Session = Depends(database.get_db),
    admin: models.User = Depends(require_admin)
):
    user = db.query(models.User).filter(models.User.id == user_id).first()
    if not user:
        raise HTTPException(status_code=404, detail="User not found")
        
    user.role = "suspended"
    db.commit()
    return {"message": f"User {user.email} suspended"}

# --- Support Tickets API ---

@app.post("/support", response_model=schemas.SupportTicketResponse)
def create_support_ticket(
    ticket: schemas.SupportTicketCreate,
    db: Session = Depends(database.get_db),
    current_user: models.User = Depends(get_current_user)
):
    new_ticket = models.SupportTicket(
        id=str(uuid.uuid4()),
        user_id=current_user.id,
        subject=ticket.subject,
        message=ticket.message,
        status="open",
        images=ticket.images
    )
    db.add(new_ticket)
    db.commit()
    db.refresh(new_ticket)
    return new_ticket

@app.get("/support", response_model=List[schemas.SupportTicketResponse])
def get_support_tickets(
    db: Session = Depends(database.get_db),
    current_user: models.User = Depends(get_current_user)
):
    if current_user.role in ["admin", "super_admin"]:
        tickets = db.query(models.SupportTicket).order_by(models.SupportTicket.created_at.desc()).all()
    else:
        tickets = db.query(models.SupportTicket).filter(models.SupportTicket.user_id == current_user.id).order_by(models.SupportTicket.created_at.desc()).all()
    return tickets

class SupportReply(BaseModel):
    admin_reply: str

@app.post("/support/{ticket_id}/reply")
def reply_to_support_ticket(
    ticket_id: str,
    reply: SupportReply,
    db: Session = Depends(database.get_db),
    admin: models.User = Depends(require_admin)
):
    ticket = db.query(models.SupportTicket).filter(models.SupportTicket.id == ticket_id).first()
    if not ticket:
        raise HTTPException(status_code=404, detail="Ticket not found")
    
    ticket.admin_reply = reply.admin_reply
    ticket.status = "resolved"
    ticket.updated_at = datetime.utcnow()
    db.commit()
    
    # Notify User
    create_notification(
        db, 
        ticket.user_id, 
        "Support Ticket Reply", 
        f"An admin has replied to your ticket: '{ticket.subject}'. Feedback: {reply.admin_reply}",
        "feedback"
    )

    db.refresh(ticket)
    return {"message": "Reply sent successfully", "status": ticket.status}



exchange_rates_cache = {
    "rates": {},
    "last_updated": 0
}

@app.get("/exchange-rates")
def get_exchange_rates():
    current_time = time.time()
    # Cache for 1 hour (3600 seconds)
    if current_time - exchange_rates_cache["last_updated"] > 3600 or not exchange_rates_cache["rates"]:
        try:
            req = urllib.request.Request("https://open.er-api.com/v6/latest/USD")
            with urllib.request.urlopen(req, timeout=5) as response:
                if response.status == 200:
                    data = json.loads(response.read().decode())
                    exchange_rates_cache["rates"] = data.get("rates", {})
                    exchange_rates_cache["last_updated"] = current_time
        except Exception as e:
            logger.warning(f"Error fetching exchange rates: {e}")
            # Silently fail and return stale rules if any exist
    return {"rates": exchange_rates_cache["rates"]}


# =====================================================
# User Safety — Report & Block (Store Compliance)
# =====================================================

class CaseFlagRequest(BaseModel):
    reason: str  # spam, harmful, misinformation

@app.post("/cases/{case_id}/flag")
async def flag_case_report(
    case_id: str,
    flag: CaseFlagRequest,
    db: Session = Depends(database.get_db),
    current_user: models.User = Depends(get_current_user)
):
    """
    Allows a user to report/flag a community case post for moderation review.
    Stores an audit log entry visible in the admin moderation panel.
    Does NOT delete the post — admin reviews first.
    """
    case = db.query(models.CaseReport).filter(models.CaseReport.id == case_id).first()
    if not case:
        raise HTTPException(status_code=404, detail="Post not found")

    # Prevent self-reporting
    if case.author_id == current_user.id:
        raise HTTPException(status_code=400, detail="You cannot report your own post")

    # Check for duplicate report from same user
    existing = db.query(models.AuditLog).filter(
        models.AuditLog.user_id == current_user.id,
        models.AuditLog.action == "flag_case",
        models.AuditLog.target_id == case_id
    ).first()
    if existing:
        return {"message": "You have already reported this post"}

    # Record the report in AuditLog (reuses existing model — no migration needed)
    log = models.AuditLog(
        id=str(uuid.uuid4()),
        user_id=current_user.id,
        action="flag_case",
        target_type="case_report",
        target_id=case_id,
        details=f"reason={flag.reason}"
    )
    db.add(log)
    db.commit()

    logger.info(f"Case {case_id} flagged by user {current_user.id} for reason: {flag.reason}")
    return {"message": "Report submitted successfully. Our moderation team will review this post."}


@app.post("/users/{user_id}/block")
async def block_user(
    user_id: str,
    db: Session = Depends(database.get_db),
    current_user: models.User = Depends(get_current_user)
):
    """
    Allows a user to block another user.
    Stores a block record in AuditLog for admin visibility.
    Blocked state is tracked client-side; blocking does not remove content.
    """
    if user_id == current_user.id:
        raise HTTPException(status_code=400, detail="You cannot block yourself")

    target_user = db.query(models.User).filter(models.User.id == user_id).first()
    if not target_user:
        raise HTTPException(status_code=404, detail="User not found")

    # Check if already blocked
    existing_block = db.query(models.AuditLog).filter(
        models.AuditLog.user_id == current_user.id,
        models.AuditLog.action == "block_user",
        models.AuditLog.target_id == user_id
    ).first()
    if existing_block:
        return {"message": "User is already blocked", "blocked": True}

    # Record block in AuditLog
    log = models.AuditLog(
        id=str(uuid.uuid4()),
        user_id=current_user.id,
        action="block_user",
        target_type="user",
        target_id=user_id,
        details=f"blocker={current_user.email} blocked={target_user.email}"
    )
    db.add(log)
    db.commit()

    logger.info(f"User {current_user.id} blocked user {user_id}")
    return {"message": f"User has been blocked successfully", "blocked": True}

